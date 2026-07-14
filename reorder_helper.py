# -*- coding: utf-8 -*-
"""订货辅助表：待发货明细表 × purchase order → 一行一品的订货决策清单。

用途（独立脚本，不进 gui.py）：
    每周收到「待发货明细表」后，对每个待订货品去 ERP purchase order 里查
    最近采购价 / 供应商 / 数量 / 当前库存，据此判断向哪家 vendor 订多少。
    本脚本把这步自动化，产出一张辅助表。订货状态 / MHD 仍人工填
    （到货量是订完才录、MHD 从 purchase 端导不出且会变化）。

数据关联（ID 优先，PZN 回退，逐行决定）：
    - 首选键 = ERP Product ID（数据库数字主键；External ID 串里嵌同一数字，归一互通）。
      仅当两侧都带产品 ID 列时可用：ERP↔ERP 直连；外部 PZN 输入配 --master 时走
      PZN → master → ID → PO 三段桥。ID 键绕开官方 PZN 空白/脏值与 IntRef 嵌旧 PZN 两坑。
    - 回退键 = PZN。purchase order 的 Internal Reference 尾部数字段即 PZN，
      两侧都去掉 PZN- 前缀、补零到 8 位再 join（6 位 EAN 类码对不上属正常）。
      待发货明细表等无 ID 输入不传 master 时走此路径，行为与旧版一致。
    - 平台裸价来自「待发货明细表」自带的 采购订单-* / 保税 分表（purchase
      order 导出本身无裸价列）。
    - 当前库存 = purchase order 的 Product/Quantity On Hand。

复用 build_excel：_short_vendor/_vendor_map（供应商简称）、_write_simple（统一版式）。

用法：
    python3 reorder_helper.py <待发货明细表.xlsx> <purchase order.xlsx> [out.xlsx]
"""

import sys, os, re
from datetime import date, datetime

import openpyxl

from build_excel import _short_vendor, _vendor_map, _write_simple, unique_path

# ---- purchase order 列名（兼容不同导出版本）----
PO_REF      = "Order Reference"
PO_VENDOR   = "Vendor"
PO_INTERNAL = "Order Lines/Product/Internal Reference"
PO_PRICE    = "Order Lines/Unit Price"
PO_QTY      = "Order Lines/Total Quantity"
PO_ONHAND   = "Product/Quantity On Hand"
PO_DATE_CANDS = ["Order Lines/Order Date", "Order Lines/Created on",
                 "Order Lines/Order Deadline"]
PO_PID_CANDS = ["Order Lines/Product/ID", "Order Lines/Product/External ID",
                "Product/ID", "Product/External ID"]  # 产品 ID 列（选填，叫法不一）

# 待发货明细表里带「平台裸价」的分表（按条形码取裸价）
PRICE_SHEETS = ["采购订单-健康", "采购订单-直营", "健康-保税"]

# ---- product.product 主数据列名（选填富化输入）----
PM_INTERNAL = "Internal Reference"
PM_BARCODE  = "Barcode"
PM_PZN      = "PZN"
PM_NAME     = "Name"
PM_ONHAND   = "Quantity On Hand"
# ERP 权威主键：PZN 会更新但 ID 不变（ERP↔ERP 切 ID 更稳）。导出叫法不一，宽松兼容。
PM_ID_CANDS = ["product/ID", "ID", "External ID"]

# 采购单里伪装成供应商的客户（实为我方销售平台，属噪音，整行剔除）。
# 比 build_excel 的 "Alibaba Health" 更宽：导出实见 Alibaba Health（港）与 Alibaba.com（新加坡）两个实体，均为客户。
CUSTOMER_PAT = "alibaba"

RECENT_N = 5  # 「近期采购记录」列展示最近几笔


def norm_pzn(x):
    """从条形码 / Internal Reference / 产品引用里提取 PZN，归一为补零 8 位的可比较键。
    按模式抓（非抓尾数），抓不到返回 None：
      1) `前缀_PZN(x件装)`：`Mucosolvan_02807988` / 销售分析 `[DOPH_12351236x2]`；
      2) 显式 `PZN-02807988`；
      3) 整格就是 7~8 位数字（纯 PZN 清单 / 条形码）。
    金额（`30623.25` 带小数、无下划线）与 12 位后端 id / 13 位 EAN（长度不符）自然抓不到→None。"""
    s = str(x).strip()
    m = re.search(r"_(\d{7,8})(?!\d)", s)           # 前缀_PZN，可带 x件装后缀
    if not m:
        m = re.search(r"PZN[-\s]*(\d{7,8})(?!\d)", s, flags=re.I)  # 显式 PZN 码
    if m:
        return m.group(1).zfill(8)
    s2 = re.sub(r"^PZN[-\s]*", "", s, flags=re.I).strip()
    if re.fullmatch(r"\d{7,8}", s2):                # 整格 7~8 位数字（条形码/PZN 清单）
        return s2.zfill(8)
    return None


def _norm_pid(x):
    """ERP 产品 ID 归一为纯数字串（product_product 表数据库主键）。兼容两种导出形态：
    纯数字 `200392` 与 External ID `__export__.product_product_200392_320ee90e`
    （后者是首次导出时按 `__export__.<模型>_<数字ID>_<哈希>` 自动生成，数字段=同一主键）。
    两形态归一后互通；非产品 ID（如 sale_order_line 的 External ID）返回 None。"""
    if x in (None, ""):
        return None
    s = str(x).strip()
    if re.fullmatch(r"\d+", s):
        return s
    m = re.search(r"product_product_(\d+)", s)
    return m.group(1) if m else None


# 需求侧产品 ID 列只认**产品作用域**表头：sale.order 导出里裸 `ID`/`External ID`
# 是订单自身的 id（sale_order_397395），认了就串键，必须排除。
PID_HDR_KEYS = ("Product/ID", "Product/External ID")


def _find_pid_col(rows, data_start):
    """在需求表里找产品 ID 列（选填，找不到返回 None，不阻断）：
    ① 表头含 Product/ID / Product/External ID（覆盖 Order Lines/ 前缀各变体）；
    ② 值形态含 `product_product_<数字>`（External ID 串无歧义可认；纯数字列有歧义
      ——数量/金额也长这样——不作形态证据，故数字形态 ID 只能靠表头认）。"""
    region = rows[:data_start] if data_start else rows[:3]
    for r in region:
        for j, v in enumerate(r):
            if v is not None and any(k in str(v) for k in PID_HDR_KEYS):
                return j
    hits = {}
    for r in rows[data_start:]:
        for j, v in enumerate(r):
            if v and re.search(r"product_product_\d+", str(v)):
                hits[j] = hits.get(j, 0) + 1
    return max(hits, key=hits.get) if hits else None


def _is_pzn(v):
    """值形态判 PZN：能被 norm_pzn 抽出 7~8 位 PZN。
    自然避开后端商品id(12 位)、采购单号(PON…长串)、EAN(13 位) 与金额（带小数）。"""
    return v is not None and norm_pzn(v) is not None


def _num(x):
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


def _pdate(x):
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(str(x)[:19] if " " in str(x) else str(x)[:10], fmt)
        except (TypeError, ValueError):
            continue
    return None


def _find_pzn_col(rows):
    """在一张 sheet 里按『值形态』找 PZN 主键列，不认死列头。
    返回 (列号, 数据起始行号) —— 该列 PZN 命中数最多；无命中返回 None。"""
    hits, first = {}, {}
    for i, r in enumerate(rows):
        for j, v in enumerate(r):
            if _is_pzn(v):
                hits[j] = hits.get(j, 0) + 1
                first.setdefault(j, i)
    if not hits:
        return None
    col = min((c for c in hits if hits[c] == max(hits.values())))  # 命中最多；并列取最左列
    return col, first[col]


def _find_label_col(rows, data_start, exact, loose=()):
    """在数据起始行之上的表头区找一列：先精确匹配 exact，再宽松 contains loose。
    找不到返回 None（该信息缺失，不阻断）。"""
    region = rows[:data_start] if data_start else rows[:3]
    for want in (exact, None):
        for r in region:
            for j, v in enumerate(r):
                if v is None:
                    continue
                s = str(v).strip()
                if want is not None and s == want:
                    return j
                if want is None and loose and any(k in s for k in loose):
                    return j
    return None


def _extract_intref(v):
    """从产品引用里抽 Internal Reference：销售分析 `[Abtei_14130309] 名称` 取中括号内；
    本身就是 `前缀_数字(x件装)` 形态则取整段；纯条形码/PZN 无 IntRef 返回 None。"""
    s = str(v).strip()
    m = re.search(r"\[([^\]]+?)\]", s)   # 中括号内(销售分析)
    if m:
        s = m.group(1).strip()
    return s if re.search(r"_\d{7,8}", s) else None


def load_demand(path):
    """商品清单 → [{pzn, name, need, intref}]，对输入宽容（不认死 sheet 名/列头）。
    优先用待发货明细表的『需求汇总-HK』；缺则回退——逐 sheet 按『值形态』找含 PZN 的列，
    第一张命中的即主键列。商品名称/总需求 按表头宽松匹配，取不到就留空（None），不报错，
    故『仅一列 PZN 的清单』也能跑。intref 从引用串里抽(有则供 Internal Reference 列/富化回退)。"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    order = (["需求汇总-HK"] if "需求汇总-HK" in wb.sheetnames else []) + \
            [s for s in wb.sheetnames if s != "需求汇总-HK"]
    for name in order:
        rows = list(wb[name].iter_rows(values_only=True))
        found = _find_pzn_col(rows)
        if found:
            c_bc, data_start = found
            break
    else:
        raise ValueError("未找到含 PZN 的列（商品清单里应有一列 PZN，如 PZN-02807988）")
    c_name = _find_label_col(rows, data_start, "商品名称", loose=("名称", "品名"))
    c_need = _find_label_col(rows, data_start, "总需求", loose=("需求",))
    c_pid = _find_pid_col(rows, data_start)   # ERP 导出才有；待发货表/纯清单为 None
    out = []
    for r in rows[data_start:]:
        bc = r[c_bc] if c_bc < len(r) else None
        pid_raw = r[c_pid] if (c_pid is not None and c_pid < len(r)) else None
        # 收 PZN 行；ERP 输入里无 PZN 但有产品 ID 的行（如 dm 系无 PZN 品）也收，走 ID 匹配
        if not _is_pzn(bc) and _norm_pid(pid_raw) is None:
            continue  # 跳过表头残留/空行/小计
        nm = r[c_name] if (c_name is not None and c_name < len(r)) else None
        out.append({"pzn": norm_pzn(bc),
                    "name": str(nm).strip() if nm else "",
                    "need": _num(r[c_need]) if (c_need is not None and c_need < len(r)) else None,
                    "intref": _extract_intref(bc),
                    "pid": _norm_pid(pid_raw),
                    "pid_raw": str(pid_raw).strip() if pid_raw not in (None, "") else ""})
    return out


def load_caps(path):
    """待发货明细表自带的采购订单/保税分表 → {pzn: 平台裸价}。"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    caps = {}
    for sh in PRICE_SHEETS:
        if sh not in wb.sheetnames:
            continue
        rows = list(wb[sh].iter_rows(values_only=True))
        if not rows:
            continue
        hdr = {str(h).strip(): i for i, h in enumerate(rows[0]) if h}
        cb, cp = hdr.get("条形码"), hdr.get("平台裸价")
        if cb is None or cp is None:
            continue
        for r in rows[1:]:
            if cb < len(r) and r[cb] and cp < len(r) and r[cp] not in (None, ""):
                v = _num(r[cp])
                k = norm_pzn(r[cb])
                if v is not None and k is not None:
                    caps.setdefault(k, v)  # 首见为准（分表内同码裸价一致）
    return caps


def load_master(path):
    """product.product 主数据(选填)富化 → {key: rec}。key **同时**索引 IntRef 嵌入 PZN 与
    官方 PZN 字段(桥接「名称/IntRef 留旧 PZN、PZN 字段是更新后新 PZN」的错位)。
    rec = {id(product/ID，可空), pzn(官方), name, barcode(EAN), intref, onhand}。"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    hdr = {str(h).strip(): i for i, h in enumerate(rows[0]) if h is not None}
    missing = [c for c in (PM_INTERNAL, PM_PZN, PM_NAME) if c not in hdr]
    if missing:
        raise ValueError(f"product.product 主数据缺列: {missing}")
    ci, cp, cn = hdr[PM_INTERNAL], hdr[PM_PZN], hdr[PM_NAME]
    cb, co = hdr.get(PM_BARCODE), hdr.get(PM_ONHAND)
    cid = next((hdr[c] for c in PM_ID_CANDS if c in hdr), None)  # product/ID 选填

    def g(r, c):
        return r[c] if (c is not None and c < len(r)) else None

    master = {}
    for r in rows[1:]:
        intref, pzn_field = g(r, ci), g(r, cp)
        rec = {
            "id": str(g(r, cid)).strip() if g(r, cid) else "",   # ERP 权威主键（选填，原样展示）
            "pid": _norm_pid(g(r, cid)),                      # 同上，归一数字形态（连接用）
            "pzn": norm_pzn(pzn_field) or norm_pzn(intref),   # 官方 PZN 优先，退回 IntRef 嵌入
            "name": str(g(r, cn)).strip() if g(r, cn) else "",
            "barcode": str(g(r, cb)).strip() if g(r, cb) else "",
            "intref": str(intref).strip() if intref else "",
            "onhand": _num(g(r, co)),
        }
        for k in {norm_pzn(intref), norm_pzn(pzn_field)}:  # 两个键都指向同一记录
            if k:
                master.setdefault(k, rec)
    return master


def load_po(path):
    """purchase order 导出（Odoo 行式：订单头字段只在每单首行）→
    (by_pzn={pzn: [line...]}, by_id={产品数字ID: [line...]})。
    双索引同源同 line；导出没带产品 ID 列时 by_id 为空 dict（旧行为）。
    line = {po, vendor(简称), price, qty, date, onhand}。"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    hdr = {str(h).strip(): i for i, h in enumerate(rows[0]) if h is not None}
    for c in (PO_REF, PO_VENDOR, PO_INTERNAL, PO_PRICE, PO_QTY):
        if c not in hdr:
            raise ValueError(f"purchase order 导出缺列: {c}")
    c_date = next((hdr[c] for c in PO_DATE_CANDS if c in hdr), None)
    c_onhand = hdr.get(PO_ONHAND)
    c_pid = next((hdr[c] for c in PO_PID_CANDS if c in hdr), None)

    # 第一遍：ffill 订单头（Vendor），收集全部 vendor 全名以建简称映射
    raw = []
    cur_ref = cur_vendor = None
    for r in rows[1:]:
        if hdr[PO_REF] < len(r) and r[hdr[PO_REF]]:
            cur_ref = r[hdr[PO_REF]]
        if hdr[PO_VENDOR] < len(r) and r[hdr[PO_VENDOR]]:
            cur_vendor = r[hdr[PO_VENDOR]]
        ref = r[hdr[PO_INTERNAL]] if hdr[PO_INTERNAL] < len(r) else None
        if not ref:
            continue
        if cur_vendor and CUSTOMER_PAT in str(cur_vendor).lower():
            continue  # Alibaba 等客户单：整行剔除（噪音，非供应商）
        raw.append((cur_ref, cur_vendor, r))
    vmap = _vendor_map({v for _, v, _ in raw if v})

    out, out_id = {}, {}
    for cur_ref, cur_vendor, r in raw:
        key = norm_pzn(r[hdr[PO_INTERNAL]])
        pid = _norm_pid(r[c_pid]) if (c_pid is not None and c_pid < len(r)) else None
        if key is None and pid is None:    # 两键皆无的采购行（运费/服务等）跳过
            continue
        line = {
            "po": cur_ref,
            "vendor": vmap.get(cur_vendor, cur_vendor) if cur_vendor else "",
            "price": _num(r[hdr[PO_PRICE]]) if hdr[PO_PRICE] < len(r) else None,
            "qty": _num(r[hdr[PO_QTY]]) if hdr[PO_QTY] < len(r) else None,
            "date": _pdate(r[c_date]) if (c_date is not None and c_date < len(r)) else None,
            "onhand": _num(r[c_onhand]) if (c_onhand is not None and c_onhand < len(r)) else None,
        }
        if key is not None:
            out.setdefault(key, []).append(line)
        if pid is not None:
            out_id.setdefault(pid, []).append(line)
    return out, out_id


# 产出列名英文化（给非中文同事）：身份字段优先取 product.product 主数据；「平台裸价/总需求」
# 是待发货表摘出的中文域词、仅中文输入时才出现（否则被 DROP_IF_EMPTY 丢掉），故保留中文。
COLS = ["Product ID", "PZN", "Name", "Barcode", "Internal Reference", "总需求",
        "Quantity On Hand", "Reorder Qty", "平台裸价", "Last Unit Price", "Price Diff",
        "Last Vendor", "Last Qty", "Last Order Date", "Recent Purchases"]
NO_PO_MARK = "No purchase record"   # 无采购记录标记（Last Vendor 列占位值）
# 纯数字列居中，其余（PZN/名称/条码/引用/日期/vendor/记录）左对齐下沉
NUM_COLS = {"总需求", "Quantity On Hand", "Reorder Qty", "平台裸价", "Last Unit Price",
            "Price Diff", "Last Qty"}
LEFT_COLS = set(COLS) - NUM_COLS
WIDTHS = {"Name": 30, "Last Vendor": 18, "Recent Purchases": 56,
          "Barcode": 15, "Internal Reference": 22, "PZN": 12, "Product ID": 12}
# 整批都无值时整列删掉（如仅 PZN 清单没 Name/Barcode/IntRef/总需求/裸价；
# 或无 master / master 不带 product/ID 时的 Product ID）；
# 有值/部分有值则保留、缺处留空，列序不变。PZN/库存/采购画像 恒在。
DROP_IF_EMPTY = ["Product ID", "Name", "Barcode", "Internal Reference", "总需求",
                 "Reorder Qty", "平台裸价", "Price Diff"]


def _round(v, n):
    return round(v, n) if v is not None else ""


def build_rows(demand, caps, po, master=None, po_id=None):
    master = master or {}
    po_id = po_id or {}
    rows = []
    for d in demand:
        m = master.get(d["pzn"])
        # 身份：有主数据取权威(官方 PZN/Name/Barcode/IntRef)，否则退回 demand
        disp_pzn = (m["pzn"] if m and m["pzn"] else d["pzn"]) or ""
        name = (m["name"] if m else d["name"]) or ""
        barcode = m["barcode"] if m else ""
        intref = (m["intref"] if m else d["intref"]) or ""
        # PO 连接：ID 优先（demand 自带 ID，或 PZN→master 桥出的 ID），逐行回退 PZN。
        # ID 命中即用（绕开 PZN 空白/脏值/IntRef 嵌旧 PZN）；无 ID 输入不传 master 时全走 PZN，旧行为不变。
        join_pid = d["pid"] or (m["pid"] if m else None)
        # PZN 回退键：有主数据用其 IntRef 嵌入 PZN(桥接待发货表官方 PZN → PO 嵌入 PZN)，否则用 demand PZN
        join_pzn = (norm_pzn(m["intref"]) if m and m["intref"] else None) or d["pzn"]
        lines = (po_id.get(join_pid) if join_pid else None) or po.get(join_pzn, [])
        dated = sorted([l for l in lines if l["date"]], key=lambda x: x["date"], reverse=True)
        # 当前库存：主数据优先，否则取采购行里任一非空（优先最近记录）
        onhand = (m["onhand"] if (m and m["onhand"] is not None) else
                  next((l["onhand"] for l in dated if l["onhand"] is not None),
                       next((l["onhand"] for l in lines if l["onhand"] is not None), None)))
        cap = caps.get(d["pzn"])
        need = d["need"]
        reorder = (need - onhand) if (need is not None and onhand is not None) else need
        # ERP product/ID 展示值：master 权威优先，退回 demand 自带（ERP 输入）；均无则空 → 整列被 DROP_IF_EMPTY 丢
        pid = (m["id"] if m and m["id"] else "") or d["pid_raw"]
        ident = [pid, disp_pzn, name, barcode, intref]

        if dated or lines:
            last = dated[0] if dated else lines[0]
            price = last["price"]
            diff = (cap - price) if (cap is not None and price is not None) else ""
            recent = "\n".join(
                "{d}|{po}|{v}|{q}|@{p}".format(
                    d=l["date"].strftime("%m-%d") if l["date"] else "??",
                    po=l["po"] or "?",  # PO 单号，便于回 ERP 按号查明细
                    v=l["vendor"] or "?",
                    q=("" if l["qty"] is None else f"{l['qty']:g}"),
                    p=("" if l["price"] is None else f"{l['price']:g}"))
                for l in (dated or lines)[:RECENT_N])
            rows.append(ident + [
                "" if need is None else int(need),
                "" if onhand is None else int(onhand),
                "" if reorder is None else int(reorder),
                _round(cap, 4), _round(price, 2), _round(diff, 2) if diff != "" else "",
                last["vendor"] or "",
                "" if last["qty"] is None else f"{last['qty']:g}",
                last["date"].strftime("%Y-%m-%d") if last["date"] else "",
                recent,
            ])
        else:
            rows.append(ident + [
                "" if need is None else int(need),
                "" if onhand is None else int(onhand),
                "" if reorder is None else int(reorder),
                _round(cap, 4), "", "", NO_PO_MARK, "", "", "",
            ])
    return rows


def build(demand_path, po_path, out_path=None, master_path=None):
    demand = load_demand(demand_path)
    caps = load_caps(demand_path)
    po, po_id = load_po(po_path)
    master = load_master(master_path) if master_path else None
    rows = build_rows(demand, caps, po, master, po_id=po_id)

    import pandas as pd
    df = pd.DataFrame(rows, columns=COLS)
    for c in DROP_IF_EMPTY:
        if df[c].map(lambda v: v in ("", None)).all():
            df = df.drop(columns=c)
    if out_path:
        outdir = os.path.dirname(out_path) or "."
        fname = os.path.basename(out_path)
    else:
        outdir = os.path.join("output", f"{date.today():%Y%m%d}")
        fname = f"订货辅助-{date.today():%Y%m%d}.xlsx"
    os.makedirs(outdir, exist_ok=True)
    path, n = _write_simple(df, outdir, fname, left_cols=LEFT_COLS, widths=WIDTHS)

    vcol = COLS.index("Last Vendor")
    matched = sum(1 for r in rows if r[vcol] != NO_PO_MARK)
    return path, n, matched


def main():
    # 位置参: <待发货明细表/PZN清单/销售分析> <purchase order> [out.xlsx]
    # 选填: --master <product.product.xlsx>  富化 PZN/Name/Barcode/Internal Reference
    args = [a for a in sys.argv[1:]]
    master_path = None
    if "--master" in args:
        i = args.index("--master")
        master_path = args[i + 1] if i + 1 < len(args) else None
        del args[i:i + 2]
    if len(args) < 2:
        print("用法: python3 reorder_helper.py <待发货明细表.xlsx> <purchase order.xlsx> "
              "[out.xlsx] [--master product.product.xlsx]")
        return
    path, n, matched = build(args[0], args[1], args[2] if len(args) > 2 else None,
                             master_path=master_path)
    print(f"订货辅助表已生成: {path}")
    print(f"  共 {n} 个货品，其中 {matched} 个有采购记录，{n - matched} 个无匹配")


if __name__ == "__main__":
    main()
