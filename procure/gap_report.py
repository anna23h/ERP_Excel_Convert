"""销售单驱动的采购缺口 + 产品 × 供应商的报价比较。

用法：
    python3 -m procure.gap_report S04018 S04029
    python3 -m procure.gap_report S04018 --months 6 -o /tmp/x

做什么：给定若干销售单，把「每个产品还缺多少、曾经在哪几家买过、上次什么价」摊平成
一张 xlsx，在「报价比较」上按产品 × 供应商横向摊开、就地问价。**选谁下单仍留给人**——
供应商会缺货、会涨价，这张表不做决定。

2026-08-30 起**一家一张「询价-x」底稿已删掉**：按供应商看是这张表上筛一次的事，
多 6 张 sheet 只增加理解成本（用户裁定）。报价比较随之从只读汇总升为**唯一录入口**，
每家 9 列 = 4 列 ERP 采购历史（刷新）+ 5 列人填（保留）。

缺口公式（四个中间量各占一列，否则没法复核）：
    需求     = Σ(product_uom_qty − qty_delivered)，目标单合并同产品
    他单占用 = 全局待出库 − 本次目标单自己的占用      ← 不剥离就把需求扣两遍
    可分配   = max(0, 在手 qty_available − 他单占用)
    缺口     = max(0, 需求 − 可分配)

两个会让表安静地骗人的坑，代码里都显式处理了：
  1. **双重扣减**：目标单自己也在全局占用里。`_occupancy_a/b` 都按 order_id 剥离目标单。
  2. **跨表重复下单**：同一产品会出现在多家供应商的底稿上，缺口都写**总缺口**。
     报价比较里同一行横着就能看到还有哪几家，汇总对账 sheet 留「已下单量/剩余」给人回填。

他单占用为什么只用一条路（口径演化见 docs/journal/2026-08-27.md）：
    封版时定的是 A（待出库 move 聚合）与 B（已确认单未交付量）两路对拍。实测 B 在本库
    **被历史欠交积压彻底污染**——确认单少发了既没补也没取消就挂着，最老到 2021-11，
    单品最多比 A 多 12102 件；抽查 `qty_delivered` 维护正常，是业务烂尾单不是字段坏。
    → **A 进公式，B 不进表**，只保留反向检查 `A > B`（有出库 move 却无对应确认单行，
    才是 picking 与 SO 脱钩的真异常），降级成运行时终端警告。

「随订货阶段自动维护」是怎么做到的：
    重跑时读回上一版 xlsx，把人填的六类内容（报价/可供/保质期/交期、分给本单、已下单量）
    按 `产品代码 (+ 供应商简称)` 搬进新表。旧版的「询价-x」sheet 仍会被读（改版不丢手填）。**ERP 那侧每跑一次刷新一次，人填的原样保留。**
    刻意没走「从 Odoo 采购单自动回填已下单量」：本库 PO 是收到账单、到货之后才建的
    （2026-08-27 实测近 90 天 707 张 PO，95% 已 done、86% 的行全额收货），拿它当下单
    进度会滞后整整一个采购周期，比手填更差。
"""
import argparse
import datetime as dt
import os
import re
import sys
from collections import defaultdict

# Windows 控制台默认 cp936/cp1252，打印 ⚠ · — 这类符号会 UnicodeEncodeError 直接崩，
# 强制 UTF-8（同 make_labels.py 的做法）。公司 Windows 上跑本脚本时必需。
for _stream in (sys.stdout, sys.stderr):
    if hasattr(_stream, "reconfigure"):
        _stream.reconfigure(encoding="utf-8")

from openpyxl import Workbook

from common import localconf
from common.vendor import short_vendor
from common.xlsx import style_sheet, unique_path
from odoo_api import stock_report as sr
from odoo_api.odoo_client import Odoo, OdooError, m2o_id, m2o_name

#: 采购单里不是真实进货的对手方，整行剔除（同 common/po.PO_NOISE_PATS 的口径）。
#: Alibaba Health 是伪装成供应商的**我方客户**；VO Test Order 是建虚拟库存映射的测试单。
VENDOR_NOISE = ["Alibaba Health", "VO Test Order"]

#: B 路线里「陈年未交付」的判定阈值（天）。超过这个岁数还没发货的确认单，
#: 在本库几乎全是没人清理的死单，不是真实占用。
STALE_DAYS = 180


def say(msg):
    print(msg, flush=True)


# --------------------------------------------------------------------------
# 取数
# --------------------------------------------------------------------------
def pull_demand(od, so_names):
    """目标销售单 → ({pid: {单号: 需求}}, {pid: 产品信息}, [单号信息])。

    只留可库存产品（`type='product'`）：服务、运费行不采购，留着只会在表里当噪音。
    `display_type` 非空的是分节标题/备注行，没有产品。
    """
    orders = od.search_read_all("sale.order", [("name", "in", list(so_names))],
                                ["name", "state", "partner_id", "date_order"])
    found = {o["name"] for o in orders}
    missing = [n for n in so_names if n not in found]
    if missing:
        raise OdooError(f"找不到销售单：{missing}（单号大小写敏感，确认没写错）")
    oids = [o["id"] for o in orders]
    lines = od.search_read_all(
        "sale.order.line", [("order_id", "in", oids)],
        ["order_id", "product_id", "product_uom_qty", "qty_delivered",
         "display_type", "product_uom"])

    pids = sorted({m2o_id(l["product_id"]) for l in lines
                   if not l.get("display_type") and l.get("product_id")})
    prods = {p["id"]: p for p in od.execute(
        "product.product", "read",
        [pids, ["default_code", "name", "type", "uom_id", "qty_available"]])}

    demand, skipped = defaultdict(lambda: defaultdict(float)), []
    for l in lines:
        if l.get("display_type") or not l.get("product_id"):
            continue
        pid = m2o_id(l["product_id"])
        if prods[pid]["type"] != "product":
            skipped.append(prods[pid]["default_code"] or prods[pid]["name"])
            continue
        qty = (l["product_uom_qty"] or 0) - (l["qty_delivered"] or 0)
        if qty > 0:
            demand[pid][m2o_name(l["order_id"])] += qty
    if skipped:
        say(f"  跳过非库存行 {len(skipped)} 条（服务/运费等）：{'、'.join(skipped[:5])}")
    say(f"  目标单 {len(orders)} 张 / 明细 {len(lines)} 行 / 需采购产品 {len(demand)} 个")
    for o in orders:
        say(f"    {o['name']}  {o['date_order'][:10]}  state={o['state']}  "
            f"{m2o_name(o['partner_id'])}")
    return demand, prods, orders, oids


def occupancy_a(od, pids, exclude_oids):
    """路线 A：待出库 stock.move 聚合 → ({pid: 他单占用}, {pid: 本单占用})。

    口径与 `stock_report.pull_in_transit` 的第二个返回值一致
    （internal → customer/transit，state not in done/cancel/draft），
    但这里**按产品收窄并按订单拆分**，因为要把目标单自己的占用剥出来（双重扣减）。
    """
    mv = od.search_read_all(
        "stock.move",
        [("state", "not in", ["done", "cancel", "draft"]),
         ("location_id.usage", "=", "internal"),
         ("location_dest_id.usage", "in", ["customer", "transit"]),
         ("product_id", "in", pids)],
        ["product_id", "product_uom_qty", "sale_line_id", "origin", "date"])
    # move → 它属于哪张 SO：优先 sale_line_id（可靠），退回 origin 字符串（PO 号可能撞，
    # 但这里只用于剥离目标单，撞了顶多少剥一点，方向是保守的）
    sl_ids = [m2o_id(m["sale_line_id"]) for m in mv if m.get("sale_line_id")]
    line_order = {}
    if sl_ids:
        for r in od.execute("sale.order.line", "read", [sorted(set(sl_ids)), ["order_id"]]):
            line_order[r["id"]] = m2o_id(r["order_id"])
    other, mine = defaultdict(float), defaultdict(float)
    for m in mv:
        pid, qty = m2o_id(m["product_id"]), m["product_uom_qty"] or 0
        oid = line_order.get(m2o_id(m["sale_line_id"])) if m.get("sale_line_id") else None
        (mine if oid in exclude_oids else other)[pid] += qty
    if sum(mine.values()):
        say(f"  已从全局待出库中剥离目标单自身占用 {sum(mine.values()):.0f} 件"
            f"（{len(mine)} 个产品）——不剥离就会把需求扣两遍")
    else:
        say("  目标单自身在全局待出库里占 0 件（未确认的单不生成出库 move，符合预期）")
    return other, mine


def occupancy_b(od, pids, exclude_oids):
    """路线 B：已确认 SO 行的未交付量 → ({pid: 全量}, {pid: 超龄部分})。

    ⚠ state 必须取 `sale` **和** `done`：`done` 是「已锁定」的确认单，它照样能有未完成的
    出库 picking。2026-08-27 实测漏了 done 就会让 A 凭空高出 B（Mykoderm 高 1417 件、
    Ursofalk 高 444 件），把一个纯口径失误伪装成数据质量问题。

    不依赖 picking 是否生成，但会漏掉非销售来源的出库（调拨、生产领料），
    且会把陈年死单算进来——所以只作对拍，不进公式。
    """
    rows = od.search_read_all(
        "sale.order.line",
        [("state", "in", ["sale", "done"]), ("product_id", "in", pids)],
        ["order_id", "product_id", "product_uom_qty", "qty_delivered"])
    oids = sorted({m2o_id(r["order_id"]) for r in rows})
    dates = {}
    if oids:
        for o in od.execute("sale.order", "read", [oids, ["date_order"]]):
            dates[o["id"]] = o["date_order"][:10]
    cut = (dt.date.today() - dt.timedelta(days=STALE_DAYS)).isoformat()
    total, stale = defaultdict(float), defaultdict(float)
    for r in rows:
        oid = m2o_id(r["order_id"])
        if oid in exclude_oids:
            continue
        qty = (r["product_uom_qty"] or 0) - (r["qty_delivered"] or 0)
        if qty <= 0:
            continue
        pid = m2o_id(r["product_id"])
        total[pid] += qty
        if dates.get(oid, "9999") < cut:
            stale[pid] += qty
    return total, stale


def pull_po_history(od, pids, months):
    """采购历史 → {pid: [{供应商, 简称, 次数, 最近日期, 最近单价, 最近数量, 窗口内?}]}。

    窗口内（默认 3 个月）为空的产品**补一行历史最近一次并标灰**——否则慢动销品在这张
    表上开天窗，等于对它没做。
    """
    cut = (dt.date.today() - dt.timedelta(days=30 * months)).isoformat()
    rows = od.search_read_all(
        "purchase.order.line",
        [("product_id", "in", pids), ("state", "in", ["purchase", "done"])],
        ["product_id", "partner_id", "product_qty", "price_unit", "date_order",
         "qty_received", "order_id"], label="采购历史")
    noise = re.compile("|".join(re.escape(n) for n in VENDOR_NOISE), re.I)
    per_vendor = defaultdict(list)          # (pid, vendor_name) → [行]
    for r in rows:
        vendor = m2o_name(r["partner_id"]) or "?"
        if noise.search(vendor):
            continue
        per_vendor[(m2o_id(r["product_id"]), vendor)].append(r)

    out = defaultdict(list)
    for (pid, vendor), rs in per_vendor.items():
        rs.sort(key=lambda r: r["date_order"])
        win = [r for r in rs if r["date_order"][:10] >= cut]
        last = (win or rs)[-1]
        out[pid].append({
            "vendor": vendor, "short": short_vendor(vendor),
            "n": len(win), "in_window": bool(win),
            "date": last["date_order"][:10],
            "price": last["price_unit"] or 0.0,
            "qty": last["product_qty"] or 0.0,
            "po": m2o_name(last["order_id"]),
        })
    # 窗口内有记录的产品：只保留窗口内的供应商；窗口内全空：保留历史最近一家（标灰）
    for pid, vs in out.items():
        win = [v for v in vs if v["in_window"]]
        if win:
            out[pid] = sorted(win, key=lambda v: (-v["n"], v["date"]), reverse=False)
        else:
            out[pid] = [max(vs, key=lambda v: v["date"])]
    n_fallback = sum(1 for pid in out if not out[pid][0]["in_window"])
    say(f"  采购历史：{len(rows)} 行 → {len(out)}/{len(pids)} 个产品有供应商；"
        f"其中 {n_fallback} 个近 {months} 个月无采购，已回落到历史最近一次（表里标『历史』）")
    return out


def pull_draft_rfq(od, pids, months):
    """未确认的询价单/草稿 PO → {pid: 数量}。只提示，不进公式。

    **必须限窗口**（2026-08-27 用户要求，同 `--months`）：不限的话陈年没清的草稿全算进来，
    实测单个产品能到 21100 件，这个提示列就成了噪音。
    """
    cut = (dt.date.today() - dt.timedelta(days=30 * months)).isoformat()
    rows = od.search_read_all(
        "purchase.order.line",
        [("product_id", "in", pids), ("state", "in", ["draft", "sent"]),
         ("date_order", ">=", cut)],
        ["product_id", "product_qty"])
    out = defaultdict(float)
    for r in rows:
        out[m2o_id(r["product_id"])] += r["product_qty"] or 0
    return out



# --------------------------------------------------------------------------
# 上一版手填内容的回收（「随阶段自动维护」就是靠这个）
# --------------------------------------------------------------------------
#: 报价比较上由人填的五列（规范键；表里带供应商前缀，见 VENDOR_FILL）
QUOTE_FIELDS = ["报价单价", "可供数量", "保质期", "交期", "备注"]

#: 报价比较里每家的列组：左 4 列 ERP 采购历史（只读、每次刷新），右 5 列人填。
#: 顺序即「参考在左、待填在右」，`VENDOR_FILL` 与 `QUOTE_FIELDS` 一一对应，不可错位。
VENDOR_HIST = ["采购次数", "最近日期", "最近单价", "最近数量"]
VENDOR_FILL = ["报价", "可供", "保质期", "交期", "备注"]
assert len(VENDOR_FILL) == len(QUOTE_FIELDS)


def _header_row(ws, first="产品代码"):
    """找到表头行号——各 sheet 上方的说明行数不一样，不能写死第 1 行。"""
    for r in range(1, min(ws.max_row, 12) + 1):
        if ws.cell(r, 1).value == first:
            return r
    return None


def _row_dicts(ws, first="产品代码"):
    hr = _header_row(ws, first)
    if hr is None:
        return []
    cols = [ws.cell(hr, c).value for c in range(1, ws.max_column + 1)]
    for r in range(hr + 1, ws.max_row + 1):
        d = {c: ws.cell(r, i + 1).value for i, c in enumerate(cols) if c}
        if d.get(first):
            yield d


#: 回收时视同空的值。`—` 是表里「这家没供过这个产品」的占位符，不是人填的报价——
#: 收进来会被当成真报价再写回新表，实测一次污染 70 处（2026-08-30 发现并修）。
BLANKS = (None, "", "—", "-")


def read_previous(path):
    """读回上一版 xlsx 的手填内容 → dict。文件不存在/读不动就返回空，不报错。

    **这是「表随订货阶段推进而自动维护」的全部实现**：ERP 那侧的数字每次重跑都刷新，
    人填的那六类内容按 `产品代码 (+ 供应商简称)` 原样搬过来。

    刻意不做的：不校验上一版的列是否与本版一致。列改了就当那列没填过——
    宁可少搬一列让人重填，也不要把错位的数据搬进新表还装作没事。
    """
    if not path or not os.path.exists(path):
        return {"quotes": {}, "chosen": {}, "ordered": {}, "alloc": {}}
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    quotes, chosen, ordered, alloc = {}, {}, {}, {}
    for name in wb.sheetnames:
        ws = wb[name]
        if name.startswith("询价-"):
            # 旧版（2026-08-30 之前）的一家一张底稿。这些 sheet 已不再生成，但**必须继续读**
            # ——上一版产出里的手填内容不能因为改版而丢。新版报价比较后读，同键覆盖它。
            vendor = name[3:]
            for d in _row_dicts(ws):
                vals = {k: d.get(k) for k in QUOTE_FIELDS if d.get(k) not in BLANKS}
                # 「近期无采购…」是上一版自己生成的说明，不是人填的；回收它会让一条
                # 已经过期的说明跟着新表一直漂下去。
                if str(vals.get("备注", "")).startswith("近期无采购"):
                    vals.pop("备注")
                if vals:
                    quotes[(d["产品代码"], vendor)] = vals
        elif name == "报价比较":
            # 2026-08-30 起这里是手填的**唯一录入口**（原先只读）。列名形如「P 报价」，
            # 供应商简称可能含空格（PHARMA LUPUS），故按最后一个空格切，后缀才是字段名。
            fill = dict(zip(VENDOR_FILL, QUOTE_FIELDS))
            for d in _row_dicts(ws):
                v = {k: d.get(k) for k in ("选定供应商", "选定数量")
                     if d.get(k) not in BLANKS}
                if v:
                    chosen[d["产品代码"]] = v
                for h, val in d.items():
                    vendor, _, suf = str(h).rpartition(" ")
                    if not vendor or suf not in fill or val in BLANKS:
                        continue
                    quotes.setdefault((d["产品代码"], vendor), {})[fill[suf]] = val
        elif name == "汇总对账":
            for d in _row_dicts(ws):
                if d.get("已下单量(手填)") not in BLANKS:
                    ordered[d["产品代码"]] = d["已下单量(手填)"]
        elif name.startswith("备货-"):
            so = name[3:]
            for d in _row_dicts(ws):
                if d.get("分给本单(手填)") not in BLANKS:
                    alloc[(so, d["产品代码"])] = d["分给本单(手填)"]
    n = len(quotes) + len(chosen) + len(ordered) + len(alloc)
    if n:
        say(f"  已从上一版回收手填内容 {n} 处："
            f"报价 {len(quotes)} / 选定 {len(chosen)} / 已下单 {len(ordered)} / 分配 {len(alloc)}")
    return {"quotes": quotes, "chosen": chosen, "ordered": ordered, "alloc": alloc}


def find_previous(outdir, orders):
    """同目录下同一组目标单的最近一版产出（按修改时间）。没有就 None。"""
    import glob
    # 只认中文版：英文版是只读快照，手填只有中文版一个真相源
    pat = os.path.join(outdir, f"采购缺口-{'+'.join(orders)}-*.xlsx")
    files = [f for f in glob.glob(pat) if not os.path.basename(f).startswith("~$")]
    return max(files, key=os.path.getmtime) if files else None


# --------------------------------------------------------------------------
# 英文版（只读快照）
# --------------------------------------------------------------------------
#: 中文是**唯一规范键**——代码里到处用中文串当列名/字典键，英文只在写盘那一刻翻译。
#: 反过来做（英文当键）会让所有取值逻辑都要先翻一道，出错面大得多。
EN = {
    # sheet 名
    "汇总对账": "Summary", "报价比较": "Quote Comparison", "缺口明细": "Gap Detail",
    # 列名
    "产品代码": "Product Code", "产品名称": "Product Name", "总缺口": "Total Gap",
    "已下单量(手填)": "Ordered Qty", "剩余": "Remaining", "候选供应商": "Candidate Suppliers",
    "本单需求": "Order Demand", "总可用": "Total Available",
    "分给本单(手填)": "Allocated to This Order", "本单缺口": "Order Gap", "备货率": "Coverage",
    "需求合计": "Total Demand", "在手": "On Hand", "他单占用A(待出库)": "Committed (Other Orders)",
    "可分配": "Allocatable", "缺口": "Gap",
    "在途入库(提示)": "Incoming (FYI)", "草稿询价(提示)": "Open RFQ (FYI)",
    "本产品另见": "Also Quoted By", "本家采购次数": "Purchases (This Supplier)",
    "最近一次日期": "Last Purchase Date", "最近一次单价": "Last Unit Price",
    "最近一次数量": "Last Qty",
    "报价单价": "Quoted Price", "可供数量": "Available Qty", "保质期": "Shelf Life / MHD",
    "交期": "Lead Time", "备注": "Notes",
    "选定供应商": "Chosen Supplier", "选定数量": "Chosen Qty",
    # 动态列的后缀
    "需求": "Demand", "报价": "Price", "可供": "Avail.",
    "采购次数": "Purchases", "最近日期": "Last Date", "最近单价": "Last Price",
    "最近数量": "Last Qty",
    # 表内文案
    "历史": "historic",
    "无采购记录": "no purchase history",
    "(历史)": " (historic)",
}
def translator(lang):
    """→ 一个把规范中文串翻成目标语言的函数。zh 时是恒等函数。"""
    if lang == "zh":
        return lambda s: s

    def t(s):
        s = str(s)
        if s in EN:
            return EN[s]
        # 「S04018 需求」「P 报价」「PHARMA LUPUS 最近单价」
        for suffix in ("需求", "采购次数", "最近日期", "最近单价", "最近数量",
                       "报价", "可供", "保质期", "交期", "备注"):
            if s.endswith(" " + suffix):
                return f"{s[:-len(suffix) - 1]} {EN.get(suffix, suffix)}"
        if s.startswith("备货-"):
            return f"Stock-up-{s[3:]}"
        return s
    return t


# --------------------------------------------------------------------------
# 组表
# --------------------------------------------------------------------------
def build_rows(demand, prods, so_names, qty_avail, occ_a, in_transit, draft, po_hist):
    """→ [dict]，每个产品一行。键即规范列名（中文），英文只在写盘时翻。"""
    rows = []
    for pid, per_so in demand.items():
        p = prods[pid]
        need = sum(per_so.values())
        a = occ_a.get(pid, 0.0)
        on_hand = qty_avail.get(pid, 0.0)
        usable = max(0.0, on_hand - a)
        gap = max(0.0, need - usable)
        row = {"产品代码": p["default_code"] or "", "产品名称": p["name"]}
        for n in so_names:
            row[f"{n} 需求"] = per_so.get(n, 0) or None
        row.update({
            "需求合计": need,
            "在手": on_hand,
            "他单占用A(待出库)": a,
            "可分配": usable,
            "缺口": gap,
            "在途入库(提示)": in_transit.get(pid, 0.0) or None,
            "草稿询价(提示)": draft.get(pid, 0.0) or None,
            "_vendors": po_hist.get(pid, []),
            "_pid": pid,
            "_per_so": dict(per_so),
        })
        rows.append(row)
    rows.sort(key=lambda r: (-r["缺口"], r["产品代码"]))
    return rows


def _candidates(r, t):
    return "、".join(t(v["short"]) + (t("(历史)") if not v["in_window"] else "")
                    for v in r["_vendors"]) or t("无采购记录")


def _finish(ws, cols, t, widths=None, left=("产品名称",), small=("产品名称",), header_row=1):
    """统一收尾。所有表的表头都在第 1 行（说明行已于 2026-08-30 全部去掉），
    `header_row` 形参留着是因为 `style_sheet` 的这个坑值得在调用侧显式可见。

    对齐一律交给 `auto_align` 按实际数据判：非纯数字左对齐、纯数字居中。
    `left` 只留给「必然是文字、但当前整列还空着」的列（备注），自动判定看不出来。"""
    style_sheet(ws, len(cols), left_cols={t(c) for c in left},
                small_cols={t(c) for c in small},
                widths={t(k): v for k, v in (widths or {"产品名称": 46}).items()},
                header_row=header_row, auto_align=True)


def sheet_summary(wb, rows, prev, t):
    """汇总对账——**唯一真相源**。已下单量手填，剩余用公式实时算。

    2026-08-30 起表上方**不再写标题/口径/时间戳**（用户裁定：不需要，且长文案会把 A 列
    撑到 140 宽）。生成日期在文件名里，口径在 README 与本文件 docstring 里。"""
    ws = wb.create_sheet(t("汇总对账"), 0)
    top = 1
    cols = ["产品代码", "产品名称", "总缺口", "已下单量(手填)", "剩余", "候选供应商"]
    for j, c in enumerate(cols, start=1):
        ws.cell(top, j, t(c))
    for i, r in enumerate(rows, start=top + 1):
        ws.cell(i, 1, r["产品代码"])
        ws.cell(i, 2, r["产品名称"])
        ws.cell(i, 3, r["缺口"])
        ws.cell(i, 4, prev["ordered"].get(r["产品代码"]))
        ws.cell(i, 5, f"=C{i}-N(D{i})")
        ws.cell(i, 6, _candidates(r, t))
    _finish(ws, cols, t, {"产品名称": 46, "候选供应商": 26},
            small=("产品名称", "候选供应商"), header_row=top)   # top 恒为 1
    return ws


def sheet_stock_up(wb, rows, so, prev, t):
    """一张销售单一张备货进度表。

    为什么「分给本单」是手填而不是算的：在手库存和到货是**几张单共用**的，同一批货
    算给谁，表决定不了（2026-08-27 用户裁定）。表只把总量摆出来，分配留给人。
    """
    ws = wb.create_sheet(t(f"备货-{so}"))
    cols = ["产品代码", "产品名称", "本单需求", "总可用", "分给本单(手填)", "本单缺口", "备货率"]
    for j, c in enumerate(cols, start=1):
        ws.cell(1, j, t(c))
    i = 1
    for r in rows:
        need = r["_per_so"].get(so)
        if not need:
            continue
        i += 1
        ws.cell(i, 1, r["产品代码"])
        ws.cell(i, 2, r["产品名称"])
        ws.cell(i, 3, need)
        ws.cell(i, 4, r["可分配"])
        ws.cell(i, 5, prev["alloc"].get((so, r["产品代码"])))
        ws.cell(i, 6, f"=MAX(0,C{i}-N(E{i}))")
        ws.cell(i, 7, f'=IF(C{i}=0,"",N(E{i})/C{i})')
        ws.cell(i, 7).number_format = "0%"
    _finish(ws, cols, t)
    return ws


def sheet_compare(wb, rows, vendors_all, prev, t):
    """报价比较——产品 × 供应商横向摊开，**唯一录入口**（2026-08-30 用户选定）。

    原先是「只读汇总」，报价填在各家的「询价-x」底稿上。那 6 张底稿已删掉：按供应商看
    本来就是这张表上筛一次的事，多 6 张 sheet 只增加理解成本（用户裁定）。
    每家 9 列，**参考在左、待填在右**——最近单价紧挨报价单价，谈价时一眼对得上。
    """
    ws = wb.create_sheet(t("报价比较"))
    cols = ["产品代码", "产品名称", "缺口"]
    for v in vendors_all:
        cols += [f"{v} {suf}" for suf in VENDOR_HIST + VENDOR_FILL]
    cols += ["选定供应商", "选定数量"]
    for j, c in enumerate(cols, start=1):
        ws.cell(1, j, t(c))
    i = 1
    for r in rows:
        if r["缺口"] <= 0:
            continue
        i += 1
        ws.cell(i, 1, r["产品代码"])
        ws.cell(i, 2, r["产品名称"])
        ws.cell(i, 3, r["缺口"])
        # 同一简称可能对应多个 ERP 供应商全名，取最近一次采购的那条
        mine = {}
        for v in r["_vendors"]:
            cur = mine.get(v["short"])
            if cur is None or v["date"] > cur["date"]:
                mine[v["short"]] = v
        for k, name in enumerate(vendors_all):
            base = 4 + k * len(VENDOR_HIST + VENDOR_FILL)
            v = mine.get(name)
            q = prev["quotes"].get((r["产品代码"], name), {})
            if v is None and not q:
                # 这家没供过这个产品，与「问了没回价」区分开
                ws.cell(i, base, "—")
                continue
            if v is not None:
                ws.cell(i, base, v["n"] if v["in_window"] else t("历史"))
                ws.cell(i, base + 1, v["date"])
                ws.cell(i, base + 2, round(v["price"], 4) or None)
                ws.cell(i, base + 3, v["qty"])
            for off, f in enumerate(QUOTE_FIELDS):
                ws.cell(i, base + len(VENDOR_HIST) + off, q.get(f))
        ch = prev["chosen"].get(r["产品代码"], {})
        ws.cell(i, len(cols) - 1, ch.get("选定供应商"))
        ws.cell(i, len(cols), ch.get("选定数量"))
    # 60+ 列不冻结没法录入：锁住产品代码/名称两列与表头
    ws.freeze_panes = "C2"
    _finish(ws, cols, t, {"产品名称": 46},
            # 「选定供应商」「x 备注」必然是文字，但录入前整列是空的，自动判定看不出来
            left=("产品名称", "选定供应商") + tuple(f"{v} 备注" for v in vendors_all))
    return ws


def sheet_gap(wb, rows, so_names, t):
    ws = wb.create_sheet(t("缺口明细"))
    cols = ["产品代码", "产品名称"] + [f"{n} 需求" for n in so_names] + [
        "需求合计", "在手", "他单占用A(待出库)", "可分配", "缺口",
        "在途入库(提示)", "草稿询价(提示)", "候选供应商"]
    ws.append([t(c) for c in cols])
    for r in rows:
        ws.append([_candidates(r, t) if c == "候选供应商" else r.get(c) for c in cols])
    _finish(ws, cols, t, {"产品名称": 46, "候选供应商": 26},
            small=("产品名称", "候选供应商"))
    ws.freeze_panes = "C2"
    return ws


def build_workbook(rows, so_names, prev, lang="zh"):
    t = translator(lang)
    wb = Workbook()
    wb.remove(wb.active)
    sheet_summary(wb, rows, prev, t)
    for so in so_names:
        sheet_stock_up(wb, rows, so, prev, t)

    by_vendor = defaultdict(list)
    for r in rows:
        if r["缺口"] <= 0:
            continue
        for v in r["_vendors"]:
            by_vendor[v["vendor"]].append((r, v))
    vendors_all = sorted({short_vendor(v) for v in by_vendor},
                         key=lambda s: -sum(len(i) for v, i in by_vendor.items()
                                            if short_vendor(v) == s))
    sheet_compare(wb, rows, vendors_all, prev, t)
    sheet_gap(wb, rows, so_names, t)
    return wb, len(vendors_all)


# --------------------------------------------------------------------------
def run(so_names, months=3, outdir=None, fresh=False, with_en=True):
    outdir = outdir or os.path.join("output", f"{dt.date.today():%Y%m%d}")
    prev_path = None if fresh else find_previous(outdir, so_names)
    prev = read_previous(prev_path)
    if prev_path:
        say(f"· 沿用上一版手填内容：{os.path.basename(prev_path)}")

    od = Odoo.connect()
    say("· 目标销售单")
    demand, prods, orders, oids = pull_demand(od, so_names)
    if not demand:
        raise OdooError("目标单里没有需要采购的库存产品（可能都已发货）。")
    pids = sorted(demand)
    names = [o["name"] for o in orders]

    say("· 在手 / 全局占用")
    qty_avail = sr.pull_qty_available(od, pids)
    occ_a, occ_mine = occupancy_a(od, pids, set(oids))
    occ_b, occ_b_stale = occupancy_b(od, pids, set(oids))
    # A/B 对拍不再进表（见 ISSUES [采购缺口] B）：本库 B 恒被历史欠交积压污染，
    # 唯一还有诊断价值的是反向的 A>B——出库 move 找不到对应确认单行。降级成终端警告。
    odd = [p for p in pids if occ_a.get(p, 0) - occ_b.get(p, 0) > 1]
    excl = sr.resolve_excluded_vendors(od, localconf.get("ODOO_EXCLUDE_VENDORS", []) or [])
    inc, _ = sr.pull_in_transit(od, excl)
    in_transit = {p: inc[p] for p in pids if p in inc}

    say("· 采购历史 / 草稿询价")
    po_hist = pull_po_history(od, pids, months)
    draft = pull_draft_rfq(od, pids, months)

    rows = build_rows(demand, prods, names, qty_avail, occ_a, in_transit, draft, po_hist)
    n_gap = sum(1 for r in rows if r["缺口"] > 0)
    n_nov = sum(1 for r in rows if r["缺口"] > 0 and not r["_vendors"])
    # 表上方不再写标题/口径/时间戳（2026-08-30 用户裁定）：口径只说给跑脚本的人听，
    # 打在终端就够了；写进表里既没人看，长文案还会把 A 列撑到 140 宽。
    say(f"  口径：缺口 = 需求 − max(0, 在手 − 他单占用)；本次剥离目标单自身占用 "
        f"{sum(occ_mine.values()):.0f} 件；采购历史回溯 {months} 个月。")

    os.makedirs(outdir, exist_ok=True)
    stamp = f"{dt.date.today():%Y%m%d}"
    outs = []
    for lang, fname in (
            ("zh", f"采购缺口-{'+'.join(names)}-{stamp}.xlsx"),
            ("en", f"Procurement-Gap-{'+'.join(names)}-{stamp}.xlsx")):
        if lang == "en" and not with_en:
            continue
        wb, n_vendors = build_workbook(rows, names, prev, lang)
        path = unique_path(os.path.join(outdir, fname))
        wb.save(path)
        outs.append(path)

    say(f"\n✓ 已写出 {outs[0]}")
    if len(outs) > 1:
        say(f"  英文版（只读快照，手填只认中文版）：{os.path.basename(outs[1])}")
    say(f"  产品 {len(rows)} 个 / 有缺口 {n_gap} 个 / 报价比较覆盖供应商 {n_vendors} 家")
    if n_nov:
        say(f"  ⚠ {n_nov} 个有缺口的产品查不到采购记录，报价比较里没有可问的家，需人工找供应商")
    if odd:
        say(f"  ⚠ {len(odd)} 个产品的待出库 move 找不到对应的确认单行（picking 与 SO 脱钩），"
            "这几个的占用数不可信，别按它下单："
            + "、".join(prods[p]["default_code"] for p in odd[:6]))
    path = outs[0]
    return path, rows


def main(argv=None):
    ap = argparse.ArgumentParser(description="销售单 → 采购缺口 + 产品×供应商报价比较")
    ap.add_argument("orders", nargs="+", help="销售单号，如 S04018 S04029")
    ap.add_argument("--months", type=int, default=3, help="采购历史回溯月数（默认 3）")
    ap.add_argument("-o", "--outdir", help="输出目录（默认 output/YYYYMMDD）")
    ap.add_argument("--fresh", action="store_true",
                    help="不回收上一版的手填内容，从零生成（默认会回收）")
    ap.add_argument("--no-en", action="store_true",
                    help="不出英文版（默认中英各一份，英文版是只读快照）")
    args = ap.parse_args(argv)
    try:
        run(args.orders, args.months, args.outdir, args.fresh, not args.no_en)
    except OdooError as e:
        print(f"✗ {e}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
