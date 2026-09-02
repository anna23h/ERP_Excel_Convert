"""销售单驱动的采购价格收集表：一个产品一行，摊开它近 N 个月的每一笔采购。

用法：
    python3 -m procure.po_price S04041 S04042
    python3 -m procure.po_price S04041 --months 6 -o /tmp/x

**这个脚本只收集整理，不做任何计算判断。** 成本、毛利、该按哪个价、该找哪家——
全部留给人。表上「成本」列**脚本留空**，由人按实际情况手填。

为什么成本不能由脚本算（2026-09-02 用户用改过的样表定的口径）：
同一个产品在窗口内往往有好几个价，销售数量常常要**跨价分摊**——实测 Gehwol 那行
人是这么填的：`70*4.63` 换行 `74*4.68`（144 件按两个实际进价拆开）。
这既不是一个数、也不是任何一种平均，脚本猜哪种都是错的。同理**不出加权均价**：
均价是算出来的结论，不是收集到的事实，摆在证据表上只会诱导人照抄。

表的三段（列序即阅读顺序）：
    销售侧   销售单 / 产品代码 / 产品名称 / 销售数量      ← 要核算的对象
    人工     成本                                        ← 脚本留空，唯一的产出格
    采购汇总 采购笔数 / 采购总量 / 最低价 / 最高价 / 最近采购日期 / 最近采购价
    逐笔     采购日期 / 供应商 / 采购单号 / 采购数量 / 采购单价  ← 多行单元格，五列行行对齐

**每次都是空白新表**（2026-09-02 用户定）：不像 `gap_report` 那样回收上一版手填内容。
一张销售单只核算一次，回收机制在这里是多余的复杂度。旧产出不会被覆盖
（`unique_path` 自动加序号），填过的表放在那儿就是账。

与 `procure/gap_report.py` 的分工：那张管「还缺多少、找谁买、什么价位」（决策前），
这张管「已经买过哪些、什么价」（核算时）。**两者不合并**——受众和时点都不同。
"""
import argparse
import datetime as dt
import os
import re
import sys
from collections import defaultdict

# Windows 控制台默认 cp936/cp1252，打印 ⚠ · — 这类符号会 UnicodeEncodeError 直接崩，
# 强制 UTF-8（同 gap_report.py / make_labels.py 的做法）。
for _stream in (sys.stdout, sys.stderr):
    if hasattr(_stream, "reconfigure"):
        _stream.reconfigure(encoding="utf-8")

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from common.xlsx import style_sheet, unique_path
from odoo_api.odoo_client import Odoo, OdooError, m2o_id, m2o_name

#: 采购单里不是真实进货的对手方，整行剔除（同 gap_report.VENDOR_NOISE 的口径）。
#: Alibaba Health 是伪装成供应商的**我方客户**，名下是 ECMS 退货包裹单，单价恒为 0；
#: VO Test Order 是建虚拟库存映射的测试单。不剔会把最低价一路拉到 0。
VENDOR_NOISE = ["Alibaba Health", "VO Test Order"]
NOISE = re.compile("|".join(re.escape(n) for n in VENDOR_NOISE), re.I)

#: 采购单显示名的开头即 ERP 单号（P11836），后面还挂着供应商发票号
#: （"P11836 (Proforma Rechnung Nr. 2007825273 (P11836) Rechnung Nr. 45107763)"，
#: 70+ 字符）。放进多行单元格会换行，把逐笔五列的行行对齐冲掉，所以只留 P 编号。
PO_CODE = re.compile(r"^\s*(P\d+)")

COLS = ["销售单", "产品代码", "产品名称", "销售数量", "成本",
        "采购笔数", "采购总量", "最低价", "最高价", "最近采购日期", "最近采购价",
        "采购日期", "供应商", "采购单号", "采购数量", "采购单价", "备注"]

#: 逐笔明细的五列：同一格内一笔一行，五列**行行对齐**，横着读就是一笔完整记录。
#: 列宽都调到不换行——一旦某格换行，这五列的对齐当场就散（这是本表唯一的排版要害）。
DETAIL_COLS = ["采购日期", "供应商", "采购单号", "采购数量", "采购单价"]

#: 固定列宽。逐笔五列按最长内容留足余量：供应商最长是
#: "GEHE Alliance Healthcare Deutschland GmbH"（41 字符，13 号字需 48）。
WIDTHS = {"产品名称": 46, "供应商": 48, "采购日期": 14, "采购单号": 12,
          "采购数量": 12, "采购单价": 12, "备注": 26}


def say(msg):
    print(msg, flush=True)


def num(v):
    """数量/单价 → 尽量短的字符串。整数不拖 `.0`，价格保留必要位数。

    逐笔五列是**文本**（一格多笔），不能靠单元格数字格式，只能在这里定型。
    """
    v = float(v or 0)
    if v == int(v):
        return str(int(v))
    return f"{v:.2f}" if round(v, 2) == round(v, 4) else f"{v:.4f}"


# --------------------------------------------------------------------------
# 取数
# --------------------------------------------------------------------------
def pull_products(od, so_names):
    """目标销售单 → [{pid, 销售单, 产品代码, 产品名称, 销售数量, 分行}]，按单内原顺序去重。

    同一产品会在一张单里出现多次——本库的 SO 用 `line_section` 按**客户的每张采购订单**
    分节，同一个产品在几张客户订单里各占一行（实测 S04042 的 Ferrum 是 5000 + 3200）。
    按产品维度统计，销售数量取**合计**，各行原始数量另记进备注：
    否则那个 8200 从哪来的，看表的人无从复核。

    `display_type` 非空的是分节标题/备注行，没有产品；服务、运费行没有采购历史，
    这里**不做 type 过滤**（与 gap_report 不同）——它们照样列出来、逐笔列为空，
    人一眼看到「这行没采购记录」比它凭空消失要好。
    """
    orders = od.search_read_all("sale.order", [("name", "in", list(so_names))],
                                ["name", "state", "partner_id", "date_order"])
    missing = [n for n in so_names if n not in {o["name"] for o in orders}]
    if missing:
        raise OdooError(f"找不到销售单：{missing}（单号大小写敏感，确认没写错）")
    for o in sorted(orders, key=lambda o: o["name"]):
        say(f"    {o['name']}  {o['date_order'][:10]}  state={o['state']}  "
            f"{m2o_name(o['partner_id'])}")

    lines = od.search_read_all(
        "sale.order.line", [("order_id", "in", [o["id"] for o in orders])],
        ["order_id", "product_id", "product_uom_qty", "display_type", "sequence"])
    lines.sort(key=lambda l: (m2o_name(l["order_id"]), l.get("sequence") or 0, l["id"]))

    out, idx = [], {}
    for l in lines:
        if l.get("display_type") or not l.get("product_id"):
            continue
        pid = m2o_id(l["product_id"])
        qty = l["product_uom_qty"] or 0
        if pid in idx:
            idx[pid]["销售数量"] += qty
            idx[pid]["分行"].append(qty)
            continue
        idx[pid] = {"pid": pid, "销售单": m2o_name(l["order_id"]),
                    "销售数量": qty, "分行": [qty]}
        out.append(idx[pid])

    info = {p["id"]: p for p in od.execute(
        "product.product", "read", [[r["pid"] for r in out], ["default_code", "name"]])}
    for r in out:
        p = info.get(r["pid"], {})
        r["产品代码"] = p.get("default_code") or ""
        r["产品名称"] = p.get("name") or ""
    say(f"  目标单 {len(orders)} 张 / 明细 {len(lines)} 行 / 产品 {len(out)} 个（已按产品合并）")
    return out


def pull_po(od, pids, months):
    """近 N 个月的采购行 → {pid: [行]}（按日期升序）。

    只认 `state ∈ {purchase, done}`：草稿/询价单还没成交，价格不作数。
    一张采购单里同一产品拆成几行的情况**逐行保留**（实测 P11255 的 20 件 + 120 件）——
    这是收集表，合并就等于替人做了判断。
    """
    cut = (dt.date.today() - dt.timedelta(days=30 * months)).isoformat()
    rows = od.search_read_all(
        "purchase.order.line",
        [("product_id", "in", pids), ("state", "in", ["purchase", "done"]),
         ("date_order", ">=", cut)],
        ["product_id", "partner_id", "product_qty", "price_unit", "date_order",
         "order_id"], label="采购历史")
    kept = [r for r in rows if not NOISE.search(m2o_name(r["partner_id"]) or "")]
    say(f"  采购历史 {len(rows)} 行（{cut} 起），剔除退货/测试单 {len(rows) - len(kept)} 行")
    by_pid = defaultdict(list)
    for r in kept:
        by_pid[m2o_id(r["product_id"])].append(r)
    for rs in by_pid.values():
        rs.sort(key=lambda r: r["date_order"])
    return by_pid, cut


# --------------------------------------------------------------------------
# 组表
# --------------------------------------------------------------------------
def build_row(prod, rs, months):
    """产品 + 它的采购行 → 表上的一行。

    「成本」列**不出现在这里**——脚本一个字都不写，留空给人填。
    汇总段只有极值与计数（笔数/总量/最低/最高/最近），**没有均价**：见模块 docstring。
    """
    notes = []
    if len(prod["分行"]) > 1:
        notes.append(f"本单 {len(prod['分行'])} 行合计："
                     + "+".join(num(q) for q in prod["分行"]))
    row = {"销售单": prod["销售单"], "产品代码": prod["产品代码"],
           "产品名称": prod["产品名称"], "销售数量": prod["销售数量"]}
    if not rs:
        notes.append(f"近 {months} 个月无采购记录")
        row["采购笔数"] = 0
        row["备注"] = "；".join(notes)
        return row

    qty = sum(r["product_qty"] or 0 for r in rs)
    prices = [r["price_unit"] or 0 for r in rs]
    last = rs[-1]
    row.update({
        "采购笔数": len(rs), "采购总量": qty,
        "最低价": round(min(prices), 4), "最高价": round(max(prices), 4),
        "最近采购日期": last["date_order"][:10],
        "最近采购价": round(last["price_unit"] or 0, 4),
        "采购日期": "\n".join(r["date_order"][:10] for r in rs),
        "供应商": "\n".join(m2o_name(r["partner_id"]) or "?" for r in rs),
        "采购单号": "\n".join(_po_code(r) for r in rs),
        "采购数量": "\n".join(num(r["product_qty"]) for r in rs),
        "采购单价": "\n".join(num(r["price_unit"]) for r in rs),
        "备注": "；".join(notes),
    })
    return row


def _po_code(r):
    name = m2o_name(r["order_id"]) or ""
    m = PO_CODE.match(name)
    return m.group(1) if m else name


def build_sheet(wb, rows):
    ws = wb.active
    ws.title = "采购价格"
    ws.append(COLS)
    for r in rows:
        ws.append([r.get(c) for c in COLS])
    price_idx = [COLS.index(c) + 1 for c in ("最低价", "最高价", "最近采购价")]
    for r in range(2, ws.max_row + 1):
        for c in price_idx:
            ws.cell(r, c).number_format = "0.0000"
    ws.freeze_panes = "D2"        # 产品代码/名称随滚动留在眼前
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{max(2, ws.max_row)}"
    # 「成本」不进 left_cols/small_cols：用户裁定它的样式与其余列一致，不做任何标记
    # （2026-09-02）。整列为空时 auto_align 判不出方向，跟着默认居中即可。
    style_sheet(ws, len(COLS),
                left_cols={"产品名称", "供应商", "备注", *DETAIL_COLS},
                small_cols={"产品名称", "备注", *DETAIL_COLS},
                widths=WIDTHS, header_row=1, auto_align=True)
    return ws


# --------------------------------------------------------------------------
def run(so_names, months=3, outdir=None):
    outdir = outdir or os.path.join("output", f"{dt.date.today():%Y%m%d}")
    od = Odoo.connect()
    say("· 目标销售单")
    prods = pull_products(od, so_names)
    if not prods:
        raise OdooError("目标单里没有产品行（可能整单都是分节/备注行）。")
    by_pid, cut = pull_po(od, [p["pid"] for p in prods], months)
    rows = [build_row(p, by_pid.get(p["pid"], []), months) for p in prods]

    wb = Workbook()
    build_sheet(wb, rows)
    say(f"· 产品 {len(rows)} 个 / 采购记录 {sum(r.get('采购笔数') or 0 for r in rows)} 笔"
        f"（窗口 {cut} ~ {dt.date.today():%Y-%m-%d}）")
    for r in rows:
        n = r.get("采购笔数") or 0
        rng = (f"{num(r['最低价'])}~{num(r['最高价'])}" if n else "—")
        say(f"    {r['产品代码']:24s} 销售 {num(r['销售数量']):>6s} · "
            f"采购 {n:2d} 笔 / {num(r.get('采购总量')):>6s} 件 · {rng}")
    empty = [r["产品代码"] for r in rows if not r.get("采购笔数")]
    if empty:
        say(f"  ⚠ 近 {months} 个月无采购记录：{'、'.join(empty)}（该行逐笔列为空，需人工另查）")

    os.makedirs(outdir, exist_ok=True)
    path = unique_path(os.path.join(
        outdir, f"采购价格-{'+'.join(so_names)}-{dt.date.today():%Y%m%d}.xlsx"))
    wb.save(path)
    say(f"✓ {path}")
    say("  「成本」列留空，按实际情况手填（可写 70*4.63 这类跨价分摊式，一格多行）")
    return path


def main():
    ap = argparse.ArgumentParser(
        description="销售单 → 近 N 个月采购价格收集表（一品一行，成本列人工填）")
    ap.add_argument("orders", nargs="+", help="销售单号，如 S04041 S04042")
    ap.add_argument("--months", type=int, default=3, help="采购历史回溯月数（默认 3）")
    ap.add_argument("-o", "--outdir", help="输出目录（默认 output/YYYYMMDD）")
    a = ap.parse_args()
    try:
        run(a.orders, a.months, a.outdir)
    except OdooError as e:
        say(f"✗ {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
