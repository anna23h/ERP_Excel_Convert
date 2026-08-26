#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""询价直通看板 · Excel Online 版生成器 (ISSUES: [dashboard] J)

把「采购任务 + 分配条目」两层模型摊成一张可共享的 .xlsx，替掉 claude.ai artifact 看板。
背景见 ISSUES.md 的 J 条：artifact 撑不住多人实时同步（协作者身份下实测也不同步），
而自建服务(F~I)成本收益倒挂，故改走 Excel Online 共享。

三个 sheet:
  采购任务  行式，一行一个采购任务；供应商横向摊 5 组；阶段/已配/缺口/已到/品名全是公式
  看板      五列泳道，只读，卡片按阶段自动归位
  产品字典  export_product_dict.py 的产出，供 VLOOKUP 带出中德品名

**为什么是行式**：Jürgen 现用的 260825 Order template.xlsx 是转置的（每个产品占一整列），
而 Excel 的筛选/排序/透视/条件格式全部假设「一行一条记录」——转置布局下做不出任何看板视图。
字段一个不改，只是躺下来。

**为什么泳道不用 FILTER**：FILTER 经 openpyxl 写出要带 `_xlfn._xlws.` 前缀，且桌面版
Excel 2019 及更早不支持。这里用经典 INDEX+MATCH+COUNTIF 辅助键，Excel Online /
桌面版 / LibreOffice 全通。兼容性优先于写法优雅。
"""
import argparse, json, os, sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule

HERE = os.path.dirname(os.path.abspath(__file__))
DEFAULT_DICT = os.path.join(HERE, "data", "product_dict.json")
DEFAULT_OUT = os.path.join(os.path.dirname(HERE), "results", "询价看板.xlsx")

N_SUPPLIER = 5          # 横向摊几组供应商。现实里一个任务拆 3~5 家（08-25 实测 Excel）
N_ROWS = 300            # 预铺多少行公式
N_CARDS = 25            # 每条泳道最多显示几张卡

# 阶段：照搬 board.html 的 STAGES 顺序与文案（zh/de 并排，两边都要看得懂）
STAGES = [
    ("待询价",              "Anzufragen",                    "E8EDEC"),
    ("询价中",              "In Anfrage",                    "DCE9E4"),
    ("已报价 · 待销售确认",  "Angebot da",                    "FBE9C8"),
    ("已下单",              "Bestellt",                      "CFE3E0"),
    ("已关闭",              "Abgeschlossen",                 "E4E4E4"),
]
# ⚠ 真实供应商名与进货价**不进公开库**（ISSUES.md 抬头的红线）。
# 两者放本地种子文件 dashboard/data/board_seed.json（该目录已 gitignore），
# 没有种子文件时用下面这套占位符——生成的表照样能用，只是名单和样例是假的。
SUPPLIERS = ["供应商 A", "供应商 B", "供应商 C", "供应商 D", "供应商 E"]

TEAL = "0C5A61"          # 沿用看板既有配色（08-25 决定：这是已成立的系统，不重造）
INK = "1B2B2A"
MUTED = "6B7F7C"

# ── 采购任务 sheet 列定义 ────────────────────────────────────────────────
# (中文, 德文, 宽度)  —— 前 13 列固定，之后是 N_SUPPLIER 组供应商块
FIXED_COLS = [
    ("阶段",     "Stage",        22),
    ("PZN",      "PZN",          11),
    ("品名（德）", "Produkt",      34),
    ("品名（中）", "品名",          26),
    ("需求量",   "Bedarf",        9),
    ("已配",     "Zugeteilt",     8),
    ("缺口",     "Fehlt",         8),
    ("已到",     "Erhalten",      8),
    ("参考价",   "AEP/HAP",       9),
    ("下单日",   "Bestelldatum", 12),
    ("要货日",   "Bedarf bis",   12),
    ("提出人",   "Anforderer",   11),
    ("备注",     "Notiz",        22),
]
SUP_BLOCK = [
    ("供应商{}",  "Lieferant {}", 16),
    ("数量{}",    "Menge",         8),
    ("单价{}",    "Preis",         8),
    ("ETA{}",     "ETA",          12),   # 非日期格式：requested / ?? 比日期还多
    ("实收{}",    "Erhalten",      8),
]

# 样例数据：默认是**脱敏占位**（价格取整、供应商用占位名），只为演示五档阶段与泳道归位。
# 想用真实数据出一张能直接给 Jürgen 看的表，就放一份 dashboard/data/board_seed.json：
#   {"suppliers": [...],
#    "sample": [[pzn, 需求量, 参考价, 下单日, 要货日, 提出人, 备注,
#                [[供应商, 数量, 单价, ETA, 实收], ...]], ...]}
# 该文件在 gitignore 里，不会被推上去。
SAMPLE = [
    ("04100371", 2200, None, "2026-08-25", "2026-08-31", "JD", "", [
        ("供应商 A", 140, 10.00, "2026-08-28", None),
        ("供应商 B", 1000, None, "requested", None),
        ("供应商 C", 1000, None, "requested", None)]),
    ("16233255", 2000, 6.00, "2026-08-25", "", "JD", "", [
        ("供应商 B", 200, 5.00, "2026-08-26", None),
        ("供应商 B", 800, 5.00, "requested", None),
        ("供应商 D", 500, 5.00, "??", None)]),
    ("12351236", 100, None, "2026-08-25", "", "Daniel", "", [
        ("供应商 A", 89, 20.00, "2026-08-27", 89)]),
    ("02766290", 1100, None, "2026-08-25", "2026-08-31", "lisa", "", []),
]

SEED = os.path.join(HERE, "data", "board_seed.json")


def load_seed():
    """本地种子（真实供应商名单与样例）覆盖占位符。没有就用占位符，不报错。"""
    if not os.path.exists(SEED):
        return False
    d = json.load(open(SEED, encoding="utf-8"))
    if d.get("suppliers"):
        SUPPLIERS[:] = d["suppliers"]
    if d.get("sample"):
        SAMPLE[:] = [tuple(row[:7]) + ([tuple(a) for a in row[7]],) for row in d["sample"]]
    return True


def col_letters():
    """→ (固定列字母 dict, 每组供应商块的字母 dict list, 辅助列字母)"""
    fixed = {}
    for i, (zh, _de, _w) in enumerate(FIXED_COLS, start=1):
        fixed[zh] = get_column_letter(i)
    blocks = []
    c = len(FIXED_COLS) + 1
    for n in range(1, N_SUPPLIER + 1):
        blocks.append({
            "supplier": get_column_letter(c),
            "qty": get_column_letter(c + 1),
            "price": get_column_letter(c + 2),
            "eta": get_column_letter(c + 3),
            "recv": get_column_letter(c + 4),
        })
        c += len(SUP_BLOCK)
    helper = {"card": get_column_letter(c), "key": get_column_letter(c + 1)}
    return fixed, blocks, helper


def build_task_sheet(ws, fixed, blocks, helper):
    head_fill = PatternFill("solid", fgColor=TEAL)
    head_font = Font(color="FFFFFF", bold=True, size=10)
    sup_fill = PatternFill("solid", fgColor="3E7C7F")
    calc_fill = PatternFill("solid", fgColor="F2F6F5")   # 公式列底色：提示「别手填」
    thin = Side(style="thin", color="C8D6D3")

    # 表头：中文 + 德文两行合一格（换行），三方都看得懂
    col = 1
    for zh, de, w in FIXED_COLS:
        c = ws.cell(1, col, "%s\n%s" % (zh, de))
        c.fill, c.font = head_fill, head_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = w
        col += 1
    for n in range(1, N_SUPPLIER + 1):
        for zh, de, w in SUP_BLOCK:
            c = ws.cell(1, col, "%s\n%s" % (zh.format(n), de.format(n)))
            c.fill, c.font = sup_fill, head_font
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(col)].width = w
            col += 1
    for key in ("card", "key"):
        c = ws.cell(1, col, "(辅助 · 勿删)")
        c.fill, c.font = PatternFill("solid", fgColor="D9D9D9"), Font(size=9, color="808080")
        ws.column_dimensions[get_column_letter(col)].width = 40
        col += 1
    ws.row_dimensions[1].height = 32

    qty_cells = ",".join("%s{r}" % b["qty"] for b in blocks)
    price_cells = ",".join("%s{r}" % b["price"] for b in blocks)
    recv_cells = ",".join("%s{r}" % b["recv"] for b in blocks)
    sup_cells = ",".join("%s{r}" % b["supplier"] for b in blocks)
    F = fixed

    for r in range(2, N_ROWS + 2):
        d = dict(r=r)
        pzn = "$%s%d" % (F["PZN"], r)
        need = "$%s%d" % (F["需求量"], r)
        alloc = "$%s%d" % (F["已配"], r)
        recv = "$%s%d" % (F["已到"], r)
        de_name = "$%s%d" % (F["品名（德）"], r)
        zh_name = "$%s%d" % (F["品名（中）"], r)

        # 品名：PZN → 产品字典。空 PZN 不查（TEXT("") 会变成 00000000 撞上真 PZN）
        # 尾部 &"" 是必须的：字典里 nameZh 为空时 VLOOKUP 返回空单元格，Excel 会显示成 0，
        # 而 0 会一路污染到看板卡片上（2026-08-26 Excel 实跑发现）。
        ws["%s%d" % (F["品名（德）"], r)] = (
            '=IF(%s="","",IFERROR(VLOOKUP(TEXT(%s,"00000000"),\'产品字典\'!$A:$C,2,FALSE)&"",""))'
            % (pzn, pzn))
        ws["%s%d" % (F["品名（中）"], r)] = (
            '=IF(%s="","",IFERROR(VLOOKUP(TEXT(%s,"00000000"),\'产品字典\'!$A:$C,3,FALSE)&"",""))'
            % (pzn, pzn))
        # 派生量：不落库，算出来（沿用 board.html 第二区口径）
        ws["%s%d" % (F["已配"], r)] = '=IF(%s="","",SUM(%s))' % (pzn, qty_cells.format(**d))
        ws["%s%d" % (F["缺口"], r)] = '=IF(%s="","",MAX(0,%s-%s))' % (need, need, alloc)
        ws["%s%d" % (F["已到"], r)] = '=IF(%s="","",SUM(%s))' % (pzn, recv_cells.format(**d))
        # 阶段推导：照搬 deriveStage() 规则，跨分配条目聚合。人不填，自己走。
        ws["%s%d" % (F["阶段"], r)] = (
            '=IF({pzn}="","",'
            'IF(AND({need}>0,N({recv})>={need}),"{s5}",'
            'IF(N({recv})>0,"{s4}",'
            'IF(COUNT({price})>0,"{s3}",'
            'IF(COUNTA({sup})>0,"{s2}","{s1}")))))'
        ).format(pzn=pzn, need=need, recv=recv,
                 price=price_cells.format(**d), sup=sup_cells.format(**d),
                 s1=STAGES[0][0], s2=STAGES[1][0], s3=STAGES[2][0],
                 s4=STAGES[3][0], s5=STAGES[4][0])
        # 卡片文本：看板 sheet 显示的就是它。德语名可能很长，截断保排版。
        ws["%s%d" % (helper["card"], r)] = (
            '=IF({pzn}="","",IF({zh}="",LEFT({de},30),{zh})&CHAR(10)&'
            '"PZN "&TEXT({pzn},"00000000")&"  ·  "&{need}&CHAR(10)&'
            '"已配 "&{alloc}&" / 已到 "&{recv})'
        ).format(pzn=pzn, zh=zh_name, de=de_name, need=need, alloc=alloc, recv=recv)
        # 辅助键：阶段#该阶段内序号 —— 泳道靠它做 MATCH，不用动态数组
        ws["%s%d" % (helper["key"], r)] = (
            '=IF($%s%d="","",$%s%d&"#"&COUNTIF($%s$2:$%s%d,$%s%d))'
            % (F["阶段"], r, F["阶段"], r, F["阶段"], F["阶段"], r, F["阶段"], r))

        for name in ("阶段", "品名（德）", "品名（中）", "已配", "缺口", "已到"):
            cell = ws["%s%d" % (F[name], r)]
            cell.fill = calc_fill
            cell.border = Border(bottom=thin)
        for name in ("PZN", "需求量", "参考价", "下单日", "要货日", "提出人", "备注"):
            ws["%s%d" % (F[name], r)].border = Border(bottom=thin)
        for b in blocks:
            for k in b.values():
                ws["%s%d" % (k, r)].border = Border(bottom=thin)
        ws["%s%d" % (F["阶段"], r)].alignment = Alignment(horizontal="center", vertical="center")

    # 样例数据
    for i, (pzn, need, ref, odate, needby, who, note, allocs) in enumerate(SAMPLE):
        r = 2 + i
        ws["%s%d" % (F["PZN"], r)] = int(pzn)
        ws["%s%d" % (F["需求量"], r)] = need
        if ref is not None:
            ws["%s%d" % (F["参考价"], r)] = ref
        ws["%s%d" % (F["下单日"], r)] = odate
        ws["%s%d" % (F["要货日"], r)] = needby
        ws["%s%d" % (F["提出人"], r)] = who
        ws["%s%d" % (F["备注"], r)] = note
        for j, (sup, q, price, eta, rec) in enumerate(allocs[:N_SUPPLIER]):
            b = blocks[j]
            ws["%s%d" % (b["supplier"], r)] = sup
            ws["%s%d" % (b["qty"], r)] = q
            if price is not None:
                ws["%s%d" % (b["price"], r)] = price
            ws["%s%d" % (b["eta"], r)] = eta       # 文本，不转日期
            if rec is not None:
                ws["%s%d" % (b["recv"], r)] = rec

    # 阶段配色（条件格式）
    for zh, _de, rgb in STAGES:
        ws.conditional_formatting.add(
            "%s2:%s%d" % (F["阶段"], F["阶段"], N_ROWS + 1),
            CellIsRule(operator="equal", formula=['"%s"' % zh],
                       fill=PatternFill("solid", fgColor=rgb)))
    # 缺口 > 0 标红：量还没配齐，是采购最该看的一件事
    ws.conditional_formatting.add(
        "%s2:%s%d" % (F["缺口"], F["缺口"], N_ROWS + 1),
        CellIsRule(operator="greaterThan", formula=["0"],
                   font=Font(color="B3261E", bold=True)))

    # 供应商下拉（可增改，名单在「供应商」sheet）
    dv_sup = DataValidation(type="list", formula1="='供应商'!$A$2:$A$60",
                            allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv_sup)
    for b in blocks:
        dv_sup.add("%s2:%s%d" % (b["supplier"], b["supplier"], N_ROWS + 1))
    # PZN 不在字典里 → 只警告不阻断（字典是药房产品快照，新品不该被挡住）
    dv_pzn = DataValidation(
        type="custom", errorStyle="warning", allow_blank=True,
        formula1='=ISNUMBER(MATCH(TEXT($%s2,"00000000"),\'产品字典\'!$A:$A,0))' % F["PZN"],
        error="这个 PZN 不在产品字典里。可能是新品，也可能填错了——确认无误可以继续。",
        errorTitle="PZN 不在字典中")
    ws.add_data_validation(dv_pzn)
    dv_pzn.add("%s2:%s%d" % (F["PZN"], F["PZN"], N_ROWS + 1))

    ws.freeze_panes = "%s2" % F["需求量"]
    ws.auto_filter.ref = "A1:%s%d" % (helper["key"], N_ROWS + 1)
    ws.column_dimensions[helper["card"]].hidden = True
    ws.column_dimensions[helper["key"]].hidden = True
    ws.sheet_view.showGridLines = False


def build_board_sheet(ws, helper):
    """五列泳道。经典 INDEX+MATCH，无动态数组。"""
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 6
    for i, (zh, de, rgb) in enumerate(STAGES, start=1):
        L = get_column_letter(i)
        ws.column_dimensions[L].width = 30
        h = ws.cell(1, i)
        h.value = '="%s  ("&COUNTIF(\'采购任务\'!$A:$A,"%s")&")"' % (zh, zh)
        h.fill = PatternFill("solid", fgColor=TEAL)
        h.font = Font(color="FFFFFF", bold=True, size=11)
        h.alignment = Alignment(horizontal="center", vertical="center")
        sub = ws.cell(2, i, de)
        sub.font = Font(color=MUTED, size=8)
        sub.alignment = Alignment(horizontal="center", vertical="center")
        for n in range(1, N_CARDS + 1):
            r = 2 + n
            c = ws.cell(r, i)
            c.value = ('=IFERROR(INDEX(\'采购任务\'!${card}:${card},'
                       'MATCH("{zh}#"&{n},\'采购任务\'!${key}:${key},0)),"")'
                       ).format(card=helper["card"], key=helper["key"], zh=zh, n=n)
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            c.font = Font(size=9, color=INK)
            # 空卡不画框：靠条件格式，卡片有内容才显示底色
            ws.conditional_formatting.add(
                "%s%d" % (get_column_letter(i), r),
                FormulaRule(formula=['LEN(%s%d)>0' % (get_column_letter(i), r)],
                            fill=PatternFill("solid", fgColor=rgb)))
            ws.row_dimensions[r].height = 44


def build_dict_sheet(ws, products):
    ws.append(["PZN", "品名（德）Produkt", "品名（中）"])
    for c, w in zip("ABC", (12, 60, 40)):
        ws.column_dimensions[c].width = w
        ws["%s1" % c].fill = PatternFill("solid", fgColor=TEAL)
        ws["%s1" % c].font = Font(color="FFFFFF", bold=True, size=10)
    for p in products:
        ws.append([p.get("pzn") or "", p.get("nameDe") or "", p.get("nameZh") or ""])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:C%d" % (len(products) + 1)


def build_supplier_sheet(ws):
    ws.append(["供应商 / Lieferant"])
    ws["A1"].fill = PatternFill("solid", fgColor=TEAL)
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=10)
    ws.column_dimensions["A"].width = 28
    for s in SUPPLIERS:
        ws.append([s])
    ws["A%d" % (len(SUPPLIERS) + 3)] = "↑ 往上面接着加就行，采购任务表的下拉会自动跟上（最多 58 家）"
    ws["A%d" % (len(SUPPLIERS) + 3)].font = Font(size=9, color=MUTED)


def main():
    ap = argparse.ArgumentParser(description="生成询价直通看板 Excel（Online 共享版）")
    ap.add_argument("--dict", default=DEFAULT_DICT, help="产品字典 JSON（export_product_dict.py 的产出）")
    ap.add_argument("--out", default=DEFAULT_OUT, help="输出 xlsx 路径")
    ap.add_argument("--no-sample", action="store_true", help="不写样例数据，出一张空表")
    a = ap.parse_args()

    seeded = load_seed()
    if a.no_sample:
        del SAMPLE[:]
    if not os.path.exists(a.dict):
        sys.exit("产品字典不存在: %s\n先跑 python3 dashboard/export_product_dict.py" % a.dict)
    products = json.load(open(a.dict, encoding="utf-8"))["products"]

    fixed, blocks, helper = col_letters()
    wb = Workbook()
    ws_board = wb.active
    ws_board.title = "看板"
    ws_task = wb.create_sheet("采购任务")
    ws_dict = wb.create_sheet("产品字典")
    ws_sup = wb.create_sheet("供应商")

    build_task_sheet(ws_task, fixed, blocks, helper)
    build_board_sheet(ws_board, helper)
    build_dict_sheet(ws_dict, products)
    build_supplier_sheet(ws_sup)

    os.makedirs(os.path.dirname(os.path.abspath(a.out)), exist_ok=True)
    wb.save(a.out)
    print("已生成: %s" % a.out)
    print("  产品字典 %d 条（有中文名 %d）" % (
        len(products), sum(1 for p in products if p.get("nameZh"))))
    print("  采购任务 %d 行预铺公式 · 供应商 %d 组 · 泳道每列 %d 张卡"
          % (N_ROWS, N_SUPPLIER, N_CARDS))
    print("  种子: %s" % ("dashboard/data/board_seed.json（真实名单/样例）" if seeded
                          else "占位符（真实供应商与价格不进公开库；放种子文件可覆盖）"))


if __name__ == "__main__":
    main()
