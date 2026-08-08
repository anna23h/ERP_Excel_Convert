#!/usr/bin/env python3
"""采购数量与频次：指定供应商的 purchase.order 导出 → 每产品的「采购次数 + 数量」+ 逐笔明细。

原始需求（2026-08-08）：同事要一张能看出「在某供应商处采购的产品 数量 + 频次」的表，
自己看规律（供应商会议用）。**纯数据整理，不做任何结论/分析**——卖穿率、价格趋势、
客户维度都不在这里（那些是同一次会话延伸出去的另一回事，见 docs/journal/2026-08-08.md）。

用法:
    python3 po_frequency/po_frequency.py <purchase.order.xlsx> [--vendor NAME] [--out PATH]

输入：purchase.order 的**行式导出**（Order Lines 粒度；订单头字段只在每单首行，脚本自动 ffill）。
      导出通常已在 ERP 里筛到单一供应商；--vendor 可再按供应商名（子串、忽略大小写）过滤。
必需列：Order Reference, Confirmation Date,
        Order Lines/Product/Internal Reference, Order Lines/Product/Display Name,
        Order Lines/Total Quantity, Order Lines/Unit Price（--vendor 时还需 Vendor）。

产出（默认 output/<Vendor>_Purchase_Quantity_and_Frequency.xlsx，英文表头）:
    Summary  每产品一行：Purchase Count(Frequency) / Total Qty / Avg·Min·Max per Purchase /
             First·Last Purchase / Span(days) / Avg Interval(days)，按频次降序。
    Details  每明细行：Order Date / PO Number / Product / Internal Reference / Qty / Unit Price €。

口径（与 2026-08-08 交付表一致）:
    ① 订单头字段向下 ffill；
    ② SKU 归一（common/po._po_base_sku，去 xN/XN·*N·_VO·_GW 尾缀，令多件装/渠道变体并到一行）；
    ③ 频次 = 不同 Order Reference 去重计数（同单同产品多行算 1 次）。
"""
import argparse
import os
import re
import sys

import pandas as pd

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))  # 让 common/ 可导入
from common.po import _po_base_sku  # noqa: E402
from common.xlsx import style_sheet, unique_path, write_df  # noqa: E402
from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

# Details 是原始明细（上万行）：用轻量表头样式，不走 common.style_sheet 的逐格描边+逐行算高
# ——那套按几百行的拣货单设计，对 1.3 万行会慢到 ~25s 且文件臃肿。Summary(约 1700 行、
# 给人读)仍走统一排版。
_DETAIL_WIDTHS = [12, 12, 44, 22, 9, 12]

F_PO = "Order Reference"
F_DATE = "Confirmation Date"
F_VENDOR = "Vendor"
F_SKU = "Order Lines/Product/Internal Reference"
F_NAME = "Order Lines/Product/Display Name"
F_QTY = "Order Lines/Total Quantity"
F_PRICE = "Order Lines/Unit Price"

SUMMARY_COLS = ["Product", "Internal Reference", "Purchase Count (Frequency)", "Total Qty",
                "Avg per Purchase", "Min per Purchase", "Max per Purchase",
                "First Purchase", "Last Purchase", "Span (days)", "Avg Interval (days)"]
DETAIL_COLS = ["Order Date", "PO Number", "Product", "Internal Reference", "Qty", "Unit Price €"]


def load(path, vendor=None):
    """读行式导出 → ffill 订单头 + 归一 SKU + 数值化。缺必需列抛 ValueError。"""
    po = pd.read_excel(path, dtype=str)
    need = [F_PO, F_DATE, F_SKU, F_NAME, F_QTY, F_PRICE] + ([F_VENDOR] if vendor else [])
    missing = [c for c in need if c not in po.columns]
    if missing:
        raise ValueError("采购导出缺列: " + ", ".join(missing))
    head = [F_PO, F_DATE] + ([F_VENDOR] if F_VENDOR in po.columns else [])
    po[head] = po[head].ffill()
    po = po.dropna(subset=[F_SKU]).copy()
    if vendor:
        po = po[po[F_VENDOR].astype(str).str.contains(vendor, case=False, na=False)].copy()
        if po.empty:
            raise ValueError(f"按 --vendor '{vendor}' 过滤后无数据；请核对供应商名。")
    po["_sku"] = po[F_SKU].map(_po_base_sku)
    po["_name"] = (po[F_NAME].astype(str)
                   .str.replace(r"^\[[^\]]*\]\s*", "", regex=True)  # 去 [Internal_Ref] 前缀
                   .str.replace(r"\n.*", "", regex=True))            # 只留首行
    po["_qty"] = pd.to_numeric(po[F_QTY], errors="coerce")
    po["_price"] = pd.to_numeric(po[F_PRICE], errors="coerce")
    po["_date"] = pd.to_datetime(po[F_DATE], errors="coerce")
    return po


def _fdate(d):
    return d.strftime("%Y-%m-%d") if pd.notna(d) else ""


def build(po):
    """归一后的 df → (summary_df, details_df)，列名即英文表头。"""
    name = po.groupby("_sku")["_name"].first()
    # 每次采购 = 每张 Order Reference 的合计（同单同产品多行先并起来）
    per = po.groupby(["_sku", F_PO])["_qty"].sum()
    g = per.groupby("_sku")
    dt = po.groupby("_sku")["_date"]
    s = pd.DataFrame({
        "cnt": g.size(),                       # 不同 PO 数 = 频次
        "tot": po.groupby("_sku")["_qty"].sum(),
        "avg": g.mean(), "mn": g.min(), "mx": g.max(),
        "first": dt.min(), "last": dt.max(),
    })
    s["span"] = (s["last"] - s["first"]).dt.days
    s = s.sort_values(["cnt", "tot"], ascending=False)

    srows = []
    for sku, r in s.iterrows():
        cnt = int(r["cnt"])
        span = int(r["span"]) if pd.notna(r["span"]) else None
        interval = round(span / (cnt - 1)) if cnt > 1 and span is not None else None
        srows.append([
            name.get(sku, sku), sku, cnt,
            int(r["tot"]) if pd.notna(r["tot"]) else None,
            round(float(r["avg"]), 1) if pd.notna(r["avg"]) else None,
            int(r["mn"]) if pd.notna(r["mn"]) else None,
            int(r["mx"]) if pd.notna(r["mx"]) else None,
            _fdate(r["first"]), _fdate(r["last"]), span, interval,
        ])
    summary = pd.DataFrame(srows, columns=SUMMARY_COLS)

    d = po.sort_values(["_sku", "_date"])
    details = pd.DataFrame({
        "Order Date": d["_date"].map(_fdate),
        "PO Number": d[F_PO],
        "Product": d["_name"],
        "Internal Reference": d["_sku"],
        "Qty": d["_qty"].map(lambda v: int(v) if pd.notna(v) else None),
        "Unit Price €": d["_price"].map(lambda v: round(float(v), 2) if pd.notna(v) else None),
    }, columns=DETAIL_COLS)
    return summary, details


def _vendor_slug(po, vendor_arg):
    if F_VENDOR in po.columns and po[F_VENDOR].notna().any():
        v = po[F_VENDOR].dropna().mode().iloc[0]
    else:
        v = vendor_arg or "Vendor"
    slug = re.sub(r"[^A-Za-z0-9]+", "_", str(v)).strip("_")
    return "_".join(slug.split("_")[:2]) or "Vendor"  # 取前两段，够辨识又不冗长


def _style_details(ws, widths):
    """轻量表头样式：蓝底白字表头 + 冻结首行 + 固定列宽，不逐格描边（省时间省体积）。"""
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_report(summary, details, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    write_df(ws, summary)
    style_sheet(ws, len(SUMMARY_COLS), left_cols={"Product", "Internal Reference"})
    ws2 = wb.create_sheet("Details")
    write_df(ws2, details)
    _style_details(ws2, _DETAIL_WIDTHS)
    path = unique_path(out_path)
    wb.save(path)
    return path


def main(argv=None):
    ap = argparse.ArgumentParser(description="指定供应商采购导出 → 数量+频次表（Summary/Details）")
    ap.add_argument("purchase_xlsx", help="purchase.order 行式导出 (.xlsx)")
    ap.add_argument("--vendor", help="按供应商名过滤（子串、忽略大小写）；导出已单供应商时可省略")
    ap.add_argument("--out", help="输出路径；默认 output/<Vendor>_Purchase_Quantity_and_Frequency.xlsx")
    args = ap.parse_args(argv)

    po = load(args.purchase_xlsx, vendor=args.vendor)
    summary, details = build(po)
    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    if args.out:
        out_path = args.out
    else:
        outdir = os.path.join(root, "output")
        os.makedirs(outdir, exist_ok=True)
        out_path = os.path.join(outdir, f"{_vendor_slug(po, args.vendor)}_Purchase_Quantity_and_Frequency.xlsx")
    path = write_report(summary, details, out_path)
    print(f"产品 {len(summary)} 个 / 明细 {len(details)} 行 → {path}")


if __name__ == "__main__":
    main()
