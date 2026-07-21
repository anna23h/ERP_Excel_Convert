#!/usr/bin/env python3
"""第二阶段：仓库反馈无货后，从"实际发货订单"生成 B/C/D。

实际发货 = 步骤4 的「待全集」 − 无货清单。

用法:
    python3 stage2.py --mmdd 0610 \
        [--erp raw_data/xxx.xlsx] [--tmall raw_data/天猫测试.xlsx] \
        [--nogoods raw_data/无货清单.xlsx] \
        [--billing raw_data/账单模板导出.xlsx] \
        [--outdir output]

- 无货清单：含 Order Reference 或 系统履约单号(15位) 任一列即可；不传则视为全部发货。
- 账单模板导出(D 用)：ERP 导出，需含 External ID(ID) + Order Reference + Terms and conditions(原值=渠道+运单)。
  不传则跳过 D。
"""
import argparse, os, re
from datetime import date
import pandas as pd
from openpyxl import Workbook

import step4_merge as s4
import build_excel as be  # 复用样式


def last15(series):
    return series.astype(str).str[-15:]


SCP_RE = re.compile(r"SCP\d+")   # 连接键: 系统履约单号(SCP…) / Order Reference(VO_TOF_SCP…)

# 出库单 Carrier/ID 固定填充值(承运商外部 ID)
CARRIER_ID = "__export__.delivery_carrier_9_066799ca"

# 取消出库单 Tracking Reference 统一标记(供 ERP 里按此筛出、批量取消 dangling picking)
CANCEL_TRK = "订单取消"

MARK_PREFIXES = ("无货", "缺货", "勾选")
TRUTHY_STR = {"1", "x", "✓", "√", "是", "y", "yes", "true", "无货", "缺货"}


def _truthy(v):
    if isinstance(v, bool):
        return v
    if pd.isna(v):
        return False
    s = str(v).strip().lower()
    if s in TRUTHY_STR:
        return True
    try:                       # 数字: 非 0 即真(兼容 1/0)
        return float(s) != 0
    except ValueError:
        return False


NOGOODS_SHEET = "无货勾选"


def read_marked(paths):
    """读无货返回文件中被标记的行。多 sheet 时优先读『无货勾选』页；有标记列则只取标记为真的行。
    paths 支持多份做冗余(同店分多次导出)，合并。"""
    if isinstance(paths, str):
        paths = [paths]
    frames = []
    for path in paths:
        xl = pd.ExcelFile(path)
        sheet = NOGOODS_SHEET if NOGOODS_SHEET in xl.sheet_names else 0
        df = pd.read_excel(path, sheet_name=sheet)
        mark = next((c for c in df.columns
                     if str(c).strip().startswith(MARK_PREFIXES)), None)
        if mark is not None:
            df = df[df[mark].apply(_truthy)].copy()
        frames.append(df)
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


def load_nogoods(paths):
    """无货返回文件 → 无货连接键集合(SCP)。有『无货/缺货/勾选』标记列则只取标记为真的行，
    否则视全部行为无货。按值模式 `SCP\\d+` 识别单号(跳过标记列)，不依赖列名/sheet。"""
    if not paths:
        return set()
    df = read_marked(paths)
    keys = set()
    for c in df.columns:
        if str(c).strip().startswith(MARK_PREFIXES):
            continue
        for v in df[c].dropna().astype(str):
            m = SCP_RE.search(v)
            if m:
                keys.add(m.group(0)[-15:])
    return keys


def _row_key(row, cols, mark):
    """从一行里按值模式 SCP\\d+ 取连接键(跳过标记列)。找不到返回 None。"""
    for c in cols:
        if c == mark:
            continue
        m = SCP_RE.search(str(row[c]))
        if m:
            return m.group(0)[-15:]
    return None


# 订单级裁决优先级：未确认 > 无货 > 有货(越保守越优先，整单只要有一行更差就降级)
_RANK = {"未确认": 2, "无货": 1, "有货": 0}


def classify_return(paths):
    """读无货勾选返回表 → {连接键: '有货'/'无货'/'未确认'}（订单级裁决）。

    每行一个 SKU；标记列(『无货/缺货/勾选』开头)按值判定：1/真=无货, 0=有货, 空=未确认。
    订单级（多品订单）取最保守：任一行未确认→整单未确认；否则任一行无货→整单不发；
    全部有货(0)才整单发货。无标记列时整表视为无货(异常，仅兜底)。"""
    if isinstance(paths, str):
        paths = [paths]
    agg = {}
    for path in paths:
        xl = pd.ExcelFile(path)
        sheet = NOGOODS_SHEET if NOGOODS_SHEET in xl.sheet_names else 0
        df = pd.read_excel(path, sheet_name=sheet)
        mark = next((c for c in df.columns
                     if str(c).strip().startswith(MARK_PREFIXES)), None)
        cols = list(df.columns)
        for _, row in df.iterrows():
            key = _row_key(row, cols, mark)
            if key is None:
                continue
            if mark is None:
                st = "无货"
            else:
                v = row[mark]
                if pd.isna(v) or str(v).strip() == "":
                    st = "未确认"
                elif _truthy(v):
                    st = "无货"
                else:
                    st = "有货"
            if key not in agg or _RANK[st] > _RANK[agg[key]]:
                agg[key] = st
    return agg


# 有货清单里运单号列的表头线索(大小写不敏感，含即认)
TRACK_HEADER_HINTS = ("tracking", "运单", "waybill", "物流单号")


def _read_tables(path):
    """读一个表格文件 → DataFrame 列表。xlsx/xls 返回所有 sheet；csv 返回单表。
    csv 以 utf-8-sig 剥 BOM、全列按 str 读、空格保留为空串(不转 NaN)，与扫码端
    产出 `有货清单{店}.csv`(build_excel 同款 BOM) 对齐，SCP 扫格逻辑无需区分来源。"""
    if str(path).lower().endswith(".csv"):
        return [pd.read_csv(path, dtype=str, keep_default_na=False, encoding="utf-8-sig")]
    return list(pd.read_excel(path, sheet_name=None).values())


def load_shipped_map(paths):
    """读「有货(真实发货)订单清单」→ {连接键(SCP后15位): 运单号}。

    入口高度模糊：不要求特定结构/sheet/列名，xlsx 与扫码端产出的 csv 均可(见 _read_tables)。
    按**行**读——连接键由该行任一单元格的 `SCP\\d+` 提取(履约单号或 Order Reference 都行)；
    运单号取表头含 tracking/运单 的列(没有该列则运单留空)。
    同键多次出现保留首个非空运单；多份文件作冗余合并(同店分多次导出)。"""
    if not paths:
        return {}
    if isinstance(paths, str):
        paths = [paths]
    mp = {}
    for path in paths:
        for df in _read_tables(path):   # 所有 sheet(xlsx) 或单表(csv)
            track_col = next((c for c in df.columns
                              if any(h in str(c).lower() for h in TRACK_HEADER_HINTS)),
                             None)
            for _, row in df.iterrows():
                key = None
                for v in row.values:
                    m = SCP_RE.search(str(v))
                    if m:
                        key = m.group(0)[-15:]   # SCP 单号(后15位)，与 ERP _key 对齐
                        break
                if not key:
                    continue
                track = ""
                if track_col is not None:
                    tv = row[track_col]
                    if pd.notna(tv) and str(tv).strip():
                        track = str(tv).strip()
                if key not in mp or (not mp[key] and track):
                    mp[key] = track
    return mp


def load_shipped_keys(paths):
    """读有货清单 → 连接键集合(只要单号，忽略运单)。"""
    return set(load_shipped_map(paths))


def channel_of(order_ref):
    """VO_TOF_SCP... -> VO ; GW_TOF_SCP... -> GW"""
    return str(order_ref).split("_", 1)[0]


def chan_suffix(channels):
    """据发货订单渠道集合给文件名后缀：单店 VO / GW，混店 VO+GW，空则无后缀。"""
    present = set(channels)
    return "+".join(c for c in ("VO", "GW") if c in present)


def stage2_name(name, ch, n, d=None):
    """阶段二统一文件名：`YYYY年MM月DD日{渠道}{n}单 {产出名}.xlsx`，对齐阶段一 make_output_name。
    ch=渠道(VO/GW/VO+GW)，n=单数，d 默认运行当天。"""
    d = d or date.today()
    return f"{d.year}年{d.month:02d}月{d.day:02d}日{ch}{n}单 {name}.xlsx"


def get_shipped_orders(erp, shipped_keys, tracking_map=None):
    """有货(真实发货)单号 ∩ ERP → 订单级 DataFrame(Order Reference, VO Tracking No, _key, channel)。
    入口直接是真实发货订单号，结合 ERP 取明细，不再由无货倒推。
    运单号来源：优先用有货清单自带的(tracking_map)，缺则回落 ERP 的 VO Tracking No 列，
    再缺则留空——某单缺运单不阻断整表生成。"""
    orders = erp[[s4.ERP_ORDER_REF, "_key"]].drop_duplicates(subset="_key").copy()
    erp_track = None
    if s4.ERP_TRACKING in erp.columns:
        erp_track = (erp.drop_duplicates(subset="_key")
                     .set_index("_key")[s4.ERP_TRACKING])
    tracking_map = tracking_map or {}

    def track_for(k):
        t = tracking_map.get(k)
        if t is not None and str(t).strip():
            return str(t).strip()
        if erp_track is not None:
            v = erp_track.get(k, "")
            if pd.notna(v) and str(v).strip():
                return str(v).strip()
        return ""

    orders["VO Tracking No"] = orders["_key"].map(track_for)
    shipped = orders[orders["_key"].isin(shipped_keys)].copy()
    shipped["channel"] = shipped[s4.ERP_ORDER_REF].map(channel_of)
    return shipped


# ---------- B: 系统履约单号 ----------
def build_B(shipped, outdir, suffix=""):
    out = pd.DataFrame({"系统履约单号": shipped["_key"].tolist()})
    wb = Workbook(); ws = wb.active; ws.title = "Sheet1"
    be.write_df(ws, out)
    be.style_sheet(ws, 1)
    path = be.unique_path(os.path.join(outdir, stage2_name("系统履约单号", suffix, len(out))))
    wb.save(path)
    return path, len(out)


# ---------- C: 发货表 (GW/VO 分 sheet) ----------
def build_C(shipped, outdir, suffix=""):
    wb = Workbook(); first = True
    counts = {}
    for ch in ["GW", "VO"]:
        sub = shipped[shipped["channel"] == ch]
        counts[ch] = len(sub)
        df = sub[[s4.ERP_ORDER_REF, s4.ERP_TRACKING]].rename(
            columns={s4.ERP_ORDER_REF: "Order Reference", s4.ERP_TRACKING: "VO Tracking No"})
        ws = wb.active if first else wb.create_sheet()
        ws.title = ch
        first = False
        be.write_df(ws, df)
        be.style_sheet(ws, 2)
    path = be.unique_path(os.path.join(outdir, stage2_name("发货表", suffix, sum(counts.values()))))
    wb.save(path)
    return path, counts


# ---------- E: 出库单 (从 stock picking 全量导出过滤，按店铺拆 VO/GW) ----------
SCP_RE = re.compile(r"(SCP\d+)")
PICK_SRC_NAMES = ("Source Document", "Referenzbeleg")
PICK_TRK_NAMES = ("Tracking Reference", "Tracking-Referenz")


def _scp(v):
    m = SCP_RE.search(str(v))
    return m.group(1) if m else None


def build_picking_writeback(picking_paths, keys, trk_value, outdir, name,
                            carrier_id=None, split_channel=True):
    """从 stock picking 全量导出(出库原始数据)生成 picking 回写文件(出库单/取消出库单同一原语)。

    过滤: Source Document 的 SCP ∈ keys(经 _scp 归一)。沿用 pool 的英文表头与原 Status，
    把 Tracking Reference 统一覆盖成 trk_value；carrier_id 非空时同时把 Carrier/ID 覆盖成它。
    split_channel=True 按 VO/GW 拆两份；False 合并成一张(文件名渠道后缀取实际覆盖到的店)。
    返回 (results{label:(path,n)}, missing[scp...])：missing = keys 里在 pool 中找不到 picking 的 SCP。"""
    results, missing = {}, []
    if not picking_paths:
        return results, missing
    frames = [pd.read_excel(p) for p in picking_paths]
    pool = pd.concat(frames, ignore_index=True)
    srccol = next((c for c in pool.columns
                   if str(c).strip() in PICK_SRC_NAMES), None)
    if srccol is None:
        raise ValueError(f"出库原始数据缺少来源单据列(任一: {PICK_SRC_NAMES})")
    trkcol = next((c for c in pool.columns
                   if str(c).strip() in PICK_TRK_NAMES), None)

    pool["_scp"] = pool[srccol].map(_scp)
    pool["_ch"] = pool[srccol].astype(str).str.split("_", n=1).str[0]
    want = {s for s in (_scp(k) for k in keys) if s}

    kept = pool[pool["_scp"].isin(want)].copy()
    if trkcol is not None and trk_value is not None:
        kept[trkcol] = trk_value
    if carrier_id is not None:      # Carrier/ID 列不存在则新建
        carcol = next((c for c in kept.columns
                       if str(c).strip() == "Carrier/ID"), "Carrier/ID")
        kept[carcol] = carrier_id

    def _emit(sub, label):
        sub = sub.drop(columns=["_scp", "_ch"])
        sub = sub.where(pd.notna(sub), "")   # 空单元格写空串，避免出现字面 nan
        if sub.empty:
            return
        wb = Workbook(); ws = wb.active; ws.title = "Sheet1"
        be.write_df(ws, sub)
        be.style_sheet(ws, len(sub.columns))
        path = be.unique_path(os.path.join(outdir, stage2_name(name, label, len(sub))))
        wb.save(path)
        results[label] = (path, len(sub))

    if split_channel:
        for ch in ["VO", "GW"]:
            _emit(kept[kept["_ch"] == ch], ch)
    else:
        _emit(kept, chan_suffix(kept["_ch"]))

    missing = sorted(want - set(pool["_scp"].dropna()))
    return results, missing


def build_E(picking_paths, shipped, shipdate, outdir):
    """出库单：过滤实际发货订单的 picking，Tracking Reference 覆盖为发货日期、
    Carrier/ID 固定承运商，按 VO/GW 拆两份。"""
    return build_picking_writeback(picking_paths, shipped["_key"].tolist(),
                                   shipdate, outdir, "出库单",
                                   carrier_id=CARRIER_ID, split_channel=True)


def build_cancel(picking_paths, cancel_keys, outdir):
    """取消出库单：过滤取消订单的 picking，Tracking Reference 覆盖为『订单取消』、
    不写 Carrier/ID，合并成唯一一张(不按 VO/GW 拆)。"""
    return build_picking_writeback(picking_paths, cancel_keys, CANCEL_TRK, outdir,
                                   "取消出库单", carrier_id=None, split_channel=False)


# ---------- 缺货记录 (明细 + SKU 汇总) ----------
def _sku_lut(erp):
    """SKU -> 条码/货位/系统在售库存 查找表(SKU 级，从 ERP 导出取)。"""
    g = erp.groupby(s4.ERP_INTERNAL)
    lut = g.agg(
        Barcode=(s4.ERP_BARCODE, "first"),
        PickingName=(s4.ERP_PICKING, "first"),
        OnHand=(s4.ERP_ONHAND, "max"),
    )
    return lut


def _merge_same(ws, df, cols, start=2):
    """对已按某列排序的 df，把 cols 中各列相同连续值纵向合并(直观分组)。"""
    names = list(df.columns)
    idx = {c: names.index(c) + 1 for c in cols}
    keys = df[cols[0]].tolist()  # 以第一列(SKU)为分组依据
    i = 0
    while i < len(keys):
        j = i
        while j + 1 < len(keys) and keys[j + 1] == keys[i]:
            j += 1
        if j > i:
            for c in idx.values():
                ws.merge_cells(start_row=start + i, start_column=c,
                               end_row=start + j, end_column=c)
        i = j + 1


def build_shortage(marked, erp, mmdd, outdir):
    """从标记无货的行生成缺货记录：明细(按SKU合并) + SKU汇总。"""
    if marked.empty or "SKU" not in marked.columns:
        return None, 0
    lut = _sku_lut(erp)

    def info(sku, field, default=""):
        return lut.loc[sku, field] if sku in lut.index else default

    rows = []
    for _, r in marked.iterrows():
        sku = r["SKU"]
        rows.append({
            "系统履约单号":    r.get("系统履约单号", ""),
            "SKU":           sku,
            "商品名":         r.get("商品名", ""),
            "Barcode":       info(sku, "Barcode"),
            "Picking Name":  info(sku, "PickingName"),
            "缺货数量":       int(pd.to_numeric(r.get("数量", 1), errors="coerce") or 1),
            "系统在售库存":    info(sku, "OnHand", ""),
            "VO Delivery Type": r.get("VO Delivery Type", ""),
            "备注":           "",
        })
    detail = pd.DataFrame(rows).sort_values(
        ["SKU", "系统履约单号"]).reset_index(drop=True)
    detail.insert(0, "序号", range(1, len(detail) + 1))

    summary = (detail.groupby("SKU", sort=False)
               .agg(商品名=("商品名", "first"),
                    Barcode=("Barcode", "first"),
                    **{"Picking Name": ("Picking Name", "first")},
                    缺货订单数=("系统履约单号", "nunique"),
                    缺货总数量=("缺货数量", "sum"),
                    系统在售库存=("系统在售库存", "first"))
               .reset_index()
               .sort_values("缺货总数量", ascending=False))

    wb = Workbook()
    ws_d = wb.active
    ws_d.title = "明细"
    be.write_df(ws_d, detail)
    be.style_sheet(ws_d, len(detail.columns))
    # 合并相同 SKU 的 SKU 级列(直观看出哪个商品缺、缺几单)
    _merge_same(ws_d, detail, ["SKU", "商品名", "Barcode", "Picking Name", "系统在售库存"])

    ws_s = wb.create_sheet("SKU汇总")
    be.write_df(ws_s, summary)
    be.style_sheet(ws_s, len(summary.columns))

    path = be.unique_path(os.path.join(outdir, "缺货记录.xlsx"))
    wb.save(path)
    return path, len(detail), len(summary)


# ---------- 货代合并发货表 (跨店, 接受 N 份发货表) ----------
FORWARDER_PREFIX = "IHTCTGMBH+IH"   # 固定前缀，给货代核对用


def _find_ref_trk(df):
    """从一张发货表里定位 (Order Reference 列, Tracking 列, 运单列是否按列名命中)。
    Order Reference：列名匹配或值含 SCP；Tracking：列名含 tracking/运单，否则取另一列。
    第三个返回值 trk_by_name：True=按列名(tracking/运单)命中；False=靠"只剩一列"兜底或没找到
    → 调用方据此告警，避免把日期等非运单列静默当成运单（错位无声无息）。"""
    cols = list(df.columns)
    refcol = next((c for c in cols if str(c).strip() == "Order Reference"), None)
    if refcol is None:                        # 按值模式兜底：含 SCP 最多的列
        best, bestn = None, 0
        for c in cols:
            n = df[c].astype(str).str.contains("SCP").sum()
            if n > bestn:
                best, bestn = c, n
        refcol = best
    trkcol = next((c for c in cols if c != refcol and
                   ("tracking" in str(c).strip().lower() or "运单" in str(c))), None)
    trk_by_name = trkcol is not None
    if trkcol is None:                        # 只有两列时，另一列即运单（兜底，需告警）
        others = [c for c in cols if c != refcol]
        trkcol = others[0] if len(others) == 1 else None
    return refcol, trkcol, trk_by_name


def build_forwarder(paths, outdir, shipdate=None):
    """N 份发货表 → 货代清单 + 天猫回执 两份产出。按 Order Reference 去重；
    同单不同运单视为冲突报警。
    返回 (货代路径, 单数, 冲突列表[(ref,旧,新)], 告警列表[str], 回执路径, 回执单数)。
    - 货代清单：Order Reference + 运单号(给货代核对)。
    - 天猫回执：所有发货 Order Reference 后15位(系统履约单号)，各渠道合并去重(上传天猫做回执)。
    告警：某表未按列名(tracking/运单)识别到运单列时提示——防止把日期等非运单列静默当运单。"""
    os.makedirs(outdir, exist_ok=True)         # GUI 货代合并直调本函数，输出目录(如 输出/YYYYMMDD)可能尚不存在
    pairs, conflicts, warnings = {}, [], []
    for path in paths:
        fname_only = os.path.basename(path)
        for sheet, df in pd.read_excel(path, sheet_name=None).items():
            if df.empty:
                continue
            refcol, trkcol, trk_by_name = _find_ref_trk(df)
            if refcol is None:
                continue
            if not trk_by_name:
                col_desc = f"『{trkcol}』" if trkcol is not None else "（无）"
                warnings.append(
                    f"⚠ {fname_only}[{sheet}] 未识别到运单列(列名不含 tracking/运单)，"
                    f"当前兜底用了 {col_desc}，请核对该列是否运单号")
            for _, r in df.iterrows():
                ref = str(r[refcol]).strip()
                if "SCP" not in ref:
                    continue
                trk = "" if (trkcol is None or pd.isna(r[trkcol])) else str(r[trkcol]).strip()
                if ref in pairs:
                    if pairs[ref] and trk and pairs[ref] != trk:
                        conflicts.append((ref, pairs[ref], trk))
                    elif not pairs[ref] and trk:
                        pairs[ref] = trk
                else:
                    pairs[ref] = trk
    out = pd.DataFrame({"Order Reference": list(pairs.keys()),
                        "Tracking number": list(pairs.values())})
    n = len(out)
    d = shipdate or date.today().strftime("%Y%m%d")
    fname = f"{FORWARDER_PREFIX}{d}+{n}.xlsx"
    wb = Workbook(); ws = wb.active; ws.title = "Sheet2"
    be.write_df(ws, out)
    be.style_sheet(ws, 2)
    path = be.unique_path(os.path.join(outdir, fname))
    wb.save(path)

    # ---- 天猫回执：所有发货 Order Reference 后15位(系统履约单号)，各渠道合并去重 ----
    # SCP 号来自天猫后台、全局唯一，不同单不会撞后15位，故按后15位顺序去重即可(无需碰撞告警)
    receipt = {}                               # 后15位(系统履约单号) -> None，仅用于保序去重
    for ref in pairs:
        m = SCP_RE.search(ref)
        if m:
            receipt.setdefault(m.group(0)[-15:], None)   # 与 ERP _key/系统履约单号 同口径
    rct = pd.DataFrame({"系统履约单号": list(receipt)})
    rn = len(rct)
    d_obj = date(int(d[:4]), int(d[4:6]), int(d[6:8])) if len(d) == 8 and d.isdigit() else date.today()
    rname = stage2_name("天猫回执", "", rn, d_obj)
    wb_r = Workbook(); ws_r = wb_r.active; ws_r.title = "Sheet1"
    be.write_df(ws_r, rct)
    be.style_sheet(ws_r, 1)
    rpath = be.unique_path(os.path.join(outdir, rname))
    wb_r.save(rpath)

    return path, n, conflicts, warnings, rpath, rn


# ---------- D: 开账单上传表 ----------
TAG_RE = re.compile(r"^(账单|开发票)\d{4}")


find_id_col = s4.find_id_col   # 复用 step4_merge 的实现(External ID/ID/外部ID)


def has_id_col(df):
    return find_id_col(df) is not None


def build_billing(src, shipped_keys, mmdd, outdir, suffix=""):
    """从含 External ID 的来源(订单导出 或 账单模板导出)生成账单上传表。
    去重到订单级；Terms = 账单MMDD + 原值(剥离旧标签)；只留实际发货订单。
    保留来源的 ID 列原名(External ID / ID)，回传 Odoo 时按其匹配。"""
    df = src.copy()
    df = df.loc[:, ~df.columns.astype(str).str.startswith("Unnamed")]
    idc = find_id_col(df)
    # 去重到订单级(订单导出是逐行的；first 跳过空值取订单头行的 ID/Date/Terms)
    agg = df.groupby("Order Reference", sort=False).agg(**{
        "Order Date":           ("Order Date", "first"),
        idc:                    (idc, "first"),
        "Terms and conditions": ("Terms and conditions", "first"),
    }).reset_index()
    agg["_key"] = last15(agg["Order Reference"])
    if shipped_keys:
        agg = agg[agg["_key"].isin(shipped_keys)]
    raw = (agg["Terms and conditions"].astype(str)
           .str.replace(TAG_RE, "", regex=True)
           .str.replace(s4.WU_TAG, "", regex=False))   # 已补运单：剥掉「无运单」恢复原值
    agg["Terms and conditions"] = f"账单{mmdd}" + raw
    out = agg[["Order Date", idc, "Order Reference", "Terms and conditions"]]
    wb = Workbook(); ws = wb.active; ws.title = "Sheet1"
    be.write_df(ws, out)
    be.style_sheet(ws, len(out.columns))
    path = be.unique_path(os.path.join(outdir, stage2_name("账单上传", suffix, len(out))))
    wb.save(path)
    return path, len(out)


def run(mmdd, erp_paths, shipped_files=None, nogoods_files=None,
        full_tmall_path=None, billing=None, outdir="output", picking=None, shipdate=None,
        cancel_list=None):
    """第二阶段核心：结合 ERP 生成 B(系统履约单号)/C(发货表)/D(账单)/E(出库) 四个产出。
    两种入口二选一(都给则优先有货)：

    - **有货清单**(shipped_files，白名单)：真实发货的履约单号/Order Reference，直接 ∩ ERP 取明细。
    - **无货勾选**(nogoods_files，黑名单)：仓库返回的无货单，直接取有货(0)作发货集合。
    full_tmall_path(完整天猫导出)：仅无货入口的交叉核对用(可选)。缺货记录已移出本阶段。
    返回结果文字行列表。供 CLI 与 GUI 共用。"""
    os.makedirs(outdir, exist_ok=True)
    log = []
    erp_df = be._load_erps(erp_paths)         # 只读一次，发货/账单/出库复用
    erp_keys = set(erp_df["_key"])
    # 拣货候选(stage1 发货集合)：无货入口交叉核对用。由完整天猫导出二段式推出。
    cand_keys = None
    if full_tmall_path:
        full = s4.load_full_tmall(full_tmall_path)
        ann = s4.classify4(erp_df, s4.done_keys_from_full(full),
                           s4.cancel_keys_from_full(full))
        cand_keys = set(ann[ann["_ship"]]["_key"]) & erp_keys

    # ---- 取消出库单：独立于发货输入，只要 取消订单清单 + 出库原始数据 就产出(可单独补跑) ----
    if cancel_list:
        if picking:
            ck = load_shipped_keys(cancel_list)
            resX, missX = build_cancel(picking, ck, outdir)
            if resX:
                for _label, (p, n) in resX.items():
                    log.append(f"取消出库单 已生成: {p}  ({n} 行，Tracking Reference={CANCEL_TRK})")
            else:
                log.append("取消出库单 跳过 (出库原始数据无匹配的取消订单 picking)")
            if missX:
                log.append(f"⚠ 取消出库单: {len(missX)} 个取消订单在出库原始数据里找不到 picking: "
                           f"{', '.join(missX[:10])}{' …' if len(missX) > 10 else ''}")
        else:
            log.append("取消出库单 跳过 (传了取消订单清单但未传出库原始数据)")

    tracking_map = {}                          # 有货清单自带的运单号(发货表优先用)
    if shipped_files:                          # —— 有货入口(白名单) ——
        tracking_map = load_shipped_map(shipped_files)
        ship_in = set(tracking_map)
        shipped_keys = ship_in & erp_keys
        log.append(f"[有货入口] 输入 {len(ship_in)} 单号，命中 ERP {len(shipped_keys)} 单")
        miss_erp = ship_in - erp_keys
        if miss_erp:
            log.append(f"⚠ {len(miss_erp)} 个有货单号在 ERP 里找不到(未结合): "
                       f"{', '.join(list(miss_erp)[:10])}{' …' if len(miss_erp) > 10 else ''}")
    elif nogoods_files:                        # —— 无货入口(黑名单) ——
        # 直接取『有货(0)』作发货，而非候选−无货：漏返回的单不会默认全发，更安全。
        ret = classify_return(nogoods_files)
        have = {k for k, v in ret.items() if v == "有货"} & erp_keys
        ng_marked = {k for k, v in ret.items() if v == "无货"}
        blank = {k for k, v in ret.items() if v == "未确认"}
        shipped_keys = have
        log.append(f"[无货入口] 返回 {len(ret)} 单(命中ERP计)：有货 {len(have)} / "
                   f"无货 {len(ng_marked)} / 未确认 {len(blank)} → 发货 {len(have)} 单")
        if blank:
            bl = sorted(blank)
            log.append(f"⚠ {len(blank)} 单未确认有无货(标记留空)，已不发货，请核对: "
                       f"{', '.join(bl[:10])}{' …' if len(bl) > 10 else ''}")
        if cand_keys is not None:            # 有候选则交叉核对(发了候选外的单很可疑)
            extra = have - cand_keys
            if extra:
                ex = sorted(extra)
                log.append(f"⚠ {len(extra)} 单标有货但不在拣货候选(名单/取消可能不一致): "
                           f"{', '.join(ex[:10])}{' …' if len(ex) > 10 else ''}")
    else:
        if cancel_list:            # 仅取消模式：已(尝试)产出取消出库单，无发货输入不算错
            return log
        log.append("✗ 未提供有货清单(shipped)或无货返回文件(nogoods)，无法确定发货集合")
        return log

    shipped = get_shipped_orders(erp_df, shipped_keys, tracking_map)
    shipped_keys = set(shipped["_key"])        # 规整(去重到订单级后的实际键)
    log.append(f"实际发货订单: {len(shipped)} 单 "
               f"(GW {sum(shipped['channel']=='GW')} / VO {sum(shipped['channel']=='VO')})")
    no_track = int((shipped["VO Tracking No"].astype(str).str.strip() == "").sum())
    if no_track:
        log.append(f"⚠ {no_track} 单缺运单号(有货清单与ERP都没有)，发货表该单运单留空")
    suffix = chan_suffix(shipped["channel"])   # 文件名渠道后缀：VO / GW / VO+GW

    # 各产出彼此独立：任一失败只记错并跳过，不阻断其他产出。
    def _step(fn):
        try:
            fn()
        except Exception as ex:
            log.append(f"❌ 该产出生成失败(已跳过，不影响其他产出): {ex}")

    def _b():
        pB, nB = build_B(shipped, outdir, suffix)
        log.append(f"系统履约单号 已生成: {pB}  ({nB} 个)")
    _step(_b)

    def _c():
        # 运单号是发货表的核心；整批一个都没有 → 跳过发货表(账单/出库不受影响)
        if len(shipped) and (shipped["VO Tracking No"].astype(str).str.strip() == "").all():
            log.append("发货表 跳过 (发货订单都没有运单号；有货清单和 ERP 都未提供，账单/出库照常生成)")
            return
        pC, cC = build_C(shipped, outdir, suffix)
        log.append(f"发货表 已生成: {pC}  (GW {cC['GW']} / VO {cC['VO']})")
    _step(_c)

    # E 出库单：从 stock picking 全量导出过滤出实际发货订单
    def _e():
        if picking:
            sd = shipdate or date.today().strftime("%Y%m%d")
            resE, missE = build_E(picking, shipped, sd, outdir)
            if resE:
                for ch, (p, n) in resE.items():
                    log.append(f"出库单{ch} 已生成: {p}  ({n} 行，发货日期 {sd})")
            else:
                log.append("出库单 跳过 (出库原始数据无匹配的发货订单)")
            if missE:
                log.append(f"⚠ 出库单: {len(missE)} 个发货订单在出库原始数据里找不到 picking: "
                           f"{', '.join(missE[:10])}{' …' if len(missE) > 10 else ''}")
        else:
            log.append("出库单 跳过 (未传出库原始数据 --picking)")
    _step(_e)

    # 账单上传：优先用含 ID 的订单导出直接生成；否则用单独的账单模板导出
    def _d():
        if has_id_col(erp_df):
            pD, nD = build_billing(erp_df, shipped_keys, mmdd, outdir, suffix)
            log.append(f"账单上传 已生成(来自订单导出含ID): {pD}  ({nD} 行，标签 账单{mmdd})")
        elif billing:
            pD, nD = build_billing(pd.read_excel(billing), shipped_keys, mmdd, outdir, suffix)
            log.append(f"账单上传 已生成(来自账单模板导出): {pD}  ({nD} 行，标签 账单{mmdd})")
        else:
            log.append("账单上传 跳过 (订单导出无ID列且未传账单模板；在 Odoo 订单导出模板勾上 External ID 列即可自动生成)")
    _step(_d)

    # 注：缺货记录已从阶段二移除。未来单独成一个阶段(无货清单 × 库存ERP 筛查)，
    # 届时复用 build_shortage()/read_marked()(暂保留)。
    return log


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--mmdd", default=date.today().strftime("%m%d"),
                    help="标签日期 MMDD，默认今天")
    ap.add_argument("--erp", nargs="+",
                    help="ERP 导出，可多份(VO/GW 各一份)")
    ap.add_argument("--forwarder", nargs="*", default=None,
                    help="货代合并模式：N 份发货表 → 一张跨店合并发货清单(其余参数忽略)")
    ap.add_argument("--shipped", nargs="*", default=None,
                    help="有货入口：真实发货订单清单(含履约单号/Order Reference 即可)，可多份冗余")
    ap.add_argument("--nogoods", nargs="*", default=None,
                    help="无货入口：仓库返回的无货勾选文件，可多份冗余(需配 --tmall-full 算拣货候选)")
    ap.add_argument("--tmall-full", dest="full", default=None,
                    help="完整天猫导出(唯一天猫输入)：二段式推发货范围，无货入口算拣货候选必需")
    ap.add_argument("--billing", default=None)
    ap.add_argument("--picking", nargs="*", default=None,
                    help="出库原始数据(stock picking 全量导出)，可传多个(VO/GW 各一份)")
    ap.add_argument("--cancel-list", dest="cancel_list", nargs="*", default=None,
                    help="取消订单清单(阶段一产出+人工补录后到的取消单)，配 --picking 生成『取消出库单』")
    ap.add_argument("--shipdate", default=None, help="发货日期 YYYYMMDD，默认今天")
    ap.add_argument("--outdir", default="output")
    a = ap.parse_args()
    os.makedirs(a.outdir, exist_ok=True)
    if a.forwarder is not None:                # 货代合并模式(独立步骤)
        p, n, conf, warns, rp, rn = build_forwarder(a.forwarder, a.outdir, a.shipdate)
        print(f"货代合并发货表 已生成: {p}  ({n} 单)")
        print(f"天猫回执(系统履约单号) 已生成: {rp}  ({rn} 单)")
        for w in warns:
            print(w)
        for ref, old, new in conf:
            print(f"⚠ 运单冲突 {ref}: {old} vs {new}(已保留先出现的)")
        return
    if not a.erp:
        ap.error("非货代合并模式需要 --erp")
    for line in run(a.mmdd, a.erp, a.shipped, a.nogoods, a.full, a.billing,
                    a.outdir, a.picking, a.shipdate, a.cancel_list):
        print(line)


if __name__ == "__main__":
    main()
