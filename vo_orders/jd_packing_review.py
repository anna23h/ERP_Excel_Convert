#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""京东「装箱复核历史查询」瘦身表。

京东后台导出的复核历史 45 列，仓库现场只看 5 列 + 一个序号。本模块把原始导出
改写成打印/核对用的瘦表，**不新建工作簿**，而是在原文件副本上删列 + 插列，
这样边框、字体、条码列的文本格式全部原样保留(条码带前导 `-`、商品编号 12 位，
一旦重写成新表极易被 Excel 当数字吃掉精度)。**只有列名行的底色是主动去掉的**：
京东给的是深色实心底，整行打出来费墨粉又没多少信息量，加粗 + 边框已经够把表头
认出来了(同 PRINT_HL 换色那次的取舍：打印件上的颜色要按灰度亮度算，不是看屏幕)。

输出列：序号 | 运单号（随单）| 商品编号 | 商品条码 | 商品名称 | 复核数量 | 备注

序号规则(与用户手改的示例一致)：一个运单 = 一个包裹 = 一个序号。一单多品时
**序号列与运单号列纵向合并单元格**，一个包裹在纸上就是一个格子，现场照着数包裹
不会把第二行当成另一单(合并前只是留空，靠视觉挂在上一行下面，打印出来容易看串)。

**按连续段分组，不是按值去重**：只有紧挨着的同运单行才算一单多品。京东导出里同一
运单的行本来就是连着的(实测 6 份导出、302 行，两处一单多品都相邻)；万一哪天不相邻，
按值去重会把远处那行的运单号也抹掉、挂到毫不相干的上一行下面，而按连续段分组
只是各自成一单，最坏情况是多一个序号，不会串行。

打印形态(仓库现场是要打出来照着核对的)：**横向 + 所有列压到一页宽**
(`fitToWidth=1` / `fitToHeight=0`，行数照常翻页)、页脚居中「第 X 页，共 Y 页」、
表头行每页重复。表头不重复的话第 2 页起就是一堆没有列名的号码。

复核数量原始存的是 1.0/2.0 这种浮点，整数的写回 int，避免打印出 "1.0"。

条码/品名两列开自动换行，行高留给 Excel 自适应(不写 customHeight)，保证打印
不截字；条码里的多码是逗号连写的(`-16233255,100380697148`)，按逗号拆成一码一行。

备注列是留白给现场手写/回填的，只有表头和边框。
"""
import os
import re
import sys
from copy import copy

from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))  # 让 common/ 可导入
from common.xlsx import apply_print  # noqa: E402

# 原始导出里要保留的列(按表头名认，京东改列序也不怕)，顺序即输出顺序
KEEP_HEADERS = ["运单号（随单）", "商品编号", "商品条码", "商品名称", "复核数量"]
SEQ_HEADER = "序号"
NOTE_HEADER = "备注"
WAYBILL_HEADER = "运单号（随单）"
QTY_HEADER = "复核数量"

# 输出表的列号(插完 A 列之后)：序号 A、运单号 B——一单多品时这两列纵向合并
SEQ_COL, WAYBILL_COL = 1, 2

#: 页脚(Excel 字段码：&P 当前页，&N 总页数)。同 common.xlsx.apply_print 的默认值，
#: 这里显式写出来是因为现场拿到的是散页，页码是唯一能确认没漏页的东西。
FOOTER = "第 &P 页，共 &N 页"

# 列宽照示例定(商品名称最宽，其余够放 15~16 位单号)
COL_WIDTHS = {"A": 10.13, "B": 16.0, "C": 16.0, "D": 16.0, "E": 65.5,
              "F": 13.2, "G": 16.0}

CENTER = Alignment(horizontal="center", vertical="center")
WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
NO_FILL = PatternFill(fill_type=None)   # 列名行去底色用


def _waybill_runs(ws, col=WAYBILL_COL):
    """运单号列 → [(首行, 末行)]，**连续同号算一段**(空号自成一段)。

    刻意不按值去重(见模块 docstring)：不相邻的同号只会各自成一单、多一个序号，
    而按值去重会把远处那行的运单号抹掉、让它挂到毫不相干的上一行下面。
    """
    runs, prev = [], None
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, col).value
        v = str(v).strip() if v is not None else ""
        if runs and v and v == prev:
            runs[-1][1] = r
        else:
            runs.append([r, r])
        prev = v
    return [tuple(x) for x in runs]


def _out_path(src):
    """`xxx_0(1).xlsx` → `xxx_0new(1).xlsx`；没有 (n) 后缀就直接接 new。"""
    d, base = os.path.split(src)
    stem, ext = os.path.splitext(base)
    m = re.search(r"\(\d+\)$", stem)
    if m:
        stem = stem[:m.start()] + "new" + m.group(0)
    else:
        stem += "new"
    return os.path.join(d, stem + ext)


def convert(src, dst=None, outdir=None):
    """把一份原始导出改写成瘦表，返回输出路径。

    outdir: 只改产出目录、文件名照旧(`xxx_0new(1).xlsx`)。命令行不传 = 就地放在
    原文件旁边(操作员一次下载一批，转完在同一个文件夹里对着看最省事)；GUI 传共用
    输出目录，否则「打开输出文件夹」按钮会指到一个根本没有产出的地方。
    """
    dst = dst or _out_path(src)
    if outdir:
        os.makedirs(outdir, exist_ok=True)
        dst = os.path.join(outdir, os.path.basename(dst))
    wb = load_workbook(src)
    ws = wb.worksheets[0]

    headers = {}
    for c in ws[1]:
        if c.value is not None and c.value not in headers:
            headers[str(c.value).strip()] = c.column
    missing = [h for h in KEEP_HEADERS if h not in headers]
    if missing:
        raise ValueError("原始表缺少列：%s（表头对不上，先确认导出模板）" % "、".join(missing))

    keep_idx = [headers[h] for h in KEEP_HEADERS]
    if keep_idx != sorted(keep_idx):
        raise ValueError("原始表列序与预期不符，需先调整 KEEP_HEADERS 逻辑")

    # 从右往左删，避免删一列后面的列号整体前移
    for col in range(ws.max_column, 0, -1):
        if col not in keep_idx:
            ws.delete_cols(col)
    ws.insert_cols(1)  # 腾出 A 列放序号

    # A1 表头沿用其余表头的样式(底色/加粗/居中/边框)
    ws["A1"].value = SEQ_HEADER
    ws["A1"]._style = copy(ws["B1"]._style)
    ws["G1"].value = NOTE_HEADER
    ws["G1"]._style = copy(ws["B1"]._style)
    # 列名行去底色(加粗/字号/边框留着)。必须在上面两行**之后**做：A1/G1 的样式是
    # 从 B1 拷来的，先清 B1 也会被拷回来的旧底色盖掉。
    for cell in ws[1]:
        cell.fill = NO_FILL

    # 序号与运单号：一个连续段(= 一个包裹)一个序号，续行清空后与首行合并
    for seq, (first, last) in enumerate(_waybill_runs(ws), start=1):
        ws.cell(first, 1).value = seq
        for r in range(first + 1, last + 1):
            ws.cell(r, 1).value = None
            ws.cell(r, 2).value = None  # 合并前必须清空：openpyxl 合并只保留左上角
        if last > first:
            for col in (SEQ_COL, WAYBILL_COL):
                letter = get_column_letter(col)
                ws.merge_cells(f"{letter}{first}:{letter}{last}")

    for r in range(2, ws.max_row + 1):
        ws.cell(r, 1)._style = copy(ws.cell(r, 6)._style)
        ws.cell(r, 1).alignment = CENTER
        ws.cell(r, 1).number_format = "General"
        # 运单号/商品编号也垂直居中：合并后的运单号不居中会贴在格子顶上，而条码、
        # 品名两列换行后本就是居中的，只剩商品编号沉底会让一行看着错位。
        for col in (2, 3):
            ws.cell(r, col).alignment = CENTER

        qty = ws.cell(r, 6).value
        if isinstance(qty, float) and qty.is_integer():
            ws.cell(r, 6).value = int(qty)
        ws.cell(r, 6).alignment = CENTER
        ws.cell(r, 6).number_format = "0"

        # 条码多码逗号连写 → 一码一行，配合自动换行打印不截
        code = ws.cell(r, 4).value
        if isinstance(code, str) and "," in code:
            ws.cell(r, 4).value = "\n".join(
                x.strip() for x in code.split(",") if x.strip())
        for col in (4, 5):  # 商品条码 / 商品名称：自动换行
            ws.cell(r, col).alignment = WRAP

        note = ws.cell(r, 7)  # 备注：留白，只给边框
        note._style = copy(ws.cell(r, 6)._style)
        note.alignment = WRAP
        note.number_format = "General"

    for letter, width in COL_WIDTHS.items():
        ws.column_dimensions[letter].width = width
    for letter in [k for k in list(ws.column_dimensions) if k not in COL_WIDTHS]:
        del ws.column_dimensions[letter]

    # 打印形态：A4 横向 + 窄边距 + 7 列压到一页宽(行数照常翻页)
    #          + 页脚页码 + 表头每页重复
    apply_print(ws, landscape=True, fit_width=True, footer=FOOTER, narrow=True)
    ws.print_title_rows = "1:1"

    wb.save(dst)
    return dst


def is_generated(path):
    """是不是本脚本自己产出的 new 版(或 Excel 的 ~$ 临时文件)。

    GUI 也用它：操作员在文件对话框里全选，很容易把上一轮的 new 版一起选进来，
    而 new 版已经只剩 7 列，再转一次会撞「原始表缺少列」直接报错。
    """
    name = os.path.basename(path)
    if not name.lower().endswith(".xlsx") or name.startswith("~$"):
        return True
    return bool(re.search(r"new(\(\d+\))?$", os.path.splitext(name)[0]))


def convert_dir(folder, outdir=None):
    """转换目录下所有原始导出(跳过已生成的 new 版和 Excel 临时文件)。"""
    return [convert(os.path.join(folder, name), outdir=outdir)
            for name in sorted(os.listdir(folder))
            if not is_generated(os.path.join(folder, name))]


if __name__ == "__main__":
    args = sys.argv[1:]
    if not args:
        print("用法：python jd_packing_review.py <xlsx 或 目录> ...")
        raise SystemExit(1)
    for a in args:
        if os.path.isdir(a):
            for p in convert_dir(a):
                print("生成:", p)
        else:
            print("生成:", convert(a))
