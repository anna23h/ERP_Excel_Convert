"""Excel 排版 / 写盘的公共部件（原先长在 vo_orders/build_excel.py 里）。

搬出来的理由：reorder_helper 一直在掏 build_excel 的下划线私有函数用，
packing_list 干脆把样式代码复制了一份。两条新流水线会把这事再重复两遍。

**刻意不 import pandas**：write_df / write_simple 只用到 .columns 和 .iterrows()，
没碰任何 pandas API。这样不产 DataFrame 的流水线（packing_list 等）不必背 pandas
依赖，将来单独打包也不会把 pandas 塞进 exe。
"""
import os
import re

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

YELLOW = PatternFill("solid", fgColor="FFFF00")
# 打印件上的标记色。灰度亮度(0.299R+0.587G+0.114B)决定黑白激光印出来的深浅与耗粉：
#   FFFF00 黄 = 226/255(89% 白) —— 几乎看不见，仓库反馈「只有轻微灰度」；
#   ED7D31 橙 = 150/255(59%)    —— 看得清但**过深、费墨粉**（2026-09-01 实测反馈）；
#   FFC000 琥珀金 = 189/255(74%) —— 取中：比黄明显得多，又比橙省粉，黑字对比度 11:1。
# 凡是要打印给人照着做的标记都用它。改色前先算灰度亮度，别只看屏幕。
PRINT_HL = PatternFill("solid", fgColor="FFC000")
THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_BOTTOM = Alignment(horizontal="left", vertical="bottom", wrap_text=True)  # 左对齐+下沉
LEFT_CENTER = Alignment(horizontal="left", vertical="center", wrap_text=True)  # 合并单元格用
FONT = Font(size=15)
SMALL_FONT = Font(size=13)  # 比正文小 2 号
HEAD_FONT = Font(size=15, bold=True)
ROW_H = 35


#: 数字列里表示「无此项」的占位符。判定对齐时跳过，否则一个 — 会把整列数字拽成左对齐。
NUM_PLACEHOLDERS = {"—", "-", "－", ""}


#: 写成字符串的日期（本仓多处是 `date_order[:10]` 切出来的 str，不是 date 对象）。
DATE_RE = re.compile(r"^\d{4}[-/]\d{1,2}[-/]\d{1,2}$")


def _is_numeric(v):
    """这个值算不算「纯数字」。日期算（2026-08-30 用户选定居中），公式串也算
    （本仓的公式产出的全是数字）。

    日期必须连**字符串形态**一起认：`2026-08-11` 这种是 `date_order[:10]` 切出来的 str，
    只判 `strftime` 会把整列日期判成文字、左对齐（第一版就是这么错的）。"""
    if isinstance(v, bool):
        return False
    if isinstance(v, (int, float)):
        return True
    if hasattr(v, "strftime"):
        return True
    s = str(v).strip()
    if s.startswith("="):
        return True
    if DATE_RE.match(s):
        return True
    try:                       # 「1000」这类写成文本的数字
        float(s)
        return True
    except ValueError:
        return False


def _disp_len(s):
    """按显示宽度估长度：CJK/全角算 1 格，ASCII 算 0.65 格。

    外面统一乘 1.5（为 15 号字与中文留的系数），对全中文的列刚好，但产品代码这种纯 ASCII
    会富余一半——`Mucosolvan_02807988` 19 字符要到 30.5 宽。这里先把 ASCII 折价，
    乘完才贴合。只在 `auto_align=True` 时启用，其他流水线的列宽一格不变。"""
    n = 0.0
    for ch in s:
        n += 1.0 if (ord(ch) > 0x2E80 or ch in "，。（）：、—") else 0.65
    return n


def _auto_left_idx(ws, n_cols, header_row):
    """逐列按**实际数据**判定该不该左对齐 → 列号集合。

    靠手写列名清单来指定左对齐，正是 header_row 写死第 1 行那个 bug 的温床
    （列名对不上就静默失效，整表居中还没人发现）。改成看值：整列非空值全是数字/日期/公式
    → 居中；掺了任何文字（产品代码、「历史」、供应商名）→ 左对齐。整列全空 → 居中
    （待填的报价/数量列按数字待遇）。"""
    left = set()
    for c in range(1, n_cols + 1):
        vals = [ws.cell(r, c).value for r in range(header_row + 1, ws.max_row + 1)]
        vals = [v for v in vals
                if v is not None and str(v).strip() not in NUM_PLACEHOLDERS]
        if vals and not all(_is_numeric(v) for v in vals):
            left.add(c)
    return left


def style_sheet(ws, n_cols, header_font=HEAD_FONT, left_cols=(), small_cols=(), widths=None,
                header_row=1, auto_align=False):
    """left_cols: 内容左对齐+下沉的列名集合；small_cols: 字号小2号的列名集合。表头始终居中。
    widths: {表头名: 列宽} 固定列宽表，命中的列用固定宽度（操作员手工调好、免二次拖列），
    未命中的列仍按内容自动算宽。

    header_row: 表头在第几行。上方若有标题/说明行必须传，否则列名匹配与列宽全部错位
    （2026-08-30 之前写死第 1 行，`procure/` 那几张带标题行的表上 left_cols 一直静默失效）。
    auto_align: 逐列按实际数据自动判定左对齐/居中，见 `_auto_left_idx`。left_cols 仍生效，
    作为**强制左对齐**的覆盖口（用于必然是文字、但当前整列还空着的列，如备注）。

    两个新参数**默认值即旧行为**——本函数被六七条流水线共用，不能顺手改了它们的产出样式。"""
    headers = [c.value for c in ws[header_row]]
    left_idx = {i + 1 for i, h in enumerate(headers) if h in left_cols}
    if auto_align:
        left_idx |= _auto_left_idx(ws, n_cols, header_row)
    small_idx = {i + 1 for i, h in enumerate(headers) if h in small_cols}
    for row in ws.iter_rows():
        for cell in row:
            if cell.row < header_row:      # 标题/说明行：左对齐、不描边（60 列时满行框线尤其难看）
                cell.alignment = LEFT_BOTTOM
                continue
            cell.border = BORDER
            if cell.row == header_row:
                cell.alignment = CENTER
                cell.font = header_font
            else:
                cell.alignment = LEFT_BOTTOM if cell.column in left_idx else CENTER
                cell.font = SMALL_FONT if cell.column in small_idx else FONT
    for r in range(1, ws.max_row + 1):
        # 含单元格内换行(\n)的行按行数放大，否则固定高度会裁掉第二行起的内容；
        # 无换行的行恒为 ROW_H(其他表无 \n 内容，行为不变)
        lines = max((str(c.value).count("\n") + 1 for c in ws[r] if c.value is not None),
                    default=1)
        ws.row_dimensions[r].height = ROW_H if lines == 1 else lines * 22
    # 列宽：固定表命中的列用固定宽度（操作员调好的成品宽），其余按内容自动算宽
    # （字号15 比默认大，需放大系数，否则日期显示为 ######）
    widths = widths or {}
    for c in range(1, n_cols + 1):
        hdr = headers[c - 1] if c - 1 < len(headers) else None
        if hdr in widths:
            ws.column_dimensions[get_column_letter(c)].width = widths[hdr]
            continue
        maxlen = 0
        # 从表头行起算：**标题/说明行不能参与算宽**。那几行长文案都写在 A1/A2，
        # 算进去会把 A 列撑到 140（2026-08-30 用户实测报出，产品代码列宽得离谱）。
        for r in range(header_row, ws.max_row + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            s = v.strftime("%Y-%m-%d %H:%M:%S") if hasattr(v, "strftime") else str(v)
            maxlen = max(maxlen, _disp_len(s) if auto_align else len(s))
        ws.column_dimensions[get_column_letter(c)].width = max(12, maxlen * 1.5 + 2)


def _cell(v):
    """把各种「空」统一成 None——openpyxl 只认 None，见到 pandas 的 NA/NaT 会直接抛
    `ValueError: Cannot convert <NA> to Excel`（用了 Int64 等 nullable dtype 就会撞上）。

    这里刻意**不 import pandas**（本模块的既定约束），改用类型名判定；
    float 的 NaN 用 `v != v` 认，它对 pd.NA 不适用（pd.NA != pd.NA 得到 NA，不是 True）。
    """
    if v is None or type(v).__name__ in ("NAType", "NaTType"):
        return None
    if isinstance(v, float) and v != v:          # NaN
        return None
    return v


def write_df(ws, df):
    ws.append(list(df.columns))
    for _, row in df.iterrows():
        ws.append([_cell(v) for v in row])


def unique_path(path):
    """目标已存在则在扩展名前加序号 (1)/(2)…，避免覆盖既有产出。"""
    if not os.path.exists(path):
        return path
    base, ext = os.path.splitext(path)
    i = 1
    while os.path.exists(f"{base}({i}){ext}"):
        i += 1
    return f"{base}({i}){ext}"


def write_simple(out, outdir, fname, n_cols=None, left_cols=(), small_cols=(), widths=None):
    """把一张 DataFrame 写成单 sheet workbook(统一样式)。返回 (路径, 行数)。
    left_cols/small_cols/widths 透传给 style_sheet（左对齐下沉/小字号/固定列宽），默认空=全居中自动宽。"""
    wb = Workbook(); ws = wb.active; ws.title = "Sheet1"
    write_df(ws, out)
    style_sheet(ws, n_cols or len(out.columns), left_cols=left_cols, small_cols=small_cols,
                widths=widths)
    path = unique_path(os.path.join(outdir, fname))
    wb.save(path)
    return path, len(out)


#: Excel 内置「窄」边距(英寸)：上下 0.75、左右 0.25。左右让出来的 1.1 英寸
#: 够多塞一列，列多的表(如京东装箱复核 7 列)压一页宽时缩放比例明显更松。
NARROW_MARGINS = {"top": 0.75, "bottom": 0.75, "left": 0.25, "right": 0.25}

#: 本仓一直在用的边距，`narrow=False` 时的取值。**不要顺手改成窄边距**——
#: apply_print 被拣货表/面单共用，改默认值等于同时改掉两份天天在打的成品。
DEFAULT_MARGINS = {"top": 0.9, "bottom": 0.9, "left": 0.8, "right": 0.8}


def apply_print(ws, landscape=False, fit_width=False, footer="第 &P 页，共 &N 页",
                top=None, bottom=None, left=None, right=None, narrow=False):
    """步骤9 打印设置。footer 用 Excel 字段码：&P=当前页码，&N=总页数。
    默认『第 &P 页，共 &N 页』(第1页，共3页...)。页边距单位为英寸。

    纸张**一律 A4**：不写死的话走打印机默认，同一份文件在设成 Letter 的机器上
    印出来右边会被裁掉一截，且事先看不出来。本仓所有成品都是 A4 打的。

    narrow: 用 Excel 的「窄」边距预设(见 NARROW_MARGINS)。**默认关**——
    top/bottom/left/right 四个形参显式传值时仍然优先，所以老调用方一格不变。"""
    if landscape:
        ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    if fit_width:  # 所有列压到一页宽
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    m = NARROW_MARGINS if narrow else DEFAULT_MARGINS
    ws.page_margins = PageMargins(
        top=m["top"] if top is None else top,
        bottom=m["bottom"] if bottom is None else bottom,
        left=m["left"] if left is None else left,
        right=m["right"] if right is None else right,
        header=0.3, footer=0.3)
    ws.oddFooter.center.text = footer
