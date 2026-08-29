"""SO -> Packing List 半成品。

吃 ERP 导出的 sale.order 行明细，产出照抄成品箱单格式的 Excel：
机器可知的列（品名/SKU/条码/HS/原产国）填好，仓库现场才知道的列
（托盘号/批次号/箱号/箱规/箱数/保质期/毛重/尺寸/体积重）留空给手填。

Quantity total 不预填订购量，写成 =SUMPRODUCT(箱规*箱数)，仓库填完自动出数。

用法:
    python packing_list.py "sale.order.xlsx" [-o 输出目录] [--spare 2]
"""

import argparse
import datetime as dt
import os
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))  # 让 common/ 可导入
from common.xlsx import BORDER, YELLOW as WARN_FILL, unique_path  # noqa: E402

# ---- 成品箱单的固定排版 ----
COMPANY = "IHTCT GmbH \nHansaallee. 189, 40549 Düsseldorf \n"
TITLE = "   Packing List"
HEADERS = [
    "Item", "Products Name", "SKU 系统产品号", "Bar Code", "HS Code",
    "批次号", "箱号", "箱规", "箱数", "保质期", "Origin\n原产国",
    "Quantity total", "Gross Weight (kg)\n毛重",
    "Paket Measurements\n(L)*(W)*(D)cm\n包裹尺寸", " Size weight/体积重",
    "核对\n箱规×箱数",   # P 列：不在成品箱单的 15 列里，是对拍用的辅助列
]
COL_WIDTHS = {"B": 36.375, "C": 22.625, "D": 18.875, "F": 18.625,
              "G": 9.0, "J": 9.0, "N": 23.625, "P": 11.0}
ORIGIN = "DE"
DEFAULT_HS = "30049000"  # ERP 主数据常缺；成品箱单里全是这个
HEADER_ROW = 6
FIRST_DATA_ROW = 7

# 脚本填的列 / 仓库手填的列（1-based）
COL_ITEM, COL_NAME, COL_SKU, COL_BARCODE, COL_HS = 1, 2, 3, 4, 5
COL_BOXSPEC, COL_BOXCOUNT = 8, 9
COL_ORIGIN, COL_QTY = 11, 12
COL_CHECK = 16  # P 列：=SUMPRODUCT(箱规×箱数)，与 L 列的 SO 订购量对拍

# ERP 导出的列名。Odoo 导出会跟界面语言走：英文一套、德文一套（用户机器语言不定，
# 德文导出会给 Auftragsreferenz/... 这类列名）。每个字段列出「英文名, 德文名」别名，
# read_orders 按别名逐一在表头里找，找到哪套算哪套。加新语言就往元组里再补一个名。
F_ORDER = ("Order Reference", "Auftragsreferenz")
F_NAME = ("Order Lines/Product/Name", "Auftragszeilen/Produkt/Name")
F_SKU = ("Order Lines/Product/Internal Reference",
         "Auftragszeilen/Produkt/Interne Referenz")
F_BARCODE = ("Order Lines/Product/Barcode", "Auftragszeilen/Produkt/Strichcode")
F_HS = ("Order Lines/Product/HS Code", "Auftragszeilen/Produkt/HS-Code")
F_QTY = ("Order Lines/Quantity", "Auftragszeilen/Menge")


def _clean(value):
    """ERP 导出里空单元格有时是 False/空串，统一成 None。"""
    if value is None or value is False or value == "":
        return None
    return str(value).strip() or None


def read_orders(path):
    """读 ERP 导出 → [(SO号, [行])]，同 SO 内同 SKU 合并求和。

    跨 SO 的同 SKU 不合并——不同 SO 的货是分开打包的。
    """
    ws = openpyxl.load_workbook(path, data_only=True).active
    header = [_clean(c.value) for c in ws[1]]

    def col(aliases):
        """在表头里按别名(英/德…)找列位；都找不到抛 ValueError。
        用 ValueError 而非 SystemExit——后者是 BaseException，会逃出 GUI 后台线程的
        `except Exception` 导致界面卡在「运行中」(run() docstring 记的就是这个坑)。"""
        for name in aliases:
            if name in header:
                return header.index(name)
        raise ValueError(aliases[0])

    try:
        idx = {f: col(f) for f in
               (F_ORDER, F_NAME, F_SKU, F_BARCODE, F_HS, F_QTY)}
    except ValueError as e:
        raise ValueError(f"导出缺列（英文/德文列名都没找到）：{e}\n实际表头：{header}")

    orders = []
    current = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        order_no = _clean(row[idx[F_ORDER]])
        if order_no:
            current = {"order": order_no, "lines": {}}
            orders.append(current)
        sku = _clean(row[idx[F_SKU]])
        if not sku or current is None:
            continue  # 小计行 / 空行
        qty = row[idx[F_QTY]] or 0
        line = current["lines"].get(sku)
        if line:
            line["qty"] += qty
        else:
            hs = _clean(row[idx[F_HS]])
            current["lines"][sku] = {
                "sku": sku,
                "name": _clean(row[idx[F_NAME]]),
                "barcode": _clean(row[idx[F_BARCODE]]),
                "hs": hs or DEFAULT_HS,
                "hs_defaulted": not hs,
                "qty": qty,
            }
    return [o for o in orders if o["lines"]]


def _write_head(ws, made_on, so_no):
    ws["A1"] = COMPANY
    ws["A2"] = TITLE
    # 合并到 P（表宽随对拍列扩了一列），否则标题会偏左半格
    ws.merge_cells("A1:P1")
    ws.merge_cells("A2:P2")
    ws["A1"].font = Font(name="Arial", size=16, bold=True)
    ws["A2"].font = Font(name="Arial", size=28, bold=True)
    for coord in ("A1", "A2"):
        ws[coord].alignment = Alignment(horizontal="center", vertical="center",
                                        wrap_text=True)
    ws.row_dimensions[1].height = 20.25
    ws.row_dimensions[2].height = 35.25

    ws["J5"] = so_no       # Invoice 左边一格：本箱单对应的 SO 号
    ws["K5"] = "Invoice："  # 发票号人工填
    ws["M5"] = f"Date: {made_on:%d.%m.%Y}"
    for coord in ("J5", "K5", "M5"):
        ws[coord].font = Font(name="Arial", size=11, bold=True)
        ws[coord].border = BORDER
        ws[coord].number_format = "@"
    ws.row_dimensions[5].height = 15

    for col, text in enumerate(HEADERS, start=1):
        cell = ws.cell(HEADER_ROW, col, text)
        cell.font = Font(name="Arial", size=11, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[HEADER_ROW].height = 60

    for col, width in COL_WIDTHS.items():
        ws.column_dimensions[col].width = width


def _write_line(ws, row, line, spare):
    """一个 SKU 占 1 主行 + spare 行空行（供仓库拆批次/保质期）。"""
    last = row + spare
    ws.cell(row, COL_NAME, line["name"])
    ws.cell(row, COL_SKU, line["sku"])
    ws.cell(row, COL_BARCODE, line["barcode"])
    ws.cell(row, COL_HS, line["hs"])
    # L 列预填 SO 订购量（2026-08-29 用户拍板）。原先这里是 =SUMPRODUCT(箱规×箱数)、
    # 刻意不预填以免与实发不符；现改为订购量打底，实发不符靠 P 列对拍列人工发现。
    ws.cell(row, COL_QTY, line["qty"])
    # P 列：仓库填完箱规/箱数后自动出数，与左边的订购量对照
    spec = get_column_letter(COL_BOXSPEC)
    count = get_column_letter(COL_BOXCOUNT)
    ws.cell(row, COL_CHECK,
            f"=SUMPRODUCT({spec}{row}:{spec}{last},{count}{row}:{count}{last})")

    for col in (COL_NAME, COL_SKU, COL_BARCODE, COL_HS, COL_QTY, COL_CHECK):
        if spare:
            ws.merge_cells(start_row=row, start_column=col,
                           end_row=last, end_column=col)
        # ERP 主数据缺失 → 标黄，提醒回 ERP 补维护而不是在这里手补
        if col in (COL_BARCODE, COL_HS) and not ws.cell(row, col).value:
            ws.cell(row, col).fill = WARN_FILL

    for r in range(row, last + 1):
        for col in range(1, len(HEADERS) + 1):
            cell = ws.cell(r, col)
            cell.border = BORDER
            cell.font = Font(name="宋体", size=11)
            if col == COL_NAME:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center",
                                           vertical="center")
        ws.row_dimensions[r].height = 30
    return last + 1


def so_label(orders):
    """S02881+2882+2886 —— 首个 SO 写全，其余省略共同前缀。
    文件名与表头 J5 的 SO 号共用这一口径，两处必须一致。"""
    nums = [o["order"] for o in orders]
    return "+".join([nums[0]] + [re.sub(r"^S0*", "", n) for n in nums[1:]])


def build(orders, spare, made_on):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet2"
    _write_head(ws, made_on, so_label(orders))

    row = FIRST_DATA_ROW
    for order in orders:
        for line in order["lines"].values():
            row = _write_line(ws, row, line, spare)
    last = row - 1

    if last >= FIRST_DATA_ROW:
        # 原产国全表恒为 DE；托盘号/毛重/尺寸/体积重按托盘合并，
        # 但托盘是仓库现场划分的，脚本不猜 → 留空不合并
        ws.merge_cells(start_row=FIRST_DATA_ROW, start_column=COL_ORIGIN,
                       end_row=last, end_column=COL_ORIGIN)
        ws.cell(FIRST_DATA_ROW, COL_ORIGIN, ORIGIN)
        ws.cell(FIRST_DATA_ROW, COL_ORIGIN).alignment = Alignment(
            horizontal="center", vertical="center")
    return wb


def run(export, outdir="output", spare=2):
    """CLI 与 GUI 共用的入口：出表并返回 (输出路径, 摘要行)。

    摘要行是给人对 SO 单用的，CLI 打到终端、GUI 打到日志区，两边同一份。
    没有产品行时抛 ValueError（不用 SystemExit——它是 BaseException，
    在 GUI 的后台线程里不会被捕获，界面会卡在「运行中」）。
    """
    orders = read_orders(export)
    if not orders:
        raise ValueError("导出里没有产品行")

    # 箱单是打包当天产生的单据，日期取制单日（当下），与 SO 何时下的无关
    made_on = dt.date.today()
    wb = build(orders, spare, made_on)

    name = f"{so_label(orders)}箱单{made_on:%d.%m}.xlsx"

    out = Path(outdir)
    out.mkdir(parents=True, exist_ok=True)
    # 防覆盖：文件名只含 SO 号 + 制单日，同一批 SO 当天重跑必然重名。
    # 旧的那份仓库可能已经在填了，不能悄悄盖掉 → 重名自动加 (1)/(2)。
    path = Path(unique_path(str(out / name)))
    wb.save(path)
    renamed = path.name != name

    all_lines = [l for o in orders for l in o["lines"].values()]
    blanks = sum(1 for l in all_lines if not l["barcode"])
    hs_defaulted = sum(1 for l in all_lines if l["hs_defaulted"])
    # L 列已预填订购量，这里再报一次总数供人工对 SO 单
    lines = []
    for o in orders:
        qty = sum(l["qty"] for l in o["lines"].values())
        lines.append(f"  {o['order']}: {len(o['lines'])} 行 SKU / 订购 {qty:g} 件")
    lines.append(f"  共 {len(orders)} 张 SO，每行预留 {spare} 行空行")
    if renamed:
        lines.append(f"  · 同名文件 {name} 已存在，本次另存为 {path.name}（不覆盖）")
    if blanks:
        lines.append(f"  ⚠ {blanks} 行的 Bar Code 在 ERP 里为空，已标黄"
                     f"——请回 ERP 补维护产品主数据")
    if hs_defaulted:
        lines.append(f"  · {hs_defaulted} 行的 HS Code 在 ERP 里为空，已按默认"
                     f" {DEFAULT_HS} 填入")
    return path, lines


def main():
    ap = argparse.ArgumentParser(description="SO 导出 → 箱单半成品")
    ap.add_argument("export", help="ERP 导出的 sale.order xlsx")
    ap.add_argument("-o", "--outdir", default="output", help="输出目录")
    ap.add_argument("--spare", type=int, default=2,
                    help="每个 SKU 预留几行空行给仓库拆批次（默认 2）")
    args = ap.parse_args()

    try:
        path, lines = run(args.export, args.outdir, args.spare)
    except ValueError as e:
        raise SystemExit(str(e))
    print(path)
    for ln in lines:
        print(ln)


if __name__ == "__main__":
    main()
