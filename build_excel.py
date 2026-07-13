#!/usr/bin/env python3
"""在 step4_merge 分流基础上生成「拣货表+面单」workbook (输出表 A)。

用法:
    python3 build_excel.py [erp] [tmall] [out.xlsx]
默认输出: output/拣货表+面单.xlsx
"""
import sys, os, re
from datetime import date
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

import step4_merge as s4

YELLOW = PatternFill("solid", fgColor="FFFF00")
THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_BOTTOM = Alignment(horizontal="left", vertical="bottom", wrap_text=True)  # 左对齐+下沉
LEFT_CENTER = Alignment(horizontal="left", vertical="center", wrap_text=True)  # 合并单元格用
FONT = Font(size=15)
SMALL_FONT = Font(size=13)  # 比正文小 2 号
HEAD_FONT = Font(size=15, bold=True)
ROW_H = 35

# 面单列（顺序即 A..F），末尾加空白「仓库备注」
FACE_COLS = [
    (s4.ERP_ORDER_REF, "Order Reference"),
    (s4.ERP_TRACKING,  "VO Tracking No"),
    (s4.ERP_INTERNAL,  "Internal Reference"),
    (s4.ERP_PICKING,   "Picking Name"),
    (s4.ERP_QTY,       "Quantity"),
    (s4.ERP_DELIVERY,  "VO Delivery Type"),
]
WAREHOUSE_NOTE = "仓库备注"

# 操作员打印前手工调好的列宽（取自样表 2026年07月03日VO162单 拣货表+面单.xlsx），
# 按表头名固定，让程序直接输出成品宽度，打印前无需再拖列。未列出的列仍按内容自动算宽。
PICK_WIDTHS = {
    "Internal Reference": 31.0,
    "Picking Name": 22.88671875,
    "Barcode": 22.77734375,
    "Quantity": 8.21875,
    "Quantity On Hand": 7.88671875,
}
FACE_WIDTHS = {
    "序号": 12.0,
    "Order Reference": 35.0,
    "VO Tracking No": 26.0,
    "Internal Reference": 29.6640625,
    "Picking Name": 18.88671875,
    "Quantity": 8.109375,
    "VO Delivery Type": 8.43,  # 样表为 Excel 默认宽（未拖动）
    "仓库备注": 12.0,
}


def build_picking(facesheet):
    """扁平单层：一行一个 SKU。"""
    g = facesheet.groupby(s4.ERP_INTERNAL, sort=False)
    out = g.agg(**{
        "Picking Name":     (s4.ERP_PICKING,  "first"),
        "Barcode":          (s4.ERP_BARCODE,  "first"),
        "Quantity":         (s4.ERP_QTY,      "sum"),
        "Quantity On Hand": (s4.ERP_ONHAND,   "max"),
    }).reset_index().rename(columns={s4.ERP_INTERNAL: "Internal Reference"})
    # 数量统一整数，消除 csv(1.0) 与 xlsx(1) 的显示差异
    for c in ("Quantity", "Quantity On Hand"):
        out[c] = pd.to_numeric(out[c], errors="coerce").round().astype("Int64")
    # 按 Internal Reference 升序，便于捡货员按货号顺序拣货(扁平表行间独立，排序不影响数据)
    out = out.sort_values("Internal Reference", key=lambda s: s.astype(str).str.lower(),
                          kind="stable").reset_index(drop=True)
    return out


ERP_NAME = "Order Lines/Product/Name"
# 订货预判用字段(员工在 ERP 订单导出时勾上这 3 个产品级列；缺则跳过订货清单)
ERP_FS     = "Order Lines/Product/FS"            # 供应商(去谁家订)
ERP_SAFETY = "Order Lines/Product/Safety Stock"  # 安全库存
ERP_REMARK = "Order Lines/Product/Supply Remark" # 备注(可能过期，仅背景)


def build_nogoods_helper(facesheet):
    """无货勾选表：与面单**完全同版式**(同列、同顺序，套用时同合并/同标色)，仅在最前面加一列
    『无货(1=缺货)』0/1 标注(默认 0=有货，缺货填 1)。
    动机：仓库返回的纸质文档就是面单版式；屏幕表与纸面单一致，操作员对着纸面单逐行填 0/1，
    不会因两表版式不同而错位/输错。
    stage2.classify_return 按标注列(表头『无货』前缀)+ 行内 SCP 键(Order Reference 含 SCP)读取，
    与本表的具体结构解耦，故改版式不影响其判定。
    注：休眠的 build_shortage()/read_marked() 仍按旧列名(SKU/商品名/数量)取数，未来复活缺货记录
    阶段时需改读面单版列名(Internal Reference/Quantity 等)。"""
    out = build_facesheet(facesheet)  # 与面单完全同列(含序号/仓库备注)
    out.insert(0, "无货(1=缺货)", 0)  # 数字 0/1，全 Excel 版本通用；标无货填 1
    return out


def _first_nonempty(s):
    """组内第一个非空值(跳过 NaN/空串)，全空返回空串。"""
    for x in s:
        if pd.notna(x) and str(x).strip():
            return x
    return ""


# 采购画像追加列(来自 purchase order 导出，见 load_po_stats)
PO_COLS = ["供应商(次数)", "最低价", "最低价供应商", "最近一次采购", "采购总量"]

# 采购单里伪装成供应商的客户(实为我方客户，属噪音，整行剔除)
PO_CUSTOMER_PAT = "Alibaba Health"

# 供应商简称：滤掉的法律形式/地名后缀词(小写比较)
VENDOR_LEGAL = {"gmbh", "gmbh,", "kg", "kgaa", "ag", "ohg", "mbh", "mbb", "co", "co.",
                "&", "e.k.", "ek", "e.u", "ltd", "ltd.", "limited", "inc", "inc.",
                "s.a.r.l.,", "s.a.r.l.", "sarl", "sas", "bv", "se",
                "niederlassung", "deutschland", "holding"}
# 首词全大写但属行业通用词，单独指代会误导(PHARMA LUPUS ≠ "PHARMA")
VENDOR_GENERIC = {"PHARMA", "APOTHEKE", "MED"}
# 个别简称覆盖(用户指定)：规则产物 → 最终简称
VENDOR_ALIAS = {"Dirk Rossmann": "Rossmann"}


def _short_vendor(name):
    """供应商全名 → 简称(2026-07-07 全量 65 家实测零碰撞)：
    去括号注记 → 滤法律/地名后缀 → 首词全大写(≥2字符、连字符取头段、非通用词)
    则单词指代(PHOENIX/AEP/GEHE/DM/UPS)，否则取前两词；结果过短再多取一词。"""
    s = re.sub(r"[（(].*?[)）]", "", str(name)).strip()
    words = [w for w in s.split() if w.lower() not in VENDOR_LEGAL]
    if not words:
        return str(name).strip()
    head = words[0].split("-")[0]
    if head.isupper() and len(head) >= 2 and head not in VENDOR_GENERIC:
        return VENDOR_ALIAS.get(head, head)
    n = 2 if len(" ".join(words[:2])) >= 4 else 3
    res = " ".join(words[:n])
    return VENDOR_ALIAS.get(res, res)


def _vendor_map(vendors):
    """全名→简称映射；不同全名缩成同一简称(前瞻防护，当前数据零碰撞)则保留全名。"""
    m = {v: _short_vendor(v) for v in vendors}
    dup = {s for s in m.values() if list(m.values()).count(s) > 1}
    return {v: (v if s in dup else s) for v, s in m.items()}


def _po_base_sku(s):
    """SKU 归一：去掉多件装 x2 / 变体 *2 / 渠道 _VO 等尾缀，对齐采购单里的基础 SKU。"""
    return re.sub(r"(x\d+|\*\d+|_VO)+$", "", str(s).strip())


def load_po_stats(path):
    """purchase order 导出(Odoo 行式：订单头只在每单首行) → 按基础 SKU 聚合采购画像。
    最低价只统计单价>0(价 0/负数是赠品/返利)；窗口=导出里有多少算多少，不写死3月。
    返回 (stats_df[_sku + PO_COLS], 窗口描述str)。缺必需列抛 ValueError。"""
    po = pd.read_excel(path, dtype=str)
    need = ["Order Reference", "Vendor", "Order Lines/Product/Internal Reference",
            "Order Lines/Unit Price", "Order Lines/Total Quantity", "Order Lines/Created on"]
    missing = [c for c in need if c not in po.columns]
    if missing:
        raise ValueError("采购单导出缺列: " + ", ".join(missing))
    po[["Order Reference", "Vendor"]] = po[["Order Reference", "Vendor"]].ffill()
    po = po.dropna(subset=["Order Lines/Product/Internal Reference", "Vendor"]).copy()
    is_cust = po["Vendor"].str.contains(PO_CUSTOMER_PAT, case=False, na=False)
    n_cust = int(is_cust.sum())
    po = po[~is_cust].copy()
    po["Vendor"] = po["Vendor"].map(_vendor_map(po["Vendor"].unique()))
    po["_sku"] = po["Order Lines/Product/Internal Reference"].map(_po_base_sku)
    po["_price"] = pd.to_numeric(po["Order Lines/Unit Price"], errors="coerce")
    po["_qty"] = pd.to_numeric(po["Order Lines/Total Quantity"], errors="coerce")
    po["_dt"] = pd.to_datetime(po["Order Lines/Created on"], errors="coerce")
    rows = []
    for sku, g in po.groupby("_sku"):
        vc = g.groupby("Vendor")["Order Reference"].nunique().sort_values(ascending=False)
        vendors = "\n".join(f"{v}×{n}" for v, n in vc.items())  # 多家纵向排开(单元格内换行)
        priced = g[g["_price"] > 0]
        if len(priced):
            low_row = priced.loc[priced["_price"].idxmin()]
            low, low_v = float(low_row["_price"]), low_row["Vendor"]
        else:
            low, low_v = None, ""
        last = ""
        if g["_dt"].notna().any():
            lr = g.loc[g["_dt"].idxmax()]
            price_s = f" @{lr['_price']:g}" if pd.notna(lr["_price"]) else ""
            # 主次排布：供应商+价格一行，日期换行
            last = f"{lr['Vendor']}{price_s}\n{lr['_dt']:%Y-%m-%d}"
        rows.append((sku, vendors, low, low_v, last, g["_qty"].sum()))
    stats = pd.DataFrame(rows, columns=["_sku"] + PO_COLS)
    stats["采购总量"] = pd.to_numeric(stats["采购总量"], errors="coerce").round().astype("Int64")
    info = (f"{po['_dt'].min():%Y-%m-%d}~{po['_dt'].max():%Y-%m-%d} "
            f"{po['Order Reference'].nunique()} 单 / {stats.shape[0]} SKU")
    if n_cust:
        info += f" (已剔除客户{PO_CUSTOMER_PAT}记录 {n_cust} 行)"
    return stats, info


def build_reorder(erp, po_stats=None):
    """补货预判清单(Solo 作战清单·模式一 step 0 盘前预判)：
    把整份 ERP 订单导出按 Internal Reference 聚合，Quantity 求和=今日需求，与 On Hand 比。
    列名与 ERP 原始字段保持一致(短名，同拣货表)，仅算出来的列(今日需求/缺口)用中文。
    前两列 Internal Reference / Picking Name 与拣货表对齐、并按 Internal Reference 升序，
    便于对照仓库反馈的拣货单逐行勾缺货。需 ERP 导出含 FS/Safety Stock/Supply Remark 三列，缺则返回 None。
    po_stats(可选，load_po_stats 产出)：按基础 SKU 尾部追加采购画像列；无记录的 SKU
    标"无采购记录"(信号：新品/长尾需人工找供应商)。缺口只能参考，补货数量始终人为定。"""
    if not all(c in erp.columns for c in (ERP_FS, ERP_SAFETY, ERP_REMARK)):
        return None
    g = erp.groupby(s4.ERP_INTERNAL, sort=False).agg(**{
        "Picking Name":     (s4.ERP_PICKING, _first_nonempty),
        "Barcode":          (s4.ERP_BARCODE, _first_nonempty),
        "Name":             (ERP_NAME,       _first_nonempty),
        "今日需求":          (s4.ERP_QTY,     "sum"),
        "Quantity On Hand": (s4.ERP_ONHAND,  "max"),
        "Safety Stock":     (ERP_SAFETY,     "max"),
        "FS":               (ERP_FS,         _first_nonempty),
        "Supply Remark":    (ERP_REMARK,     _first_nonempty),
    }).reset_index().rename(columns={s4.ERP_INTERNAL: "Internal Reference"})
    for c in ("今日需求", "Quantity On Hand", "Safety Stock"):
        g[c] = pd.to_numeric(g[c], errors="coerce").round().astype("Int64")
    g["缺口"] = g["今日需求"] - g["Quantity On Hand"]
    # 按 Internal Reference 升序(大小写不敏感、stable)，与拣货表同序，便于逐行对照
    g = g.sort_values("Internal Reference", key=lambda s: s.astype(str).str.lower(),
                      kind="stable").reset_index(drop=True)
    cols = ["Internal Reference", "Picking Name", "Barcode", "Name", "今日需求",
            "Quantity On Hand", "缺口", "Safety Stock", "FS", "Supply Remark"]
    if po_stats is not None:
        m = (po_stats.set_index("_sku")
             .reindex(g["Internal Reference"].map(_po_base_sku)))
        for c in PO_COLS:
            g[c] = pd.Series(m[c].values, dtype=object)
        # 无记录 SKU：数字列的 NA/NaN 写不进 openpyxl，置空串
        g[PO_COLS] = g[PO_COLS].where(g[PO_COLS].notna(), "")
        g.loc[g["供应商(次数)"] == "", "供应商(次数)"] = "无采购记录"
        cols += PO_COLS
    return g[cols]


def build_facesheet(facesheet):
    df = facesheet[[c for c, _ in FACE_COLS]].copy()
    df.columns = [h for _, h in FACE_COLS]
    df["Quantity"] = pd.to_numeric(df["Quantity"], errors="coerce").round().astype("Int64")
    df.insert(0, "序号", facesheet["序号"].values)  # 与无货勾选同一编号
    df[WAREHOUSE_NOTE] = ""
    return df


def is_multipack(v):
    # x 后跟数字结尾 = 多件装（x2/x3/x4/x10…），标色提醒打包员
    return bool(re.search(r"x\d+$", str(v), re.I))


def style_sheet(ws, n_cols, header_font=HEAD_FONT, left_cols=(), small_cols=(), widths=None):
    """left_cols: 内容左对齐+下沉的列名集合；small_cols: 字号小2号的列名集合。表头始终居中。
    widths: {表头名: 列宽} 固定列宽表，命中的列用固定宽度（操作员手工调好、免二次拖列），
    未命中的列仍按内容自动算宽。"""
    headers = [c.value for c in ws[1]]
    left_idx = {i + 1 for i, h in enumerate(headers) if h in left_cols}
    small_idx = {i + 1 for i, h in enumerate(headers) if h in small_cols}
    for row in ws.iter_rows():
        for cell in row:
            cell.border = BORDER
            if cell.row == 1:
                cell.alignment = CENTER
                cell.font = header_font
            else:
                cell.alignment = LEFT_BOTTOM if cell.column in left_idx else CENTER
                cell.font = SMALL_FONT if cell.column in small_idx else FONT
    for r in range(1, ws.max_row + 1):
        # 含单元格内换行(\n)的行按行数放大，否则固定高度会裁掉第二行起的内容；
        # 无换行的行恒为 ROW_H(其他表无 \n 内容，行为不变)
        lines = max((str(c.value).count("\n") + 1 for c in ws[r] if c.value is not None),
                    default=1)
        ws.row_dimensions[r].height = ROW_H if lines == 1 else lines * 22
    # 列宽：固定表命中的列用固定宽度（操作员调好的成品宽），其余按内容自动算宽
    # （字号15 比默认大，需放大系数，否则日期显示为 ######）
    widths = widths or {}
    for c in range(1, n_cols + 1):
        hdr = headers[c - 1] if c - 1 < len(headers) else None
        if hdr in widths:
            ws.column_dimensions[get_column_letter(c)].width = widths[hdr]
            continue
        maxlen = 0
        for r in range(1, ws.max_row + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            s = v.strftime("%Y-%m-%d %H:%M:%S") if hasattr(v, "strftime") else str(v)
            maxlen = max(maxlen, len(s))
        ws.column_dimensions[get_column_letter(c)].width = max(12, maxlen * 1.5 + 2)


def write_df(ws, df):
    ws.append(list(df.columns))
    for _, row in df.iterrows():
        ws.append(list(row))


def highlight_facesheet(ws, df):
    """三条规则，标黄触发的单元格（让仓库知道原因）。
    Delivery Type 目标值按店区分：VO 的包裹标记是 CC，GW 的是 SYB。"""
    col = {name: i + 1 for i, name in enumerate(df.columns)}
    for ridx, (_, row) in enumerate(df.iterrows(), start=2):
        if is_multipack(row["Internal Reference"]):
            ws.cell(ridx, col["Internal Reference"]).fill = YELLOW
        if pd.notna(row["Quantity"]) and float(row["Quantity"]) > 1:
            ws.cell(ridx, col["Quantity"]).fill = YELLOW
        ch = str(row["Order Reference"]).split("_", 1)[0]
        if str(row["VO Delivery Type"]) == ("SYB" if ch == "GW" else "CC"):
            ws.cell(ridx, col["VO Delivery Type"]).fill = YELLOW


def merge_multiproduct(ws, df, extra_cols=()):
    """多品订单：订单级列(序号 / Order Reference / VO Tracking No / VO Delivery Type / 仓库备注)纵向合并。
    extra_cols: 额外一并按订单合并的列名(如无货勾选的 无货标注)，缺列则跳过。
    按列名定位，避免加序号列后错位。"""
    cols = list(df.columns)
    merge_cols = {n: cols.index(n) + 1
                  for n in ("序号", "Order Reference", "VO Tracking No", "VO Delivery Type",
                            WAREHOUSE_NOTE, *extra_cols) if n in cols}
    start = 2
    refs = df["Order Reference"].tolist()
    i = 0
    while i < len(refs):
        j = i
        while j + 1 < len(refs) and refs[j + 1] == refs[i]:
            j += 1
        if j > i:  # 多行同单
            for c in merge_cols.values():
                ws.merge_cells(start_row=start + i, start_column=c,
                               end_row=start + j, end_column=c)
        i = j + 1


def fix_merged_alignment(ws, left_cols):
    """合并后调用：左对齐列若为合并单元格，首格改为 左对齐+垂直居中（非下沉）。"""
    headers = [c.value for c in ws[1]]
    left_idx = {i + 1 for i, h in enumerate(headers) if h in left_cols}
    for rng in ws.merged_cells.ranges:
        if rng.min_col in left_idx:
            ws.cell(rng.min_row, rng.min_col).alignment = LEFT_CENTER


def make_output_name(facesheet, outdir):
    """步骤9 命名：YYYY年MM月DD日{渠道}{n}单 拣货表+面单.xlsx"""
    n = facesheet[s4.ERP_ORDER_REF].nunique()
    chans = sorted({str(r).split("_", 1)[0] for r in facesheet[s4.ERP_ORDER_REF].dropna()})
    ch = "+".join(chans)
    d = date.today()
    fname = f"{d.year}年{d.month:02d}月{d.day:02d}日{ch}{n}单 拣货表+面单.xlsx"
    return os.path.join(outdir, fname)


def apply_print(ws, landscape=False, fit_width=False, footer="第 &P 页，共 &N 页",
                top=0.9, bottom=0.9, left=0.8, right=0.8):
    """步骤9 打印设置。footer 用 Excel 字段码：&P=当前页码，&N=总页数。
    默认『第 &P 页，共 &N 页』(第1页，共3页...)。页边距单位为英寸。"""
    if landscape:
        ws.page_setup.orientation = "landscape"
    if fit_width:  # 所有列压到一页宽
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins = PageMargins(top=top, bottom=bottom, left=left, right=right,
                                  header=0.3, footer=0.3)
    ws.oddFooter.center.text = footer


def _load_erps(erp_paths):
    """接受单个路径或路径列表(VO/GW 各一份)，concat 成一张 ERP 行级表。"""
    if isinstance(erp_paths, str):
        erp_paths = [erp_paths]
    return pd.concat([s4.load_erp(p) for p in erp_paths], ignore_index=True)


def _order_level(df, idcol):
    """ERP 行级 → 订单级(多品合一行，取订单头行)。"""
    return df.groupby(s4.ERP_ORDER_REF, sort=False).agg(**{
        "Order Date":           ("Order Date", "first"),
        idcol:                  (idcol, "first"),
        "Terms and conditions": ("Terms and conditions", "first"),
    }).reset_index()


def _wu_add(series):
    """无运单：Terms 前缀『无运单』(已带则不重复加)。"""
    raw = series.astype(str)
    return raw.where(raw.str.contains(s4.WU_TAG), s4.WU_TAG + raw).values


def _wu_strip(series):
    """已补运单：剥掉『无运单』恢复原值。"""
    return series.astype(str).str.replace(s4.WU_TAG, "", regex=False).str.strip().values


def unique_path(path):
    """目标已存在则在扩展名前加序号 (1)/(2)…，避免覆盖既有产出。"""
    if not os.path.exists(path):
        return path
    base, ext = os.path.splitext(path)
    i = 1
    while os.path.exists(f"{base}({i}){ext}"):
        i += 1
    return f"{base}({i}){ext}"


def _write_simple(out, outdir, fname, n_cols=None, left_cols=(), small_cols=(), widths=None):
    """把一张 DataFrame 写成单 sheet workbook(统一样式)。返回 (路径, 行数)。
    left_cols/small_cols/widths 透传给 style_sheet（左对齐下沉/小字号/固定列宽），默认空=全居中自动宽。"""
    wb = Workbook(); ws = wb.active; ws.title = "Sheet1"
    write_df(ws, out)
    style_sheet(ws, n_cols or len(out.columns), left_cols=left_cols, small_cols=small_cols,
                widths=widths)
    path = unique_path(os.path.join(outdir, fname))
    wb.save(path)
    return path, len(out)


def _upload_terms(cat_df, idcol, terms):
    """某一类(取消/无运单/已补运单)的 ERP 上传行：订单级 4 列，Terms 取传入值/series。"""
    orders = _order_level(cat_df, idcol)
    out = pd.DataFrame({
        "Order Date":           orders["Order Date"].values,
        idcol:                  orders[idcol].values,
        "Order Reference":      orders[s4.ERP_ORDER_REF].values,
        "Terms and conditions": terms(orders) if callable(terms) else terms,
    })
    return out


def _write_pickface(facesheet, outdir, out_arg=None):
    """把一个店铺的发货集合写成一个「拣货表+面单+无货勾选」workbook。
    facesheet 需已含 `序号`。返回 (路径, SKU数, 订单数)。"""
    pick_df = build_picking(facesheet)
    face_df = build_facesheet(facesheet)
    wb = Workbook()
    ws_pick = wb.active; ws_pick.title = "拣货表"
    write_df(ws_pick, pick_df)
    style_sheet(ws_pick, len(pick_df.columns),
                left_cols={"Internal Reference", "Picking Name", "Barcode"},
                small_cols={"Picking Name"}, widths=PICK_WIDTHS)
    apply_print(ws_pick, fit_width=True)
    ws_face = wb.create_sheet("面单")
    write_df(ws_face, face_df)
    face_left = {"Order Reference", "VO Tracking No", "Internal Reference", "Picking Name"}
    style_sheet(ws_face, len(face_df.columns),
                left_cols=face_left,
                small_cols={"Internal Reference", "Picking Name"}, widths=FACE_WIDTHS)
    highlight_facesheet(ws_face, face_df)
    merge_multiproduct(ws_face, face_df)
    fix_merged_alignment(ws_face, face_left)
    apply_print(ws_face, landscape=True)
    chk_df = build_nogoods_helper(facesheet)
    ws_chk = wb.create_sheet("无货勾选")
    write_df(ws_chk, chk_df)
    style_sheet(ws_chk, len(chk_df.columns),
                left_cols=face_left,
                small_cols={"Internal Reference", "Picking Name"})
    highlight_facesheet(ws_chk, chk_df)
    # 无货勾选：无货标注也按订单合并(每单一格，操作员按单勾一次；序号已是固定订单级列)
    merge_multiproduct(ws_chk, chk_df, extra_cols=("无货(1=缺货)",))
    fix_merged_alignment(ws_chk, face_left)
    path = unique_path(out_arg or make_output_name(facesheet, outdir))
    wb.save(path)
    return path, len(pick_df), int(face_df["Order Reference"].nunique())


def build(erp_paths, full_tmall_path, out_arg=None, outdir="output", po_path=None):
    """阶段一核心(步骤4+7/8/9)：分流 + 生成交付。返回 (log行列表, stats)。

    输入单店 ERP（天猫两店混合，经 ∩ERP 收敛到单店）+ 一份完整天猫导出，产出全部按店带后缀：
    - 新订单获单清单：履约单状态=新订单 ∩ ERP 的系统履约单号(去天猫批量获单)。
    - 拣货表+面单：只含「发货」订单(已剔除无运单/取消)。
    - 回传ERP销售上传表：取消/无运单/已补运单三类 Terms 写回**合并一张**(External ID 匹配键)。
    完整天猫导出是唯一天猫输入：发货范围由二段式(履约∈{新订单,商家已接单}∧面单已完成)推出，
    负集再按履约状态拆 取消/无运单。供 CLI(main) 与 GUI 共用。"""
    os.makedirs(outdir, exist_ok=True)
    log = []
    erp = _load_erps(erp_paths)
    full = s4.load_full_tmall(full_tmall_path)         # 唯一天猫输入
    done = s4.done_keys_from_full(full)                # 发货范围(二段式)
    status_map = full[s4.TM_STATUS] if not full.empty else pd.Series(dtype=object)
    cancel_keys = s4.cancel_keys_from_full(full)
    ann = s4.classify4(erp, done, cancel_keys)
    idcol = s4.find_id_col(ann)

    o = ann.drop_duplicates("_key")
    log.append("分流(订单级): " + " / ".join(
        f"{k} {v}" for k, v in o["_cat"].value_counts().items()))

    # ---- 拣货表 + 面单 (发货集合，按店铺 VO/GW 各出一份) ----
    facesheet = ann[ann["_ship"]].copy()
    main_paths = []
    if facesheet.empty:
        log.append("⚠ 无发货订单(全部无运单/取消)，未生成拣货表+面单")
    else:
        facesheet["_ch"] = (facesheet[s4.ERP_ORDER_REF].astype(str)
                            .str.split("_", n=1).str[0])
        chans = sorted(facesheet["_ch"].unique())
        for ch in chans:
            sub = (facesheet[facesheet["_ch"] == ch]
                   .drop(columns="_ch").reset_index(drop=True))
            sub.insert(0, "序号", range(1, len(sub) + 1))   # 序号按店内独立编号
            p, nsku, nord = _write_pickface(
                sub, outdir, out_arg if len(chans) == 1 else None)
            main_paths.append(p)
            log.append(f"拣货表+面单[{ch}] 已生成: {p}  ({nsku} SKU / {nord} 单)")

    # ---- 新订单获单清单 (履约单状态=新订单 ∩ ERP；复制履约单号去天猫批量获单；按店各一份) ----
    d = date.today()
    if len(status_map):
        new_keys = set(status_map[status_map == s4.NEW_ORDER_STATUS].index)
        no = ann.drop_duplicates("_key")
        no = no[no["_key"].isin(new_keys)].copy()
        if not no.empty:
            no["_ch"] = no[s4.ERP_ORDER_REF].astype(str).str.split("_", n=1).str[0]
            for ch in sorted(no["_ch"].unique()):
                keys = no[no["_ch"] == ch]["_key"].tolist()
                p, n = _write_simple(pd.DataFrame({"系统履约单号": keys}),
                                     outdir, f"新订单获单清单{ch}.xlsx", n_cols=1)
                log.append(f"新订单获单清单{ch} 已生成: {p}  ({n} 单)")
        else:
            log.append("新订单获单清单: 0 单")
    else:
        log.append("新订单获单清单: 跳过 (未传完整天猫导出，无法识别新订单)")

    # ---- 回传ERP销售上传表 (取消/无运单/已补运单 三类 Terms 写回一张，按店各一份) ----
    tag = f"{d.year}年{d.month:02d}月{d.day:02d}日平台订单取消"
    cats = {
        "取消":     (ann[ann["_cat"] == "取消"],     tag),
        "无运单":    (ann[ann["_cat"] == "无运单"],    lambda o: _wu_add(o["Terms and conditions"])),
        "已补运单":  (ann[ann["_cat"] == "已补运单"],  lambda o: _wu_strip(o["Terms and conditions"])),
    }
    present = {k: df for k, (df, _) in cats.items() if not df.empty}
    if present and idcol:
        parts = [_upload_terms(df, idcol, terms)
                 for k, (df, terms) in cats.items() if not df.empty]
        allup = pd.concat(parts, ignore_index=True)
        allup["_ch"] = allup["Order Reference"].astype(str).str.split("_", n=1).str[0]
        for ch in sorted(allup["_ch"].unique()):
            sub = allup[allup["_ch"] == ch].drop(columns="_ch")
            p, n = _write_simple(sub, outdir, f"回传ERP销售上传表{ch}.xlsx")
            log.append(f"回传ERP销售上传表{ch} 已生成: {p}  ({n} 单)")
        log.append("  └ 含 " + " / ".join(
            f"{k} {df['_key'].nunique()}" for k, df in present.items()))
    elif present:
        log.append("⚠ 有需回传订单(取消/无运单/已补运单)但 ERP 无 External ID 列，"
                   "无法生成上传表(请在订单导出勾上 External ID)")
    else:
        log.append("回传ERP销售上传表: 0 单 (无取消/无运单/已补运单)"
                   + ("" if full_tmall_path else " (未传完整天猫导出，取消无法识别)"))

    # ---- 取消订单清单 (种子表：供阶段二生成『取消出库单』批量清理 dangling picking) ----
    # 取消是滚动产生的(回传天猫后仍会冒出 1~5 单)，此表只是阶段一时的初始集；
    # 用户在回传天猫后把后到的取消单(填系统履约号 SCP)手工 append 进去，再喂给阶段二。
    cxl = ann[ann["_cat"] == "取消"].drop_duplicates("_key")
    if not cxl.empty:
        seed = pd.DataFrame({"系统履约单号": cxl["_key"].tolist(),
                             "Order Reference": cxl[s4.ERP_ORDER_REF].tolist()})
        p, n = _write_simple(seed, outdir, "取消订单清单.xlsx",
                             left_cols={"系统履约单号", "Order Reference"})
        log.append(f"取消订单清单 已生成: {p}  ({n} 单；后到的取消单请手工补录后喂阶段二)")

    # ---- 补货预判清单 (Solo 作战清单·模式一 step 0；需 ERP 含 FS/Safety/Remark) ----
    # 只在传了采购单导出时才产出：不带真实采购参考的补货预判清单信息滞后、无参考意义，
    # 未传采购单导出则整张跳过，仅出其余 3 份(拣货表+面单/新订单获单/回传ERP上传表)。
    if not po_path:
        log.append("补货预判清单: 跳过 (未传采购单导出)")
    else:
        po_stats = None
        try:
            po_stats, po_info = load_po_stats(po_path)
            log.append(f"采购参考已加载: {po_info}")
        except Exception as e:
            log.append(f"⚠ 采购单导出读取失败，补货预判清单不带采购参考: {e}")
        reorder = build_reorder(erp, po_stats)
        if reorder is None:
            log.append("补货预判清单: 跳过 (ERP 订单导出未含 FS/Safety Stock/Supply Remark 列；"
                       "在 Odoo 订单导出模板勾上这 3 列即可生成)")
        elif reorder.empty:
            log.append("补货预判清单: 0 SKU")
        else:
            reorder["_ch"] = (erp.drop_duplicates(s4.ERP_INTERNAL)
                              .set_index(s4.ERP_INTERNAL)[s4.ERP_ORDER_REF]
                              .reindex(reorder["Internal Reference"]).astype(str)
                              .str.split("_", n=1).str[0].values)
            for ch in sorted(reorder["_ch"].dropna().unique()):
                sub = reorder[reorder["_ch"] == ch].drop(columns="_ch")
                short = int((sub["缺口"] > 0).sum())
                # 非数字列左对齐+下沉；数字列(今日需求/On Hand/缺口/Safety Stock/最低价/采购总量)保持居中
                # 采购画像的长文本列固定宽度让 wrap 生效，否则供应商名单会把列撑得极宽
                p, n = _write_simple(sub, outdir, f"{ch}补货预判清单.xlsx",
                                     left_cols={"Internal Reference", "Picking Name",
                                                "Barcode", "Name", "FS", "Supply Remark",
                                                "供应商(次数)", "最低价供应商", "最近一次采购"},
                                     widths={"供应商(次数)": 45, "最低价供应商": 22,
                                             "最近一次采购": 30})
                log.append(f"{ch}补货预判清单 已生成: {p}  ({n} SKU，其中缺口>0 {short} 个)")

    # ---- 异常上报(不静默) ----
    dup = erp.drop_duplicates(s4.ERP_ORDER_REF)["_key"].duplicated().sum()
    if dup:
        log.append(f"⚠ 连接键冲突: {dup} 个订单的 Order Reference 后15位与他单相同(可能误判状态)")
    # 注：发货范围由二段式从活单状态(新订单/商家已接单)推出，已结构性排除已发货/已收货/
    # 取消的历史单，故无需再做「名单过期」反查护栏。
    if not facesheet.empty:
        empty_dt = facesheet.drop_duplicates("_key")[s4.ERP_DELIVERY].isna().sum()
        if empty_dt:
            log.append(f"⚠ 发货订单中 VO Delivery Type 为空: {empty_dt} 单")

    stats = {
        "ship": int(o["_ship"].sum()),
        "cancel": int((o["_cat"] == "取消").sum()),
        "nowaybill": int((o["_cat"] == "无运单").sum()),
        "refill": int((o["_cat"] == "已补运单").sum()),
        "main_paths": main_paths,
    }
    return log, stats


def main():
    args = sys.argv[1:]
    if len(args) < 2:
        print("用法: python3 build_excel.py <erp[,erp2...]> <完整天猫导出> [out.xlsx]")
        return
    erp_paths = args[0].split(",")
    full_tmall = args[1]
    out_arg = args[2] if len(args) > 2 else None
    outdir = os.path.dirname(out_arg) if out_arg else "output"
    log, st = build(erp_paths, full_tmall, out_arg, outdir)
    for ln in log:
        print(ln)


if __name__ == "__main__":
    main()
