#!/usr/bin/env python3
"""京东订单 · 通用选列导出。

京东后台导出的原始 xlsx(如「复核历史查询-SKU汇总」)列很多，最终要哪些列、
出成几版仍在摸索。本模块把"出哪些列"交给操作时勾选/排序，并支持存成命名预设，
日常一键出表。逻辑与 VO/GW 那套固定管线完全独立。

核心约定：
- 长数字列(订单号16位/运单号/生产单号 等)一律以文本写出，防 Excel 科学计数法丢精度。
- 预设列按"列名"记；某列在原始数据里缺失(京东改名/去列) → 跳过并告警，不报错崩。
- aggregate 字段当前保留不实现(留档版是否按商品编号聚合出货件数未定)；置值仅告警。
"""
import os
import json
from datetime import date

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---- 内置预设(始终可用；用户另存的预设叠加在其上) ----
BUILTIN_PRESETS = [
    {"name": "货代版", "columns": ["订单号", "运单号"],
     "dedup": True, "aggregate": None, "out_name": "京东货代"},
    {"name": "内部留档版",
     "columns": ["运单号", "订单号", "商品编号", "商品名称",
                 "复核数量", "生产日期", "到期日期"],
     "dedup": False, "aggregate": None, "out_name": "京东留档"},
]

HEAD_FONT = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_thin = Side(style="thin", color="D0D0D0")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


# ---------- 预设管理 ----------
def presets_path(base_dir):
    return os.path.join(base_dir, "presets", "京东预设.json")


def load_presets(base_dir):
    """返回预设列表：内置在前，用户另存的按名覆盖/追加。"""
    presets = [dict(p) for p in BUILTIN_PRESETS]
    idx = {p["name"]: i for i, p in enumerate(presets)}
    path = presets_path(base_dir)
    if os.path.exists(path):
        try:
            with open(path, encoding="utf-8") as f:
                for p in json.load(f).get("presets", []):
                    if p.get("name") in idx:
                        presets[idx[p["name"]]] = p        # 同名覆盖内置
                    else:
                        idx[p["name"]] = len(presets)
                        presets.append(p)
        except (json.JSONDecodeError, OSError):
            pass                                            # 坏文件不阻断，用内置
    return presets


def save_preset(base_dir, preset):
    """把一个预设写入用户预设文件(按名替换或追加)。仅存用户的，不写内置。"""
    path = presets_path(base_dir)
    os.makedirs(os.path.dirname(path), exist_ok=True)
    data = {"presets": []}
    if os.path.exists(path):
        try:
            with open(path, encoding="utf-8") as f:
                data = json.load(f)
        except (json.JSONDecodeError, OSError):
            data = {"presets": []}
    users = [p for p in data.get("presets", []) if p.get("name") != preset["name"]]
    users.append(preset)
    with open(path, "w", encoding="utf-8") as f:
        json.dump({"presets": users}, f, ensure_ascii=False, indent=2)


# ---------- 读列 / 导出 ----------
def read_columns(path):
    """读原始 xlsx 第一个 sheet 的列名(丢掉 pandas 对空表头生成的 Unnamed 占位列)。"""
    df = pd.read_excel(path, sheet_name=0, dtype=str, nrows=0)
    return [c for c in df.columns if not str(c).startswith("Unnamed:")]


def _out_name(name, n, d=None):
    """YYYY年MM月DD日{n}单 {name}.xlsx，对齐阶段一/二命名(京东无渠道，故不带渠道段)。"""
    d = d or date.today()
    return f"{d.year}年{d.month:02d}月{d.day:02d}日{n}单 {name}.xlsx"


def _unique_path(path):
    if not os.path.exists(path):
        return path
    base, ext = os.path.splitext(path)
    i = 1
    while os.path.exists(f"{base}({i}){ext}"):
        i += 1
    return f"{base}({i}){ext}"


def export(path, columns, outdir, out_name="京东导出",
           dedup=False, aggregate=None, d=None):
    """按 columns(顺序即输出列序)从原始 xlsx 产出一张 xlsx。

    返回 (输出路径, 行数, warnings)。缺列跳过并进 warnings；aggregate 保留未实现。
    """
    warnings = []
    df = pd.read_excel(path, sheet_name=0, dtype=str).fillna("")

    present = [c for c in columns if c in df.columns]
    missing = [c for c in columns if c not in df.columns]
    for c in missing:
        warnings.append(f"⚠ 预设列「{c}」在原始数据里找不到，已跳过(京东是否改了列名?)")
    if not present:
        raise ValueError("选中的列在原始数据里一个都没匹配上，无法导出")

    out = df[present].copy()
    if dedup:
        before = len(out)
        out = out.drop_duplicates().reset_index(drop=True)
        if len(out) < before:
            warnings.append(f"去重：{before} 行 → {len(out)} 行")
    if aggregate:
        warnings.append("⚠ 该预设设了 aggregate(聚合)，当前版本尚未实现，已按原样输出未聚合")

    os.makedirs(outdir, exist_ok=True)
    out_path = _unique_path(os.path.join(outdir, _out_name(out_name, len(out), d)))
    _write(out, out_path)
    return out_path, len(out), warnings


def _write(df, out_path):
    """单 sheet 写出；所有单元格锁文本格式('@')，长数字不被转科学计数法。"""
    wb = Workbook(); ws = wb.active; ws.title = "Sheet1"
    ws.append(list(df.columns))
    for _, row in df.iterrows():
        ws.append(["" if v is None else str(v) for v in row])
    for r_i, row in enumerate(ws.iter_rows(), start=1):
        for cell in row:
            cell.border = BORDER
            if r_i == 1:
                cell.font = HEAD_FONT
                cell.alignment = CENTER
            else:
                cell.number_format = "@"          # 文本，防精度丢失
                cell.alignment = LEFT
    for c in range(1, len(df.columns) + 1):
        maxlen = max([len(str(df.columns[c - 1]))] +
                     [len(str(v)) for v in df.iloc[:, c - 1]] + [4])
        ws.column_dimensions[get_column_letter(c)].width = min(60, maxlen * 1.2 + 2)
    wb.save(out_path)


# ---------- 便于命令行自测 ----------
if __name__ == "__main__":
    import argparse
    ap = argparse.ArgumentParser(description="京东选列导出(自测用)")
    ap.add_argument("raw", help="原始 xlsx 路径")
    ap.add_argument("--preset", default="货代版", help="预设名(默认 货代版)")
    ap.add_argument("--out", default=".", help="输出目录")
    a = ap.parse_args()
    base = os.path.dirname(os.path.abspath(__file__))
    ps = {p["name"]: p for p in load_presets(base)}
    if a.preset not in ps:
        raise SystemExit(f"没有预设 {a.preset}；可用：{list(ps)}")
    p = ps[a.preset]
    op, n, w = export(a.raw, p["columns"], a.out, p.get("out_name", a.preset),
                      dedup=p.get("dedup", False), aggregate=p.get("aggregate"))
    print(f"已生成：{op}  ({n} 行)")
    for line in w:
        print(line)
